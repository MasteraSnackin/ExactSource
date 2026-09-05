#!/usr/bin/env python3
"""Freeze a deterministic, metadata-stratified public development split.

This tool deliberately inspects initial workbooks only. It never searches for or
opens golden workbooks, so choosing the hold-out cannot be influenced by answers.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries

FAMILY_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("delete", ("delete", "remove", "clear", "excluding", "except")),
    ("filter", ("filter", "extract", "criteria", "matching rows", "select")),
    ("merge", ("merge", "combine", "consolidat", "append", "multiple sheets")),
    ("dedupe", ("duplicate", "deduplicat", "unique", "distinct")),
    ("sort", ("sort", "rank", "order by", "ascending", "descending")),
    ("lookup", ("lookup", "index", "match", "xlookup", "reference")),
    ("conditional", ("if ", "condition", "iferror", "ifna", "case")),
    ("aggregate", ("sum", "count", "average", "total", "aggregate")),
    ("text", ("concaten", "substring", "text", "character", "word")),
    ("date", ("date", "month", "year", "weekday", "time")),
    ("reshape", ("transpose", "pivot", "reshape", "rows into columns")),
)


@dataclass(frozen=True, slots=True)
class Features:
    task_id: str
    kind: str
    family: str
    answer_size_bucket: str
    has_formula: bool
    multi_sheet: bool
    macro_wording: bool
    exceeds_baseline_window: bool
    has_table: bool
    has_defined_name: bool

    @property
    def stratum(self) -> tuple[str, str, str]:
        return self.kind, self.family, self.answer_size_bucket


def stable_rank(seed: str, task_id: str) -> str:
    return hashlib.sha256(f"{seed}\0{task_id}".encode()).hexdigest()


def classify_family(instruction: str) -> str:
    folded = f" {instruction.casefold()} "
    for family, keywords in FAMILY_KEYWORDS:
        if any(keyword in folded for keyword in keywords):
            return family
    return "other"


def size_bucket(size: int) -> str:
    if size <= 10:
        return "01-10"
    if size <= 50:
        return "11-50"
    if size <= 250:
        return "051-250"
    if size <= 1_000:
        return "0251-1000"
    if size <= 10_000:
        return "1001-10000"
    return "10001+"


def repaired_range(value: str) -> str:
    if ":" not in value:
        return value
    start, end = value.split(":", 1)
    if end.isdigit():
        column = "".join(character for character in start if character.isalpha())
        return f"{start}:{column}{end}"
    return value


def parsed_answer_ranges(task: dict) -> list[tuple[str | None, str]]:
    raw = str(task["answer_position"])
    cleaned = raw.replace("'", "").replace('"', "")
    tokens = [cleaned] if cleaned.count("!") == 1 else cleaned.split(",")
    parsed: list[tuple[str | None, str]] = []
    for token in tokens:
        token = token.strip()
        if "!" in token:
            sheet, cells = token.rsplit("!", 1)
            parsed.append((sheet, repaired_range(cells)))
        else:
            parsed.append((task.get("answer_sheet"), repaired_range(token)))
    return parsed


def range_size(cell_range: str, max_row: int) -> int:
    min_col, min_row, max_col, final_row = range_boundaries(cell_range)
    first_row = min_row or 1
    final_row = final_row or max_row or first_row
    return (max_col - min_col + 1) * (final_row - first_row + 1)


def resolve_initial_workbook(dataset_dir: Path, task: dict) -> Path:
    folder = (dataset_dir / str(task["spreadsheet_path"])).resolve()
    try:
        folder.relative_to(dataset_dir.resolve())
    except ValueError as exc:
        raise ValueError(f"task {task.get('id')} path escapes dataset") from exc
    matches = sorted(folder.glob("*init*.xlsx"))
    if len(matches) != 1:
        raise ValueError(f"task {task.get('id')} has {len(matches)} initial workbooks")
    return matches[0]


def inspect_task(dataset_dir: Path, task: dict) -> Features:
    task_id = str(task["id"])
    workbook_path = resolve_initial_workbook(dataset_dir, task)
    workbook = openpyxl.load_workbook(
        workbook_path, data_only=False, read_only=False, keep_links=False
    )
    try:
        has_formula = False
        exceeds_window = False
        has_table = False
        for sheet in workbook.worksheets:
            exceeds_window |= sheet.max_row > 120 or sheet.max_column > 30
            has_table |= bool(sheet.tables)
            if not has_formula:
                has_formula = any(
                    isinstance(cell.value, str) and cell.value.startswith("=")
                    for row in sheet.iter_rows()
                    for cell in row
                )

        answer_size = 0
        for sheet_name, cells in parsed_answer_ranges(task):
            if sheet_name and sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            answer_size += range_size(cells, sheet.max_row)

        instruction = str(task.get("instruction", ""))
        instruction_type = str(task.get("instruction_type", ""))
        kind = "cell" if instruction_type.casefold().startswith("cell") else "sheet"
        return Features(
            task_id=task_id,
            kind=kind,
            family=classify_family(instruction),
            answer_size_bucket=size_bucket(answer_size),
            has_formula=has_formula,
            multi_sheet=len(workbook.sheetnames) > 1,
            macro_wording=any(marker in instruction.casefold() for marker in ("vba", "macro")),
            exceeds_baseline_window=exceeds_window,
            has_table=has_table,
            has_defined_name=any(True for _ in workbook.defined_names.values()),
        )
    finally:
        workbook.close()


def apportion_holdout(
    features: Iterable[Features], *, count: int, seed: str
) -> tuple[list[str], list[str]]:
    rows = list(features)
    if not 0 < count < len(rows):
        raise ValueError("hold-out count must be between zero and dataset size")
    by_kind: dict[str, list[Features]] = defaultdict(list)
    for row in rows:
        by_kind[row.kind].append(row)
    kind_quotas = _hamilton_quotas(
        {kind: len(group) for kind, group in by_kind.items()},
        total=count,
        seed=seed,
    )

    groups: dict[tuple[str, str, str], list[Features]] = defaultdict(list)
    for row in rows:
        groups[row.stratum].append(row)
    quotas: dict[tuple[str, str, str], int] = {}
    for kind, _kind_rows in by_kind.items():
        sizes = {stratum: len(group) for stratum, group in groups.items() if stratum[0] == kind}
        quotas.update(
            _hamilton_quotas(
                sizes,
                total=kind_quotas[kind],
                seed=f"{seed}|{kind}",
            )
        )

    holdout: list[str] = []
    for stratum, group in groups.items():
        ranked = sorted(group, key=lambda row: stable_rank(seed, row.task_id))
        holdout.extend(row.task_id for row in ranked[: quotas[stratum]])
    holdout_set = set(holdout)
    dev = [row.task_id for row in rows if row.task_id not in holdout_set]
    return sorted(dev), sorted(holdout)


def _hamilton_quotas(sizes: dict[object, int], *, total: int, seed: str) -> dict[object, int]:
    """Allocate ``total`` slots proportionally using largest remainders."""

    population = sum(sizes.values())
    if total < 0 or total > population:
        raise ValueError("quota total is outside the population")
    if population == 0:
        return {}
    exact = {key: size * total / population for key, size in sizes.items()}
    quotas = {key: math.floor(value) for key, value in exact.items()}
    remaining = total - sum(quotas.values())
    order = sorted(
        sizes,
        key=lambda key: (
            -(exact[key] - quotas[key]),
            stable_rank(seed, repr(key)),
        ),
    )
    for key in order[:remaining]:
        quotas[key] += 1
    return quotas


def counts(rows: Iterable[Features], selected: set[str]) -> dict[str, dict[str, int]]:
    chosen = [row for row in rows if row.task_id in selected]
    return {
        "kind": dict(sorted(Counter(row.kind for row in chosen).items())),
        "family": dict(sorted(Counter(row.family for row in chosen).items())),
        "answer_size_bucket": dict(
            sorted(Counter(row.answer_size_bucket for row in chosen).items())
        ),
        "boolean_features": {
            field: sum(bool(getattr(row, field)) for row in chosen)
            for field in (
                "has_formula",
                "multi_sheet",
                "macro_wording",
                "exceeds_baseline_window",
                "has_table",
                "has_defined_name",
            )
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset-dir", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--holdout-count", type=int, default=80)
    parser.add_argument("--seed", default="exactsource-public-v1")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    dataset_path = args.dataset_dir / "dataset.json"
    raw = dataset_path.read_bytes()
    dataset = json.loads(raw)
    if not isinstance(dataset, list):
        raise ValueError("dataset.json must contain a list")
    features = [inspect_task(args.dataset_dir.resolve(), task) for task in dataset]
    if len({row.task_id for row in features}) != len(features):
        raise ValueError("dataset task ids are not unique")
    dev, holdout = apportion_holdout(features, count=args.holdout_count, seed=args.seed)
    document = {
        "schema_version": 1,
        "seed": args.seed,
        "dataset_sha256": hashlib.sha256(raw).hexdigest(),
        "development_ids": dev,
        "holdout_ids": holdout,
        "summary": {
            "all": counts(features, {row.task_id for row in features}),
            "development": counts(features, set(dev)),
            "holdout": counts(features, set(holdout)),
        },
        "features": [asdict(row) for row in features],
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(document, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"wrote {len(dev)} development and {len(holdout)} hold-out ids to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
