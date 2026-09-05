#!/usr/bin/env python3
"""Check ExactSource's modern formula storage names with LibreOffice.

The workbooks are synthetic and contain no SpreadsheetBench data. Each case is
written to a temporary xlsx file, recalculated by the requested ``soffice``
executable, and reopened with cached-value reading enabled.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import shutil
import subprocess
import sys
import tempfile
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "Compatibility"
RESULT_CELL = "F2"
SCHEMA_VERSION = 1
SOFFICE_TIMEOUT_SECONDS = 120


class CompatibilityError(RuntimeError):
    """Raised when the compatibility check cannot be completed."""


JsonScalar = str | int | float | bool | None


@dataclass(frozen=True)
class FormulaCase:
    """One synthetic formula and its independently specified expected value."""

    case_id: str
    function: str
    stored_name: str
    formula: str
    expected: JsonScalar
    inputs: tuple[tuple[str, JsonScalar], ...] = ()


FORMULA_CASES: tuple[FormulaCase, ...] = (
    FormulaCase(
        "aggregate",
        "AGGREGATE",
        "_xlfn.AGGREGATE",
        "=_xlfn.AGGREGATE(9,4,A2:A4)",
        6,
        (("A2", 1), ("A3", 2), ("A4", 3)),
    ),
    FormulaCase(
        "choosecols",
        "CHOOSECOLS",
        "_xlfn.CHOOSECOLS",
        "=INDEX(_xlfn.CHOOSECOLS(A2:C2,2),1,1)",
        20,
        (("A2", 10), ("B2", 20), ("C2", 30)),
    ),
    FormulaCase(
        "concat",
        "CONCAT",
        "_xlfn.CONCAT",
        "=_xlfn.CONCAT(A2:C2)",
        "redblue",
        (("A2", "red"), ("B2", "blue"), ("C2", "")),
    ),
    FormulaCase(
        "drop",
        "DROP",
        "_xlfn.DROP",
        "=INDEX(_xlfn.DROP(A2:A4,1),1)",
        20,
        (("A2", 10), ("A3", 20), ("A4", 30)),
    ),
    FormulaCase(
        "filter",
        "FILTER",
        "_xlfn._xlws.FILTER",
        '=INDEX(_xlfn._xlws.FILTER(A2:A4,B2:B4="keep"),1)',
        "one",
        (
            ("A2", "one"),
            ("A3", "two"),
            ("A4", "three"),
            ("B2", "keep"),
            ("B3", "skip"),
            ("B4", "keep"),
        ),
    ),
    FormulaCase(
        "filterxml",
        "FILTERXML",
        "_xlfn.FILTERXML",
        '=_xlfn.FILTERXML("<root><v>alpha</v></root>","//v")',
        "alpha",
    ),
    FormulaCase(
        "ifna",
        "IFNA",
        "_xlfn.IFNA",
        "=_xlfn.IFNA(NA(),7)",
        7,
    ),
    FormulaCase(
        "let",
        "LET",
        "_xlfn.LET",
        "=_xlfn.LET(x,A2*2,x+1)",
        7,
        (("A2", 3),),
    ),
    FormulaCase(
        "maxifs",
        "MAXIFS",
        "_xlfn.MAXIFS",
        '=_xlfn.MAXIFS(A2:A4,B2:B4,"yes")',
        9,
        (
            ("A2", 3),
            ("A3", 9),
            ("A4", 5),
            ("B2", "yes"),
            ("B3", "yes"),
            ("B4", "no"),
        ),
    ),
    FormulaCase(
        "minifs",
        "MINIFS",
        "_xlfn.MINIFS",
        '=_xlfn.MINIFS(A2:A4,B2:B4,"yes")',
        3,
        (
            ("A2", 3),
            ("A3", 9),
            ("A4", 5),
            ("B2", "yes"),
            ("B3", "yes"),
            ("B4", "no"),
        ),
    ),
    FormulaCase(
        "sort",
        "SORT",
        "_xlfn._xlws.SORT",
        "=INDEX(_xlfn._xlws.SORT(A2:A4),1)",
        1,
        (("A2", 3), ("A3", 1), ("A4", 2)),
    ),
    FormulaCase(
        "sortby",
        "SORTBY",
        "_xlfn.SORTBY",
        "=INDEX(_xlfn.SORTBY(A2:A4,B2:B4,1),1)",
        "second",
        (
            ("A2", "first"),
            ("A3", "second"),
            ("A4", "third"),
            ("B2", 20),
            ("B3", 10),
            ("B4", 30),
        ),
    ),
    FormulaCase(
        "textjoin",
        "TEXTJOIN",
        "_xlfn.TEXTJOIN",
        '=_xlfn.TEXTJOIN("-",TRUE,A2:C2)',
        "red-blue",
        (("A2", "red"), ("B2", ""), ("C2", "blue")),
    ),
    FormulaCase(
        "textsplit",
        "TEXTSPLIT",
        "_xlfn.TEXTSPLIT",
        '=INDEX(_xlfn.TEXTSPLIT(A2,","),1,2)',
        "blue",
        (("A2", "red,blue,green"),),
    ),
    FormulaCase(
        "unique",
        "UNIQUE",
        "_xlfn.UNIQUE",
        "=INDEX(_xlfn.UNIQUE(A2:A5),2)",
        "beta",
        (("A2", "alpha"), ("A3", "beta"), ("A4", "alpha"), ("A5", "gamma")),
    ),
    FormulaCase(
        "xlookup",
        "XLOOKUP",
        "_xlfn.XLOOKUP",
        '=_xlfn.XLOOKUP("b",A2:A4,B2:B4)',
        20,
        (
            ("A2", "a"),
            ("A3", "b"),
            ("A4", "c"),
            ("B2", 10),
            ("B3", 20),
            ("B4", 30),
        ),
    ),
)


Recalculator = Callable[[Sequence[Path], Path, Path, Path], None]


def _resolve_soffice(value: str | os.PathLike[str]) -> Path:
    raw = os.fspath(value)
    if not raw.strip():
        raise CompatibilityError("--soffice must name an executable")
    contains_separator = os.sep in raw or (os.altsep is not None and os.altsep in raw)
    resolved = Path(raw).expanduser().resolve() if contains_separator else None
    if resolved is None:
        found = shutil.which(raw)
        if found is None:
            raise CompatibilityError(f"cannot find soffice executable: {raw}")
        resolved = Path(found).resolve()
    if not resolved.is_file() or not os.access(resolved, os.X_OK):
        raise CompatibilityError(f"soffice path is not an executable file: {resolved}")
    return resolved


def _soffice_version(soffice: Path) -> str:
    try:
        completed = subprocess.run(
            [str(soffice), "--version"],
            capture_output=True,
            text=True,
            check=False,
            timeout=30,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise CompatibilityError(f"cannot read LibreOffice version: {exc}") from exc
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout).strip()
        raise CompatibilityError(
            f"LibreOffice version command exited {completed.returncode}: {detail}"
        )
    version = next((line.strip() for line in completed.stdout.splitlines() if line.strip()), "")
    if not version:
        raise CompatibilityError("LibreOffice version command returned no version text")
    return version


def _create_case_workbook(case: FormulaCase, path: Path) -> None:
    workbook = openpyxl.Workbook()
    try:
        sheet = workbook.active
        sheet.title = SHEET_NAME
        for coordinate, value in case.inputs:
            sheet[coordinate] = value
        sheet[RESULT_CELL] = case.formula
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(path)
    finally:
        workbook.close()


def _recalculate_with_soffice(
    sources: Sequence[Path], output_dir: Path, profile_dir: Path, soffice: Path
) -> None:
    command = [
        str(soffice),
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        *(str(path) for path in sources),
    ]
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False,
            timeout=SOFFICE_TIMEOUT_SECONDS,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise CompatibilityError(f"LibreOffice recalculation failed: {exc}") from exc
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout).strip()
        raise CompatibilityError(
            f"LibreOffice recalculation exited {completed.returncode}: {detail}"
        )
    missing = [path.name for path in sources if not (output_dir / path.name).is_file()]
    if missing:
        raise CompatibilityError(
            "LibreOffice did not create recalculated files: " + ", ".join(missing)
        )


def _read_cached_result(path: Path) -> JsonScalar:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CompatibilityError(f"cannot open recalculated workbook {path.name}: {exc}") from exc
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise CompatibilityError(
                f"recalculated workbook {path.name} has no {SHEET_NAME!r} sheet"
            )
        value = workbook[SHEET_NAME][RESULT_CELL].value
    finally:
        workbook.close()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    raise CompatibilityError(
        f"cached result in {path.name}!{RESULT_CELL} is not JSON-compatible: {type(value).__name__}"
    )


def _values_equal(actual: JsonScalar, expected: JsonScalar) -> bool:
    if isinstance(actual, bool) or isinstance(expected, bool):
        return type(actual) is type(expected) and actual == expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        if not math.isfinite(float(actual)) or not math.isfinite(float(expected)):
            return False
        return math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-12)
    return type(actual) is type(expected) and actual == expected


def run_compatibility_check(
    soffice: Path,
    *,
    recalculator: Recalculator = _recalculate_with_soffice,
    engine_version: str | None = None,
) -> dict[str, Any]:
    """Build, recalculate, and evaluate all synthetic formula cases."""

    version = engine_version if engine_version is not None else _soffice_version(soffice)
    with tempfile.TemporaryDirectory(prefix="exactsource-formula-compat-") as raw_temp:
        temp_root = Path(raw_temp)
        source_dir = temp_root / "source"
        output_dir = temp_root / "recalculated"
        profile_dir = temp_root / "libreoffice-profile"
        source_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()

        source_paths: list[Path] = []
        for case in FORMULA_CASES:
            source_path = source_dir / f"{case.case_id}.xlsx"
            _create_case_workbook(case, source_path)
            source_paths.append(source_path)

        recalculator(tuple(source_paths), output_dir, profile_dir, soffice)

        case_results: list[dict[str, Any]] = []
        for case, source_path in zip(FORMULA_CASES, source_paths, strict=True):
            actual = _read_cached_result(output_dir / source_path.name)
            passed = _values_equal(actual, case.expected)
            case_results.append(
                {
                    "actual": actual,
                    "expected": case.expected,
                    "formula": case.formula,
                    "function": case.function,
                    "id": case.case_id,
                    "passed": passed,
                    "stored_name": case.stored_name,
                }
            )

    passed_count = sum(bool(result["passed"]) for result in case_results)
    total = len(case_results)
    return {
        "cases": case_results,
        "engine": {
            "executable": str(soffice.resolve()),
            "version": version,
        },
        "schema_version": SCHEMA_VERSION,
        "summary": {
            "failed": total - passed_count,
            "passed": passed_count,
            "total": total,
        },
    }


def render_json(result: dict[str, Any]) -> str:
    """Return the stable on-disk and stdout representation."""

    return json.dumps(result, indent=2, sort_keys=True, ensure_ascii=False) + "\n"


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--soffice",
        required=True,
        help="Path or command name for the LibreOffice soffice executable.",
    )
    parser.add_argument(
        "--out",
        type=Path,
        help="Optional path for the deterministic JSON result (stdout is always written).",
    )
    return parser.parse_args(argv)


def main(
    argv: Sequence[str] | None = None,
    *,
    recalculator: Recalculator = _recalculate_with_soffice,
    engine_version: str | None = None,
) -> int:
    args = parse_args(argv)
    try:
        soffice = _resolve_soffice(args.soffice)
        result = run_compatibility_check(
            soffice,
            recalculator=recalculator,
            engine_version=engine_version,
        )
        rendered = render_json(result)
        if args.out is not None:
            args.out.parent.mkdir(parents=True, exist_ok=True)
            args.out.write_text(rendered, encoding="utf-8")
        sys.stdout.write(rendered)
    except (CompatibilityError, OSError) as exc:
        print(f"formula compatibility check failed: {exc}", file=sys.stderr)
        return 2
    return 0 if result["summary"]["failed"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
