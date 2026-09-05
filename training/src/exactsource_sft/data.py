"""Build auditable SFT conversations from hand-authored synthetic workbooks."""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import tempfile
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path
from typing import Any, Literal

import openpyxl
from exactsource.artifacts import atomic_write_text
from exactsource.context import build_context
from exactsource.contracts import JsonScalar, QualifiedRange, SolvePlan, TaskSpec
from exactsource.plans import apply_operations
from exactsource.prompts import (
    CELL_SYSTEM_PROMPT,
    PROMPT_PLAN_SCHEMA_TEXT,
    SYSTEM_PROMPT,
    build_messages,
)
from exactsource.ranges import iter_range_coordinates, normalise_a1_range
from exactsource.sandbox import run_transform
from pydantic import BaseModel, ConfigDict, Field, model_validator

MODEL_ID = "Qwen/Qwen3.8-27B"
RENDERER_NAME = "qwen3_8_xhigh_reasoning"
TRAIN_ON_WHAT = "LAST_ASSISTANT_MESSAGE"
MAX_LENGTH = 32_768
TRAINING_SEED = 240_905
TOKENIZER_REVISION = "1d4bf0f2ff6012fd82039f2fa52739d0dd7c60c0"
TOKENIZER_VOCAB_SIZE = 248_044
TINKER_VERSION = "0.27.1"
COOKBOOK_VERSION = "0.5.7"
TRANSFORMERS_VERSION = "5.5.4"
OPENPYXL_VERSION = "3.1.5"
PYDANTIC_VERSION = "2.13.5"
SOFFICE_TIMEOUT_SECONDS = 120
SOFFICE_ENV = "EXACTSOURCE_SOFFICE"

EXPECTED_DEPENDENCIES = {
    "openpyxl": OPENPYXL_VERSION,
    "pydantic": PYDANTIC_VERSION,
    "tinker": TINKER_VERSION,
    "tinker-cookbook": COOKBOOK_VERSION,
    "transformers": TRANSFORMERS_VERSION,
}

TRAINING_DIR = Path(__file__).resolve().parents[2]
REPOSITORY_DIR = TRAINING_DIR.parent
DEFAULT_CASE_FILE = TRAINING_DIR / "cases" / "synthetic_v1.json"
DEFAULT_OUTPUT_DIR = REPOSITORY_DIR / "scratch" / "sft"


class _StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True, allow_inf_nan=False)


class SheetSpec(_StrictModel):
    name: str = Field(min_length=1, max_length=31)
    cells: dict[str, JsonScalar]

    @model_validator(mode="after")
    def excel_safe_name_and_cells(self) -> SheetSpec:
        if re.search(r"[\\/*?:\[\]!]", self.name):
            raise ValueError(f"worksheet name contains a prohibited character: {self.name!r}")
        for coordinate in self.cells:
            if ":" in coordinate or "!" in coordinate:
                raise ValueError("workbook cells must use one local A1 address")
            if normalise_a1_range(coordinate) != coordinate.upper().replace("$", ""):
                raise ValueError(f"workbook cell is not canonical A1 notation: {coordinate!r}")
        return self


class WorkbookSpec(_StrictModel):
    sheets: list[SheetSpec] = Field(min_length=1)

    @model_validator(mode="after")
    def unique_sheet_names(self) -> WorkbookSpec:
        names = [sheet.name.casefold() for sheet in self.sheets]
        if len(names) != len(set(names)):
            raise ValueError(
                "synthetic workbook contains case-insensitive duplicate worksheet names"
            )
        return self


class RangeSpec(_StrictModel):
    sheet: str = Field(min_length=1, max_length=31)
    range: str = Field(min_length=1)


class Expectations(_StrictModel):
    cells: dict[str, JsonScalar] = Field(min_length=1)
    preserved: dict[str, JsonScalar] = Field(min_length=1)
    calculated: dict[str, JsonScalar]


class SyntheticCase(_StrictModel):
    id: str = Field(pattern=r"^synthetic-[a-z0-9][a-z0-9-]*$")
    split: Literal["train", "tune"]
    description: str = Field(min_length=1)
    instruction_type: Literal["Cell-Level Manipulation", "Sheet-Level Manipulation"]
    instruction: str = Field(min_length=1)
    answer_ranges: list[RangeSpec] = Field(min_length=1)
    data_position: str | None = None
    workbook: WorkbookSpec
    plan: SolvePlan
    expectations: Expectations

    @model_validator(mode="after")
    def validate_task_and_expectation_contract(self) -> SyntheticCase:
        is_cell = self.instruction_type.startswith("Cell")
        if is_cell and self.plan.route != "operations":
            raise ValueError("cell-level synthetic cases must use the operations route")

        initial_sheets = {sheet.name.casefold() for sheet in self.workbook.sheets}
        answer_cells: set[str] = set()
        for answer_range in self.answer_ranges:
            if re.search(r"[\\/*?:\[\]!]", answer_range.sheet):
                raise ValueError("answer-range worksheet name contains a prohibited character")
            if is_cell and answer_range.sheet.casefold() not in initial_sheets:
                raise ValueError("cell-level answer ranges must use an initial worksheet")
            for coordinate in iter_range_coordinates(answer_range.range, max_cells=1_000):
                answer_cells.add(f"{answer_range.sheet}!{coordinate}")
        expected_cells = set(self.expectations.cells)
        if expected_cells != answer_cells:
            missing = sorted(answer_cells - expected_cells)
            extra = sorted(expected_cells - answer_cells)
            raise ValueError(
                "expectations.cells must cover every declared answer cell exactly; "
                f"missing={missing}, extra={extra}"
            )
        preserved = set(self.expectations.preserved)
        if preserved & answer_cells:
            raise ValueError("preserved expectations must be outside declared answer ranges")
        formula_cells = {
            reference
            for reference, value in self.expectations.cells.items()
            if isinstance(value, str) and value.startswith("=")
        }
        if set(self.expectations.calculated) != formula_cells:
            raise ValueError("expectations.calculated must cover every formula answer cell exactly")
        return self


class CaseFile(_StrictModel):
    schema_version: Literal[1]
    cases: list[SyntheticCase] = Field(min_length=2)

    @model_validator(mode="after")
    def unique_ids_and_explicit_splits(self) -> CaseFile:
        ids = [case.id for case in self.cases]
        if len(ids) != len(set(ids)):
            raise ValueError("synthetic case file contains duplicate ids")
        splits = {case.split for case in self.cases}
        if splits != {"train", "tune"}:
            raise ValueError("synthetic case file must contain train and tune cases")
        return self


@dataclass(frozen=True, slots=True)
class PreparedRecord:
    id: str
    split: Literal["train", "tune"]
    messages: list[dict[str, str]]


FormulaVerifier = Callable[
    [Mapping[str, tuple[Path, Mapping[str, JsonScalar]]], Path, Path | str | None],
    str,
]


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _reject_json_constant(value: str) -> None:
    raise ValueError(f"non-standard JSON numeric constant is forbidden: {value}")


def _load_json(text: str, *, label: str) -> Any:
    try:
        return json.loads(text, parse_constant=_reject_json_constant)
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise ValueError(f"cannot parse {label}: {exc}") from exc


def _installed_dependency_versions() -> dict[str, str]:
    installed: dict[str, str] = {}
    for distribution, expected in EXPECTED_DEPENDENCIES.items():
        try:
            actual = version(distribution)
        except PackageNotFoundError as exc:
            raise RuntimeError(
                f"required training package is not installed: {distribution}"
            ) from exc
        if actual != expected:
            raise RuntimeError(
                f"training package version drift for {distribution}: "
                f"expected {expected}, got {actual}"
            )
        installed[distribution] = actual
    return installed


def _safe_case_path(path: Path, *, cases_root: Path | None = None) -> Path:
    root = (cases_root or (TRAINING_DIR / "cases")).resolve(strict=True)
    candidate = Path(path)
    if candidate.is_symlink():
        raise ValueError("synthetic case file must not be a symbolic link")
    try:
        resolved = candidate.resolve(strict=True)
        resolved.relative_to(root)
    except (FileNotFoundError, OSError, ValueError) as exc:
        raise ValueError(
            "synthetic case file must be a regular file beneath training/cases"
        ) from exc
    if not resolved.is_file():
        raise ValueError("synthetic case path is not a regular file")
    return resolved


def load_case_file(path: Path, *, cases_root: Path | None = None) -> tuple[Path, CaseFile]:
    resolved = _safe_case_path(path, cases_root=cases_root)
    try:
        text = resolved.read_text(encoding="utf-8")
    except (OSError, UnicodeError) as exc:
        raise ValueError(f"cannot read synthetic case JSON: {exc}") from exc
    payload = _load_json(text, label="synthetic case JSON")
    return resolved, CaseFile.model_validate(payload)


def _write_workbook(case: SyntheticCase, destination: Path) -> None:
    workbook = openpyxl.Workbook()
    try:
        workbook.remove(workbook.active)
        for spec in case.workbook.sheets:
            sheet = workbook.create_sheet(spec.name)
            for coordinate, value in spec.cells.items():
                sheet[coordinate] = value
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(destination)
    finally:
        workbook.close()


def _task(case: SyntheticCase, workbook: Path) -> TaskSpec:
    return TaskSpec(
        id=case.id,
        instruction_type=case.instruction_type,
        instruction=case.instruction,
        spreadsheet_path=f"synthetic/{case.id}",
        init_xlsx=workbook,
        answer_ranges=tuple(
            QualifiedRange(sheet=item.sheet, cells=item.range) for item in case.answer_ranges
        ),
        data_position=case.data_position,
    )


def _qualified_cell(workbook: openpyxl.Workbook, reference: str):
    if reference.count("!") != 1:
        raise ValueError(f"expected a Sheet!A1 reference, got {reference!r}")
    sheet_name, coordinate = reference.split("!", 1)
    if sheet_name not in workbook.sheetnames or ":" in coordinate:
        raise ValueError(f"invalid expected cell reference {reference!r}")
    return workbook[sheet_name][coordinate]


def _assert_cells(
    workbook_path: Path,
    expected: Mapping[str, JsonScalar],
    *,
    case_id: str,
    label: str,
    data_only: bool = False,
) -> None:
    workbook = openpyxl.load_workbook(
        workbook_path,
        data_only=data_only,
        keep_links=False,
    )
    try:
        for reference, value in expected.items():
            actual = _qualified_cell(workbook, reference).value
            if not _values_equal(actual, value):
                raise ValueError(
                    f"case {case_id}: {label} assertion failed at {reference}; "
                    f"expected {value!r}, got {actual!r}"
                )
    finally:
        workbook.close()


def _values_equal(actual: object, expected: object) -> bool:
    if isinstance(actual, bool) or isinstance(expected, bool):
        return type(actual) is type(expected) and actual == expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return (
            math.isfinite(float(actual))
            and math.isfinite(float(expected))
            and math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-12)
        )
    return type(actual) is type(expected) and actual == expected


def _answer_cell_references(case: SyntheticCase) -> set[str]:
    return {
        f"{item.sheet}!{coordinate}"
        for item in case.answer_ranges
        for coordinate in iter_range_coordinates(item.range, max_cells=1_000)
    }


def _cell_state(cell: Any) -> tuple[object, ...] | None:
    hyperlink = cell.hyperlink.target if cell.hyperlink is not None else None
    comment = None if cell.comment is None else (cell.comment.author, cell.comment.text)
    if cell.value is None and not cell.has_style and hyperlink is None and comment is None:
        return None
    return (
        cell.value,
        cell.data_type,
        cell.style_id,
        cell.number_format,
        hyperlink,
        comment,
    )


def _source_sheet_snapshot(
    workbook_path: Path,
    *,
    source_sheets: Sequence[str],
    excluded: set[str],
) -> dict[str, object]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, keep_links=False)
    try:
        snapshots: dict[str, object] = {}
        for sheet_name in source_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"output workbook removed source worksheet {sheet_name!r}")
            sheet = workbook[sheet_name]
            if sheet.max_row * sheet.max_column > 20_000:
                raise ValueError("synthetic preservation snapshot exceeds 20,000 cells")
            cells: dict[str, tuple[object, ...]] = {}
            for row in sheet.iter_rows():
                for cell in row:
                    reference = f"{sheet_name}!{cell.coordinate}"
                    if reference in excluded:
                        continue
                    state = _cell_state(cell)
                    if state is not None:
                        cells[cell.coordinate] = state
            snapshots[sheet_name] = {
                "cells": cells,
                "merged_cells": tuple(sorted(str(item) for item in sheet.merged_cells.ranges)),
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "sheet_state": sheet.sheet_state,
                "auto_filter": sheet.auto_filter.ref,
            }
        return snapshots
    finally:
        workbook.close()


def _assert_scope_preserved(case: SyntheticCase, source: Path, output: Path) -> None:
    source_names = [sheet.name for sheet in case.workbook.sheets]
    allowed_cells = _answer_cell_references(case)
    before = _source_sheet_snapshot(source, source_sheets=source_names, excluded=allowed_cells)
    after = _source_sheet_snapshot(output, source_sheets=source_names, excluded=allowed_cells)
    if after != before:
        raise ValueError(f"case {case.id}: cells outside declared answer ranges changed")

    workbook = openpyxl.load_workbook(output, data_only=False, keep_links=False)
    try:
        allowed_new_sheets = {
            item.sheet for item in case.answer_ranges if item.sheet not in source_names
        }
        expected_names = [*source_names, *sorted(allowed_new_sheets)]
        if workbook.sheetnames != expected_names:
            raise ValueError(
                f"case {case.id}: unexpected output worksheets; "
                f"expected {expected_names}, got {workbook.sheetnames}"
            )
        for sheet_name in allowed_new_sheets:
            sheet = workbook[sheet_name]
            if sheet.max_row * sheet.max_column > 20_000:
                raise ValueError("synthetic output worksheet exceeds 20,000 cells")
            for row in sheet.iter_rows():
                for cell in row:
                    state = _cell_state(cell)
                    reference = f"{sheet_name}!{cell.coordinate}"
                    if state is not None and reference not in allowed_cells:
                        raise ValueError(
                            f"case {case.id}: populated output {reference} is outside answer ranges"
                        )
    finally:
        workbook.close()


def _exercise_plan(case: SyntheticCase, task: TaskSpec, source: Path, output: Path) -> None:
    _assert_cells(
        source,
        case.expectations.preserved,
        case_id=case.id,
        label="source preservation",
    )
    if case.plan.route == "operations":
        apply_operations(case.plan, task, source, output)
    else:
        assert case.plan.python_code is not None
        run_transform(case.plan.python_code, source, output)
    _assert_cells(output, case.expectations.cells, case_id=case.id, label="output")
    _assert_cells(
        output,
        case.expectations.preserved,
        case_id=case.id,
        label="preservation",
    )
    _assert_scope_preserved(case, source, output)


def _resolve_soffice(value: Path | str | None) -> Path:
    candidates: list[Path | str] = []
    if value is not None:
        candidates.append(value)
    elif os.environ.get(SOFFICE_ENV, "").strip():
        candidates.append(os.environ[SOFFICE_ENV])
    else:
        found = shutil.which("soffice")
        if found is not None:
            candidates.append(found)
        candidates.append("/Applications/LibreOffice.app/Contents/MacOS/soffice")

    for candidate in candidates:
        path = Path(candidate).expanduser().resolve(strict=False)
        if path.is_file() and os.access(path, os.X_OK):
            return path
    raise ValueError(
        "LibreOffice is required for independent formula-label verification; "
        f"set {SOFFICE_ENV} to the soffice executable"
    )


def _soffice_version(soffice: Path) -> str:
    try:
        completed = subprocess.run(
            [str(soffice), "--version"],
            capture_output=True,
            text=True,
            check=False,
            timeout=30,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise ValueError(f"cannot read LibreOffice version: {exc}") from exc
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout).strip()
        raise ValueError(f"LibreOffice version command exited {completed.returncode}: {detail}")
    version_text = next(
        (line.strip() for line in completed.stdout.splitlines() if line.strip()),
        "",
    )
    if not version_text:
        raise ValueError("LibreOffice version command returned no output")
    return version_text


def _verify_formula_outputs(
    outputs: Mapping[str, tuple[Path, Mapping[str, JsonScalar]]],
    work_root: Path,
    soffice_value: Path | str | None,
) -> str:
    if not outputs:
        return "not-required"
    soffice = _resolve_soffice(soffice_value)
    recalculated = work_root / "recalculated"
    profile = work_root / "libreoffice-profile"
    recalculated.mkdir()
    profile.mkdir()
    command = [
        str(soffice),
        f"-env:UserInstallation={profile.resolve().as_uri()}",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(recalculated),
        *(str(output[0]) for output in outputs.values()),
    ]
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False,
            timeout=SOFFICE_TIMEOUT_SECONDS,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise ValueError(f"LibreOffice formula recalculation failed: {exc}") from exc
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout).strip()
        raise ValueError(
            f"LibreOffice formula recalculation exited {completed.returncode}: {detail}"
        )
    for case_id, (source, expected) in outputs.items():
        result = recalculated / source.name
        if not result.is_file():
            raise ValueError(f"LibreOffice did not produce recalculated output for {case_id}")
        _assert_cells(
            result,
            expected,
            case_id=case_id,
            label="recalculated formula",
            data_only=True,
        )
    return _soffice_version(soffice)


def _load_renderer():
    # Dataset preparation must never contact Tinker or Hugging Face. The exact
    # tokenizer revision must already be present in the local cache.
    os.environ["HF_HUB_OFFLINE"] = "1"
    os.environ["TRANSFORMERS_OFFLINE"] = "1"
    from tinker_cookbook import model_info, renderers
    from transformers import AutoTokenizer

    recommended = model_info.get_recommended_renderer_name(MODEL_ID)
    if recommended != RENDERER_NAME:
        raise RuntimeError(
            f"unexpected recommended renderer {recommended!r}; expected {RENDERER_NAME!r}"
        )
    tokenizer = AutoTokenizer.from_pretrained(
        MODEL_ID,
        revision=TOKENIZER_REVISION,
        local_files_only=True,
    )
    if tokenizer.vocab_size != TOKENIZER_VOCAB_SIZE:
        raise RuntimeError(
            "pinned tokenizer vocabulary does not match the recorded revision: "
            f"expected {TOKENIZER_VOCAB_SIZE}, got {tokenizer.vocab_size}"
        )
    return renderers.get_renderer(RENDERER_NAME, tokenizer)


def _render_lengths(renderer, messages: list[dict[str, str]]) -> tuple[int, int]:
    from tinker_cookbook.renderers import TrainOnWhat

    model_input, weights = renderer.build_supervised_example(
        messages,
        train_on_what=TrainOnWhat.LAST_ASSISTANT_MESSAGE,
    )
    rendered_tokens = model_input.length
    target_tokens = int((weights > 0).sum().item())
    if rendered_tokens > MAX_LENGTH:
        raise ValueError(
            f"rendered training example has {rendered_tokens} tokens; limit is {MAX_LENGTH}"
        )
    if target_tokens < 1:
        raise ValueError("rendered training example has no positive-weight target tokens")
    return rendered_tokens, target_tokens


def _jsonl(records: Sequence[Mapping[str, object]]) -> str:
    return "".join(
        json.dumps(
            record,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
            allow_nan=False,
        )
        + "\n"
        for record in records
    )


def _build_records(
    case_file: CaseFile,
    *,
    renderer: object | None,
    formula_verifier: FormulaVerifier | None,
    soffice: Path | str | None = None,
) -> tuple[
    dict[str, list[dict[str, object]]],
    dict[str, dict[str, int]],
    str | None,
]:
    records: dict[str, list[dict[str, object]]] = {"train": [], "tune": []}
    token_counts: dict[str, dict[str, int]] = {}

    with tempfile.TemporaryDirectory(prefix="exactsource-sft-build-") as directory:
        work_root = Path(directory)
        source_dir = work_root / "source"
        output_dir = work_root / "output"
        source_dir.mkdir()
        output_dir.mkdir()
        formula_outputs: dict[str, tuple[Path, Mapping[str, JsonScalar]]] = {}

        for case in case_file.cases:
            source = source_dir / f"{case.id}.xlsx"
            output = output_dir / f"{case.id}.xlsx"
            _write_workbook(case, source)
            task = _task(case, source)
            context = build_context(task)
            messages = build_messages(task, context)
            assistant = case.plan.model_dump_json(exclude_none=False)
            conversation = [*messages, {"role": "assistant", "content": assistant}]
            _exercise_plan(case, task, source, output)
            if case.expectations.calculated:
                formula_outputs[case.id] = (output, case.expectations.calculated)
            if renderer is not None:
                rendered_tokens, target_tokens = _render_lengths(renderer, conversation)
                token_counts[case.id] = {
                    "rendered_tokens": rendered_tokens,
                    "target_tokens": target_tokens,
                }

            records[case.split].append(
                {
                    "id": case.id,
                    "messages": conversation,
                    "provenance": {
                        "case": case.id,
                        "kind": "synthetic",
                        "split": case.split,
                    },
                }
            )

        calculation_engine = None
        if formula_verifier is not None:
            calculation_engine = formula_verifier(formula_outputs, work_root, soffice)
    return records, token_counts, calculation_engine


def prepare_dataset(
    case_path: Path = DEFAULT_CASE_FILE,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    *,
    renderer=None,
    formula_verifier: FormulaVerifier = _verify_formula_outputs,
    soffice: Path | str | None = None,
) -> dict[str, object]:
    """Validate synthetic plans, render them and publish deterministic JSONL."""

    dependency_versions = _installed_dependency_versions()
    resolved_case_path, case_file = load_case_file(case_path)
    active_renderer = renderer or _load_renderer()
    records, token_counts, calculation_engine = _build_records(
        case_file,
        renderer=active_renderer,
        formula_verifier=formula_verifier,
        soffice=soffice,
    )
    assert calculation_engine is not None

    train_text = _jsonl(records["train"])
    tune_text = _jsonl(records["tune"])
    output_dir = Path(output_dir)
    train_path = output_dir / "train.jsonl"
    tune_path = output_dir / "tune.jsonl"
    manifest_path = output_dir / "manifest.json"
    atomic_write_text(train_path, train_text)
    atomic_write_text(tune_path, tune_text)

    manifest: dict[str, object] = {
        "schema_version": 1,
        "method": (
            "hand-authored synthetic workbooks; labels executed, scope-compared "
            "and formula results independently recalculated before export"
        ),
        "model": MODEL_ID,
        "renderer": RENDERER_NAME,
        "train_on_what": TRAIN_ON_WHAT,
        "max_length": MAX_LENGTH,
        "seed": TRAINING_SEED,
        "empty_thinking_target_ablation": True,
        "benchmark_goldens_used": False,
        "benchmark_initial_workbooks_used": False,
        "formula_verification": {
            "engine": "LibreOffice",
            "version": calculation_engine,
            "case_ids": [case.id for case in case_file.cases if case.expectations.calculated],
        },
        "source_case": str(resolved_case_path.relative_to(REPOSITORY_DIR)),
        "source_case_sha256": _sha256_file(resolved_case_path),
        "train_jsonl_sha256": _sha256_bytes(train_text.encode("utf-8")),
        "tune_jsonl_sha256": _sha256_bytes(tune_text.encode("utf-8")),
        "system_prompt_sha256": {
            "cell": _sha256_bytes(CELL_SYSTEM_PROMPT.encode("utf-8")),
            "sheet": _sha256_bytes(SYSTEM_PROMPT.encode("utf-8")),
        },
        "solve_schema_sha256": _sha256_bytes(PROMPT_PLAN_SCHEMA_TEXT.encode("utf-8")),
        "ordered_train_ids": [case.id for case in case_file.cases if case.split == "train"],
        "ordered_tune_ids": [case.id for case in case_file.cases if case.split == "tune"],
        "rendered_token_counts": token_counts,
        "dependencies": dependency_versions,
        "tokenizer": {
            "repository": MODEL_ID,
            "revision": TOKENIZER_REVISION,
            "vocabulary_size": TOKENIZER_VOCAB_SIZE,
        },
    }
    atomic_write_text(
        manifest_path,
        json.dumps(
            manifest,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            allow_nan=False,
        )
        + "\n",
    )
    return manifest


def verify_prepared_dataset(
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    *,
    case_path: Path = DEFAULT_CASE_FILE,
) -> dict[str, object]:
    """Rebuild and verify prepared records from source before any paid call."""

    output_dir = Path(output_dir)
    manifest_path = output_dir / "manifest.json"
    if manifest_path.is_symlink():
        raise ValueError("SFT manifest must not be a symbolic link")
    try:
        manifest_text = manifest_path.read_text(encoding="utf-8")
    except (OSError, UnicodeError) as exc:
        raise ValueError(f"cannot read SFT manifest: {exc}") from exc
    manifest = _load_json(manifest_text, label="SFT manifest")
    if not isinstance(manifest, dict) or manifest.get("schema_version") != 1:
        raise ValueError("unsupported SFT manifest")
    expected_scalars = {
        "model": MODEL_ID,
        "renderer": RENDERER_NAME,
        "train_on_what": TRAIN_ON_WHAT,
        "max_length": MAX_LENGTH,
        "seed": TRAINING_SEED,
        "empty_thinking_target_ablation": True,
        "benchmark_goldens_used": False,
        "benchmark_initial_workbooks_used": False,
    }
    for key, expected in expected_scalars.items():
        if manifest.get(key) != expected:
            raise ValueError(f"SFT manifest field {key!r} does not match the fixed configuration")

    resolved_case, case_file = load_case_file(case_path)
    expected_records, _token_counts, _engine = _build_records(
        case_file,
        renderer=None,
        formula_verifier=None,
    )
    expected_texts = {
        "train": _jsonl(expected_records["train"]),
        "tune": _jsonl(expected_records["tune"]),
    }
    for split, expected_text in expected_texts.items():
        path = output_dir / f"{split}.jsonl"
        if path.is_symlink():
            raise ValueError(f"prepared {split} JSONL must not be a symbolic link")
        try:
            actual = path.read_text(encoding="utf-8")
        except (OSError, UnicodeError) as exc:
            raise ValueError(f"cannot read prepared {split} JSONL: {exc}") from exc
        if actual != expected_text:
            raise ValueError(
                f"prepared {split} JSONL does not reproduce from the reviewed case source"
            )

    expected_train_ids = [case.id for case in case_file.cases if case.split == "train"]
    expected_tune_ids = [case.id for case in case_file.cases if case.split == "tune"]
    if manifest.get("ordered_train_ids") != expected_train_ids:
        raise ValueError("SFT manifest train IDs do not match the reviewed case source")
    if manifest.get("ordered_tune_ids") != expected_tune_ids:
        raise ValueError("SFT manifest tune IDs do not match the reviewed case source")
    expected_source_name = str(resolved_case.relative_to(REPOSITORY_DIR))
    if manifest.get("source_case") != expected_source_name:
        raise ValueError("SFT manifest source path does not match the reviewed case source")

    checks = {
        "source_case_sha256": _sha256_file(resolved_case),
        "train_jsonl_sha256": _sha256_file(output_dir / "train.jsonl"),
        "tune_jsonl_sha256": _sha256_file(output_dir / "tune.jsonl"),
        "solve_schema_sha256": _sha256_bytes(PROMPT_PLAN_SCHEMA_TEXT.encode("utf-8")),
    }
    for key, actual in checks.items():
        if manifest.get(key) != actual:
            raise ValueError(f"SFT manifest hash mismatch for {key}")
    expected_prompts = {
        "cell": _sha256_bytes(CELL_SYSTEM_PROMPT.encode("utf-8")),
        "sheet": _sha256_bytes(SYSTEM_PROMPT.encode("utf-8")),
    }
    if manifest.get("system_prompt_sha256") != expected_prompts:
        raise ValueError("SFT manifest system-prompt hashes do not match the current source")
    if manifest.get("dependencies") != _installed_dependency_versions():
        raise ValueError("SFT manifest dependencies do not match the locked environment")
    expected_tokenizer = {
        "repository": MODEL_ID,
        "revision": TOKENIZER_REVISION,
        "vocabulary_size": TOKENIZER_VOCAB_SIZE,
    }
    if manifest.get("tokenizer") != expected_tokenizer:
        raise ValueError("SFT manifest tokenizer does not match the fixed configuration")

    formula_case_ids = [case.id for case in case_file.cases if case.expectations.calculated]
    formula_verification = manifest.get("formula_verification")
    if (
        not isinstance(formula_verification, dict)
        or formula_verification.get("engine") != "LibreOffice"
        or not isinstance(formula_verification.get("version"), str)
        or not formula_verification["version"].strip()
        or formula_verification.get("case_ids") != formula_case_ids
    ):
        raise ValueError("SFT manifest formula-verification evidence is incomplete")

    token_counts = manifest.get("rendered_token_counts")
    all_ids = [*expected_train_ids, *expected_tune_ids]
    if not isinstance(token_counts, dict) or set(token_counts) != set(all_ids):
        raise ValueError("SFT manifest rendered-token IDs are incomplete")
    for case_id in all_ids:
        item = token_counts[case_id]
        if (
            not isinstance(item, dict)
            or type(item.get("rendered_tokens")) is not int
            or not 1 <= item["rendered_tokens"] <= MAX_LENGTH
            or type(item.get("target_tokens")) is not int
            or not 1 <= item["target_tokens"] <= item["rendered_tokens"]
        ):
            raise ValueError(f"SFT manifest has invalid token counts for {case_id}")

    read_prepared_records(output_dir / "train.jsonl", expected_split="train")
    read_prepared_records(output_dir / "tune.jsonl", expected_split="tune")
    return manifest


def read_prepared_records(
    path: Path,
    *,
    expected_split: Literal["train", "tune"],
) -> list[PreparedRecord]:
    """Read prepared rows after enforcing IDs, provenance and messages."""

    source = Path(path)
    if source.is_symlink():
        raise ValueError(f"prepared SFT file must not be a symbolic link: {path}")
    raw = source.read_bytes()
    if not raw or not raw.endswith(b"\n"):
        raise ValueError(f"prepared SFT file is empty or lacks a final newline: {path}")
    records: list[PreparedRecord] = []
    seen_ids: set[str] = set()
    for line_number, line in enumerate(raw.decode("utf-8").splitlines(), start=1):
        row = _load_json(line, label=f"prepared SFT row {line_number}")
        if not isinstance(row, dict) or set(row) != {"id", "messages", "provenance"}:
            raise ValueError(f"prepared SFT row {line_number} has invalid top-level fields")
        case_id = row["id"]
        if (
            not isinstance(case_id, str)
            or re.fullmatch(r"synthetic-[a-z0-9][a-z0-9-]*", case_id) is None
            or case_id in seen_ids
        ):
            raise ValueError(f"prepared SFT row {line_number} has an invalid or duplicate id")
        provenance = row["provenance"]
        expected_provenance = {
            "case": case_id,
            "kind": "synthetic",
            "split": expected_split,
        }
        if provenance != expected_provenance:
            raise ValueError(f"prepared SFT row {line_number} has invalid provenance")
        messages = row["messages"]
        if not isinstance(messages, list) or not all(
            isinstance(message, dict) and set(message) == {"role", "content"}
            for message in messages
        ):
            raise ValueError(f"prepared SFT row {line_number} has invalid message objects")
        roles = [message["role"] for message in messages]
        if roles != ["system", "user", "assistant"]:
            raise ValueError(f"prepared SFT row {line_number} has an invalid conversation shape")
        if not all(
            isinstance(message["role"], str) and isinstance(message["content"], str)
            for message in messages  # type: ignore[union-attr]
        ):
            raise ValueError(f"prepared SFT row {line_number} contains invalid message content")
        typed_messages = [
            {"role": str(message["role"]), "content": str(message["content"])}
            for message in messages
        ]
        SolvePlan.model_validate_json(typed_messages[-1]["content"])
        records.append(PreparedRecord(id=case_id, split=expected_split, messages=typed_messages))
        seen_ids.add(case_id)
    return records


def read_conversations(path: Path) -> list[list[dict[str, str]]]:
    """Compatibility wrapper returning validated training conversations."""

    return [record.messages for record in read_prepared_records(path, expected_split="train")]
