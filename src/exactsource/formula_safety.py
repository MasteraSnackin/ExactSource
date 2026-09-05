"""Conservative integrity and external-capability checks for generated formulae.

Excel formulae are normally pure calculations over workbook data, but a small
set of functions and link syntaxes can contact external services, load another
workbook, or invoke legacy integration mechanisms. Generated formulae are
screened before a workbook is saved so opening or recalculating the result does
not unexpectedly exercise those capabilities. Quote-aware delimiter checks and
confidently parsed worksheet references reject basic malformed output without
pretending to implement Excel's complete and evolving formula grammar.

The scanner deliberately ignores function-looking text inside Excel string
literals and quoted worksheet names. Generated ``HYPERLINK`` and ``IMAGE`` calls
are rejected outright because a dynamic target cannot be proved internal from
formula text alone. Other data-dependent references, such as ``INDIRECT(A1)``,
cannot be semantically proved from formula text and remain allowed unless static
external syntax is present. Existing formulae are fingerprinted before a Python
transform and are grandfathered only when their type, text and stable workbook
location all remain unchanged.
"""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from typing import TypeAlias

from openpyxl.formula.tokenizer import Token, Tokenizer, TokenizerError
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

FormulaCell: TypeAlias = tuple[str, str]
FormulaSnapshot: TypeAlias = dict[tuple[str, str], FormulaCell]
MetadataLocation: TypeAlias = tuple[object, ...]
FormulaMetadataSnapshot: TypeAlias = dict[MetadataLocation, FormulaCell]
HyperlinkSnapshot: TypeAlias = dict[tuple[str, str], tuple[str | None, str | None]]

_BLOCKED_FUNCTION_RE = re.compile(
    r"(?<![A-Za-z0-9_.])"
    r"(?:(?:_xlfn|_xlws)\.)*"
    r"(?:WEBSERVICE|RTD|CALL|REGISTER(?:\.ID)?|EXEC|RUN|EVALUATE)\s*\(",
    flags=re.IGNORECASE,
)
_LINK_FUNCTION_RE = re.compile(
    r"(?<![A-Za-z0-9_.])"
    r"(?:(?:_xlfn|_xlws)\.)*"
    r"(?P<name>HYPERLINK|IMAGE)\s*\(",
    flags=re.IGNORECASE,
)
_INDIRECT_RE = re.compile(
    r"(?<![A-Za-z0-9_.])(?:(?:_xlfn|_xlws)\.)*INDIRECT\s*\(",
    flags=re.IGNORECASE,
)

# A structured table reference never owns the following ``!`` worksheet
# separator. Requiring that separator avoids treating ``Sales[Amount]`` or
# ``Sales[[#Data],[Amount]]`` as links to another workbook.
_EXTERNAL_WORKBOOK_RE = re.compile(
    r"(?:'(?:[^'\r\n]|'')*\[[^\[\]\r\n]+\](?:[^'\r\n]|'')*'|"
    r"\[[^\[\]\r\n]+\][A-Za-z0-9_.]+)\s*!",
    flags=re.IGNORECASE,
)
_FILE_URI_RE = re.compile(r"(?i)(?<![A-Za-z0-9+.-])file:(?:/{0,3})")
_UNC_RE = re.compile(r"\\\\[^\\\r\n]")
_INTERNAL_HYPERLINK_RE = re.compile(
    r"^(?:"
    r"(?:'(?:[^'\[\]\\/:?*\r\n]|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!"
    r"\$?[A-Za-z]{1,3}\$?[1-9][0-9]*"
    r"(?::\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)?"
    r"|\$?[A-Za-z]{1,3}\$?[1-9][0-9]*"
    r"(?::\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)?"
    r"|[A-Za-z_\\][A-Za-z0-9_.\\]*"
    r")$"
)
_EXPLICIT_SHEET_CELL_RE = re.compile(
    r"(?<![A-Za-z0-9_.:])"
    r"(?:'(?P<quoted>(?:[^'\r\n]|'')+)'|(?P<bare>[A-Za-z_\\][A-Za-z0-9_.]*))"
    r"\s*!\s*\$?[A-Za-z]{1,3}\$?[1-9][0-9]*",
    flags=re.IGNORECASE,
)
_STATIC_A1_RANGE_RE = re.compile(
    r"^(?:"
    r"(?:'(?:[^'\[\]\\/:?*\r\n]|'')+'|[A-Za-z_\\][A-Za-z0-9_.]*)!"
    r")?"
    r"\$?(?P<start_column>[A-Za-z]{1,3})\$?(?P<start_row>[1-9][0-9]*)"
    r"(?:"
    r":\$?(?P<end_column>[A-Za-z]{1,3})\$?(?P<end_row>[1-9][0-9]*)"
    r")?$"
)
_LITERAL_INTEGER_RE = re.compile(r"^[0-9]+$")
_LOOKUP_FUNCTION_RE = re.compile(
    r"^@?(?:(?:_xlfn|_xlws)\.)?(VLOOKUP|HLOOKUP)$",
    flags=re.IGNORECASE,
)
_OPENING_DELIMITERS = {"(": ")", "[": "]", "{": "}"}
_CLOSING_DELIMITERS = frozenset(_OPENING_DELIMITERS.values())
_EXCEL_MAX_COLUMN = 16_384
_EXCEL_MAX_ROW = 1_048_576


class FormulaSafetyError(ValueError):
    """Raised when generated formula text is unsafe or structurally invalid.

    Only a fixed category is included in the message. Formula text, URI text,
    sheet names and cell contents are intentionally omitted because they are
    untrusted workbook data and may include secrets.
    """

    def __init__(self, category: str) -> None:
        self.category = category
        super().__init__(f"formula rejected: {category} is not allowed")


def _outside_string_mask(formula: str) -> str:
    """Return a same-length view with Excel double-quoted strings blanked.

    Excel escapes a quote inside a string by doubling it. Keeping the result
    the same length lets the function-call scanner map matched parentheses back
    to the original formula without exposing or attempting to parse the string
    contents.
    """

    masked = list(formula)
    index = 0
    while index < len(masked):
        if masked[index] != '"':
            index += 1
            continue

        masked[index] = " "
        index += 1
        while index < len(masked):
            if masked[index] != '"':
                masked[index] = " "
                index += 1
                continue
            masked[index] = " "
            if index + 1 < len(masked) and masked[index + 1] == '"':
                masked[index + 1] = " "
                index += 2
                continue
            index += 1
            break
    return "".join(masked)


def _outside_quoted_token_mask(formula: str, string_mask: str) -> str:
    """Also blank single-quoted worksheet/name tokens outside strings.

    A worksheet may legitimately be named ``WEBSERVICE(archive)`` or contain a
    pipe character. Neither is a function call nor DDE syntax when the whole
    identifier is single-quoted. DDE separators sit *between* quoted tokens and
    therefore remain visible. Excel escapes apostrophes in an identifier by
    doubling them, mirroring its double-quoted string escaping.
    """

    masked = list(string_mask)
    index = 0
    while index < len(masked):
        # A quote which was blanked in string_mask belonged to a double-quoted
        # string, so only a still-visible apostrophe opens an identifier here.
        if formula[index] != "'" or masked[index] != "'":
            index += 1
            continue

        masked[index] = " "
        index += 1
        while index < len(masked):
            if formula[index] != "'":
                masked[index] = " "
                index += 1
                continue
            masked[index] = " "
            if index + 1 < len(masked) and formula[index + 1] == "'":
                masked[index + 1] = " "
                index += 2
                continue
            index += 1
            break
    return "".join(masked)


def _validate_formula_delimiters(formula: str) -> None:
    """Reject only confidently malformed quote and delimiter structure.

    This is intentionally not an Excel grammar parser. It ignores delimiters in
    double-quoted strings and single-quoted worksheet/name tokens, including
    Excel's doubled-character escaping. Parentheses, array-constant braces and
    structured-reference brackets must otherwise be properly nested.
    """

    stack: list[str] = []
    index = 1
    while index < len(formula):
        token = formula[index]
        if stack and stack[-1] == "]":
            # Structured-reference contents are column/name tokens rather than
            # ordinary formula grammar. Excel uses an apostrophe to escape a
            # special character there, including a literal closing bracket.
            if token == "'" and index + 1 < len(formula):
                index += 2
                continue
            if token == "[":
                stack.append("]")
            elif token == "]":
                stack.pop()
            index += 1
            continue
        if token in {'"', "'"}:
            delimiter = token
            category = (
                "unterminated formula string"
                if delimiter == '"'
                else "unterminated quoted identifier"
            )
            index += 1
            while index < len(formula):
                if formula[index] != delimiter:
                    index += 1
                    continue
                if index + 1 < len(formula) and formula[index + 1] == delimiter:
                    index += 2
                    continue
                index += 1
                break
            else:
                raise FormulaSafetyError(category)
            continue
        if token in _OPENING_DELIMITERS:
            stack.append(_OPENING_DELIMITERS[token])
        elif token in _CLOSING_DELIMITERS and (not stack or stack.pop() != token):
            raise FormulaSafetyError("unbalanced formula delimiter")
        index += 1
    if stack:
        raise FormulaSafetyError("unbalanced formula delimiter")


def _validate_explicit_sheet_references(formula: str, sheetnames: Sequence[str]) -> None:
    """Reject confidently parsed A1 references to absent worksheets.

    Formula strings are masked before matching, while quoted worksheet names are
    decoded using Excel's doubled-apostrophe rule. Ambiguous constructs such as
    INDIRECT, defined names, table references and 3-D ranges are deliberately not
    treated as proof of a missing worksheet.
    """

    available = {str(name).casefold() for name in sheetnames}
    string_mask = _outside_string_mask(formula)
    for match in _EXPLICIT_SHEET_CELL_RE.finditer(string_mask):
        quoted = match.group("quoted")
        worksheet = quoted.replace("''", "'") if quoted is not None else match.group("bare")
        if quoted is not None and worksheet is not None and ":" in worksheet:
            # Excel quotes the two endpoints together when a 3-D reference uses
            # worksheet names which require quoting, for example
            # ``'Jan 2024:Dec 2024'!B2``. A colon cannot occur in an actual
            # worksheet name, but resolving the endpoint order requires a real
            # Excel parser. Match the deliberately conservative treatment of
            # unquoted 3-D ranges and do not claim that this compound token is a
            # missing single worksheet.
            continue
        if worksheet is not None and worksheet.casefold() not in available:
            raise FormulaSafetyError("missing worksheet reference")


def validate_formula_integrity(
    formula: str,
    *,
    sheetnames: Sequence[str] | None = None,
) -> None:
    """Apply conservative safety, delimiter and explicit-reference checks.

    Function names, arity, operators and calculated results generally remain
    outside this static contract because Excel permits names, UDFs and evolving
    function sets. The one semantic invariant checked here is a provably invalid
    literal VLOOKUP or HLOOKUP index against a bounded static A1 range. Broader
    semantics require engine recalculation rather than an unsound allow-list.
    """

    validate_formula_safety(formula)
    if not isinstance(formula, str) or not formula.startswith("="):
        return
    _validate_formula_delimiters(formula)
    _validate_literal_lookup_bounds(formula)
    if sheetnames is not None:
        _validate_explicit_sheet_references(formula, sheetnames)


def _trim_token_whitespace(tokens: list[Token]) -> list[Token]:
    """Remove only whitespace surrounding a token sequence.

    Whitespace between range operands is Excel's intersection operator, so it
    must remain visible to the conservative argument checks below.
    """

    start = 0
    end = len(tokens)
    while start < end and tokens[start].type == Token.WSPACE:
        start += 1
    while end > start and tokens[end - 1].type == Token.WSPACE:
        end -= 1
    return tokens[start:end]


def _top_level_lookup_arguments(formula: str) -> tuple[str, list[list[Token]]] | None:
    """Return arguments for a whole top-level VLOOKUP or HLOOKUP call.

    The tokenizer identifies nested functions, parentheses and arrays. Calls
    wrapped in another expression, followed by an operator, or shaped in a way
    the tokenizer cannot confidently interpret are deliberately ignored.
    """

    try:
        # Leading whitespace after ``=`` has no left-hand range with which to
        # form an intersection, so it is safe to normalise before tokenising.
        token_formula = f"={formula[1:].lstrip()}"
        tokens = _trim_token_whitespace(list(Tokenizer(token_formula).items))
    except (IndexError, TokenizerError):
        return None
    if len(tokens) < 2:
        return None

    if tokens[0].type == Token.OP_PRE and tokens[0].value == "+":
        tokens = _trim_token_whitespace(tokens[1:])
    if len(tokens) < 2:
        return None

    opening = tokens[0]
    if opening.type != Token.FUNC or opening.subtype != Token.OPEN:
        return None
    function_match = _LOOKUP_FUNCTION_RE.fullmatch(opening.value[:-1])
    if function_match is None:
        return None
    function_name = function_match.group(1).upper()

    arguments: list[list[Token]] = [[]]
    depth = 1
    closing_index: int | None = None
    for index, token in enumerate(tokens[1:], start=1):
        if token.subtype == Token.OPEN and token.type in {Token.FUNC, Token.PAREN, Token.ARRAY}:
            depth += 1
        elif token.subtype == Token.CLOSE and token.type in {Token.FUNC, Token.PAREN, Token.ARRAY}:
            depth -= 1
            if depth == 0:
                closing_index = index
                break
        if token.type == Token.SEP and token.subtype == Token.ARG and depth == 1:
            arguments.append([])
        else:
            arguments[-1].append(token)

    if closing_index != len(tokens) - 1 or len(arguments) not in {3, 4}:
        return None
    return function_name, arguments


def _static_a1_range_dimensions(tokens: list[Token]) -> tuple[int, int] | None:
    """Return width and height for one unambiguous, bounded A1 reference."""

    tokens = _trim_token_whitespace(tokens)
    if len(tokens) != 1:
        return None
    token = tokens[0]
    if token.type != Token.OPERAND or token.subtype != Token.RANGE:
        return None
    match = _STATIC_A1_RANGE_RE.fullmatch(token.value)
    if match is None:
        return None

    try:
        start_column = column_index_from_string(match.group("start_column"))
        end_column_text = match.group("end_column") or match.group("start_column")
        end_column = column_index_from_string(end_column_text)
    except (TypeError, ValueError):
        return None

    start_row = _bounded_excel_row(match.group("start_row"))
    end_row = _bounded_excel_row(match.group("end_row") or match.group("start_row"))
    if start_row is None or end_row is None:
        return None
    if (
        start_column > _EXCEL_MAX_COLUMN
        or end_column > _EXCEL_MAX_COLUMN
        or end_column < start_column
        or end_row < start_row
    ):
        return None
    return end_column - start_column + 1, end_row - start_row + 1


def _bounded_excel_row(value: str) -> int | None:
    """Parse one row number without exposing or converting oversized text."""

    normalized = value.lstrip("0") or "0"
    if len(normalized) > len(str(_EXCEL_MAX_ROW)):
        return None
    try:
        row = int(normalized)
    except (TypeError, ValueError):
        return None
    return row if 1 <= row <= _EXCEL_MAX_ROW else None


def _literal_integer(tokens: list[Token]) -> int | None:
    """Return a lexically literal integer, including one unary sign."""

    tokens = [token for token in tokens if token.type != Token.WSPACE]
    sign = 1
    if len(tokens) == 2:
        prefix, number = tokens
        if prefix.type != Token.OP_PRE or prefix.value not in {"+", "-"}:
            return None
        sign = -1 if prefix.value == "-" else 1
    elif len(tokens) == 1:
        number = tokens[0]
    else:
        return None

    value = number.value
    if (
        number.type != Token.OPERAND
        or number.subtype != Token.NUMBER
        or _LITERAL_INTEGER_RE.fullmatch(value) is None
    ):
        return None
    value = value.lstrip("0") or "0"
    if len(value) > len(str(_EXCEL_MAX_ROW)):
        return sign * (_EXCEL_MAX_ROW + 1)
    try:
        return sign * int(value)
    except (TypeError, ValueError):
        return None


def _validate_literal_lookup_bounds(formula: str) -> None:
    """Reject a provably invalid literal VLOOKUP/HLOOKUP index.

    Only a whole top-level call with a static A1 table reference and lexical
    integer index is considered. Named ranges, structured references, dynamic
    arrays, calculated arguments, unions and nested calls remain untouched.
    """

    parsed = _top_level_lookup_arguments(formula)
    if parsed is None:
        return
    function_name, arguments = parsed
    dimensions = _static_a1_range_dimensions(arguments[1])
    index = _literal_integer(arguments[2])
    if dimensions is None or index is None:
        return
    width, height = dimensions
    limit = width if function_name == "VLOOKUP" else height
    if index < 1 or index > limit:
        raise FormulaSafetyError("literal lookup index outside static range")


def _function_arguments(formula: str, masked: str, opening: int) -> str:
    """Return one function's argument source, or its safe remaining suffix."""

    depth = 0
    for index in range(opening, len(masked)):
        token = masked[index]
        if token == "(":
            depth += 1
        elif token == ")":
            depth -= 1
            if depth == 0:
                return formula[opening + 1 : index]
    return formula[opening + 1 :]


def _contains_external_workbook_reference(expression: str) -> bool:
    """Detect direct and basically concatenated external workbook syntax."""

    if _EXTERNAL_WORKBOOK_RE.search(expression):
        return True
    # INDIRECT commonly builds a reference from quoted fragments. Removing
    # Excel's string delimiters, concatenation operators and whitespace joins
    # basic literal/dynamic fragments conservatively. A table structured
    # reference still lacks the worksheet ``!`` required by the link pattern.
    collapsed = re.sub(r'["&\s]', "", expression)
    return bool(_EXTERNAL_WORKBOOK_RE.search(collapsed))


def validate_formula_safety(formula: str) -> None:
    """Reject statically identifiable external or legacy capabilities.

    The exception is deliberately categorical: callers can report the failure
    without copying potentially sensitive formula or URL text into logs.
    """

    if not isinstance(formula, str) or not formula.startswith("="):
        return

    string_mask = _outside_string_mask(formula)
    token_mask = _outside_quoted_token_mask(formula, string_mask)
    if _BLOCKED_FUNCTION_RE.search(token_mask):
        raise FormulaSafetyError("external-capability function")
    if _LINK_FUNCTION_RE.search(token_mask):
        # A dynamic IMAGE/HYPERLINK target may be assembled from cells or
        # concatenated fragments. There is no sound formula-only proof that it
        # is internal, so generated uses are rejected in full.
        raise FormulaSafetyError("hyperlink or image function")
    if "|" in token_mask:
        raise FormulaSafetyError("DDE link")
    if _EXTERNAL_WORKBOOK_RE.search(string_mask):
        raise FormulaSafetyError("external workbook link")

    # File and UNC references are dangerous even inside a string because they
    # can be consumed by an otherwise innocuous-looking formula expression.
    if _FILE_URI_RE.search(formula) or _UNC_RE.search(formula):
        raise FormulaSafetyError("file or UNC reference")

    for match in _INDIRECT_RE.finditer(token_mask):
        opening = token_mask.find("(", match.start(), match.end())
        arguments = _function_arguments(formula, token_mask, opening)
        if _contains_external_workbook_reference(arguments):
            raise FormulaSafetyError("external workbook link")


def _formula_cell(value: object) -> FormulaCell | None:
    if isinstance(value, ArrayFormula):
        text = value.text
        if isinstance(text, str) and text:
            return (f"array:{value.ref}", text)
        return None
    if isinstance(value, DataTableFormula):
        attributes = (
            value.ref,
            value.r1,
            value.r2,
            value.ca,
            value.dt2D,
            value.dtr,
            value.del1,
            value.del2,
        )
        return ("data-table", repr(attributes))
    if isinstance(value, str) and value.startswith("="):
        return ("formula", value)
    return None


def _validate_cell_formula_safety(value: object, sheetnames: Sequence[str]) -> None:
    if isinstance(value, ArrayFormula):
        text = value.text
        if not isinstance(text, str) or not text:
            return
        validate_formula_integrity(_formula_expression(text), sheetnames=sheetnames)
        if not text.startswith("="):
            # openpyxl serialises array formula text as ``text[1:]``. Check the
            # actual expression that a malformed prefix would place in XML too.
            validate_formula_integrity(f"={text[1:]}", sheetnames=sheetnames)
        return
    if isinstance(value, DataTableFormula):
        for reference in (value.r1, value.r2):
            if isinstance(reference, str) and reference.strip():
                validate_formula_integrity(
                    _formula_expression(reference),
                    sheetnames=sheetnames,
                )
        return
    if isinstance(value, str):
        validate_formula_integrity(value, sheetnames=sheetnames)


def snapshot_formula_texts(workbook: object) -> FormulaSnapshot:
    """Fingerprint materialised formula cells without expanding sheet bounds.

    The stable worksheet-title/coordinate location is part of the fingerprint.
    Moving or copying legacy external formula text therefore creates a formula
    at a new location which must pass current safety checks.
    """

    snapshot: FormulaSnapshot = {}
    for worksheet in getattr(workbook, "worksheets", ()):
        for cell in worksheet._cells.values():
            formula_cell = _formula_cell(cell.value)
            if formula_cell is not None:
                snapshot[(worksheet.title, cell.coordinate)] = formula_cell
    return snapshot


def validate_changed_formula_safety(
    workbook: object,
    before: Mapping[tuple[str, str], FormulaCell],
) -> int:
    """Validate only formulae added or changed since ``before`` was captured.

    Returning the number checked makes the child-worker evidence auditable.
    Formulae which pre-dated the transform remain usable when their type, text,
    worksheet and coordinate are unchanged.
    """

    checked = 0
    worksheets = tuple(getattr(workbook, "worksheets", ()))
    sheetnames = tuple(str(worksheet.title) for worksheet in worksheets)
    for worksheet in worksheets:
        for cell in worksheet._cells.values():
            formula_cell = _formula_cell(cell.value)
            if formula_cell is None:
                continue
            key = (worksheet.title, cell.coordinate)
            if before.get(key) == formula_cell:
                continue
            _validate_cell_formula_safety(cell.value, sheetnames)
            checked += 1
    return checked


def _metadata_formula_text(value: object) -> str | None:
    if isinstance(value, str):
        return value if value.strip() else None
    text = getattr(value, "attr_text", None)
    return text if isinstance(text, str) and text.strip() else None


def _defined_name_type(defined_name: object) -> str:
    try:
        return str(getattr(defined_name, "type", ""))
    except (AttributeError, IndexError, TypeError, ValueError):
        return "UNKNOWN"


def _defined_name_is_formula(name_type: str, text: str) -> bool:
    if text.lstrip().startswith("="):
        return True
    if name_type == "UNKNOWN":
        return True
    return name_type in {
        "FUNC",
        "OPERATOR-INFIX",
        "OPERATOR-POSTFIX",
        "OPERATOR-PREFIX",
        "RANGE",
    }


def _formula_expression(text: str) -> str:
    stripped = text.lstrip()
    return stripped if stripped.startswith("=") else f"={stripped}"


def snapshot_formula_metadata(workbook: object) -> FormulaMetadataSnapshot:
    """Fingerprint formula-bearing names, tables, validation and formatting.

    These locations are mutable through openpyxl without writing ordinary cell
    formulae. Stable semantic locations ensure a moved, renamed or newly added
    metadata formula is validated rather than grandfathered by object identity.
    """

    snapshot: FormulaMetadataSnapshot = {}

    defined_name_scopes: list[tuple[tuple[object, ...], object]] = [
        (("workbook",), getattr(workbook, "defined_names", {}))
    ]
    for worksheet in getattr(workbook, "worksheets", ()):
        defined_name_scopes.append(
            (("worksheet", worksheet.title), getattr(worksheet, "defined_names", {}))
        )
    for scope, names in defined_name_scopes:
        for mapping_name, defined_name in names.items():
            text = _metadata_formula_text(getattr(defined_name, "attr_text", None))
            name_type = _defined_name_type(defined_name)
            executable_flags = any(
                bool(getattr(defined_name, attribute, None))
                for attribute in ("xlm", "function", "vbProcedure")
            )
            if text is None or not (executable_flags or _defined_name_is_formula(name_type, text)):
                continue
            actual_name = str(getattr(defined_name, "name", mapping_name))
            local_sheet_id = getattr(defined_name, "localSheetId", None)
            key = ("defined-name", *scope, str(mapping_name), actual_name, local_sheet_id)
            flags = (
                name_type,
                getattr(defined_name, "function", None),
                getattr(defined_name, "vbProcedure", None),
                getattr(defined_name, "xlm", None),
            )
            capability = "executable" if executable_flags else "ordinary"
            snapshot[key] = (f"defined-name:{capability}:{flags!r}", text)

    for worksheet in getattr(workbook, "worksheets", ()):
        for table in getattr(worksheet, "tables", {}).values():
            table_name = str(getattr(table, "displayName", None) or getattr(table, "name", ""))
            table_ref = str(getattr(table, "ref", ""))
            for column_index, column in enumerate(getattr(table, "tableColumns", ())):
                column_identity = (
                    column_index,
                    getattr(column, "id", None),
                    str(getattr(column, "name", "")),
                )
                for attribute in ("calculatedColumnFormula", "totalsRowFormula"):
                    formula_object = getattr(column, attribute, None)
                    text = _metadata_formula_text(formula_object)
                    if text is None:
                        continue
                    key = (
                        "table-formula",
                        worksheet.title,
                        table_name,
                        table_ref,
                        *column_identity,
                        attribute,
                    )
                    kind = f"table-formula:{attribute}:{getattr(formula_object, 'array', None)!r}"
                    snapshot[key] = (kind, text)

        validations = getattr(getattr(worksheet, "data_validations", None), "dataValidation", ())
        for validation_index, validation in enumerate(validations):
            location = str(getattr(validation, "sqref", ""))
            validation_type = getattr(validation, "type", None)
            operator = getattr(validation, "operator", None)
            for attribute in ("formula1", "formula2"):
                text = _metadata_formula_text(getattr(validation, attribute, None))
                if text is None:
                    continue
                key = (
                    "data-validation",
                    worksheet.title,
                    validation_index,
                    location,
                    attribute,
                )
                snapshot[key] = (f"data-validation:{validation_type}:{operator}", text)

        formatting = getattr(worksheet, "conditional_formatting", None)
        rules_by_range = getattr(formatting, "_cf_rules", {})
        for conditional, rules in rules_by_range.items():
            location = str(getattr(conditional, "sqref", conditional))
            for rule_index, rule in enumerate(rules):
                rule_type = getattr(rule, "type", None)
                operator = getattr(rule, "operator", None)
                for formula_index, value in enumerate(getattr(rule, "formula", ()) or ()):
                    text = _metadata_formula_text(value)
                    if text is None:
                        continue
                    key = (
                        "conditional-formatting",
                        worksheet.title,
                        location,
                        rule_index,
                        formula_index,
                    )
                    snapshot[key] = (f"conditional-formatting:{rule_type}:{operator}", text)

    return snapshot


def validate_changed_formula_metadata_safety(
    workbook: object,
    before: Mapping[MetadataLocation, FormulaCell],
) -> dict[str, int]:
    """Validate formula-bearing metadata added, moved or changed since capture."""

    counts = {
        "conditional_formatting": 0,
        "data_validations": 0,
        "defined_names": 0,
        "table_formulae": 0,
    }
    current = snapshot_formula_metadata(workbook)
    sheetnames = tuple(str(name) for name in getattr(workbook, "sheetnames", ()))
    for key, formula_cell in current.items():
        if before.get(key) == formula_cell:
            continue
        if formula_cell[0].startswith("defined-name:executable:"):
            raise FormulaSafetyError("executable defined name")
        validate_formula_integrity(
            _formula_expression(formula_cell[1]),
            sheetnames=sheetnames,
        )
        category = key[0]
        if category == "conditional-formatting":
            counts["conditional_formatting"] += 1
        elif category == "data-validation":
            counts["data_validations"] += 1
        elif category == "defined-name":
            counts["defined_names"] += 1
        elif category == "table-formula":
            counts["table_formulae"] += 1
    return counts


def snapshot_cell_hyperlinks(workbook: object) -> HyperlinkSnapshot:
    """Fingerprint hyperlink destinations at stable worksheet locations."""

    snapshot: HyperlinkSnapshot = {}
    for worksheet in getattr(workbook, "worksheets", ()):
        for cell in worksheet._cells.values():
            hyperlink = cell.hyperlink
            if hyperlink is None:
                continue
            target = getattr(hyperlink, "target", None)
            location = getattr(hyperlink, "location", None)
            snapshot[(worksheet.title, cell.coordinate)] = (
                target if isinstance(target, str) else None,
                location if isinstance(location, str) else None,
            )
    return snapshot


def _is_internal_hyperlink_destination(value: str, *, require_fragment: bool) -> bool:
    candidate = value.strip()
    if _FILE_URI_RE.search(candidate) or _UNC_RE.search(candidate):
        return False
    if require_fragment:
        if not candidate.startswith("#"):
            return False
        candidate = candidate[1:]
    elif candidate.startswith("#"):
        candidate = candidate[1:]
    return bool(candidate and _INTERNAL_HYPERLINK_RE.fullmatch(candidate))


def validate_changed_cell_hyperlinks(
    workbook: object,
    before: Mapping[tuple[str, str], tuple[str | None, str | None]],
) -> int:
    """Reject newly introduced external cell hyperlink relationships."""

    checked = 0
    current = snapshot_cell_hyperlinks(workbook)
    for key, (target, location) in current.items():
        if before.get(key) == (target, location):
            continue
        if (
            target is not None
            and target.strip()
            and not _is_internal_hyperlink_destination(target, require_fragment=True)
        ):
            raise FormulaSafetyError("external cell hyperlink")
        if (
            location is not None
            and location.strip()
            and not _is_internal_hyperlink_destination(location, require_fragment=False)
        ):
            raise FormulaSafetyError("external cell hyperlink")
        checked += 1
    return checked
