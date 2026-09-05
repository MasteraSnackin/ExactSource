"""Strict SolvePlan parsing and deterministic declarative plan application."""

from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import tempfile
from pathlib import Path
from typing import NoReturn

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from pydantic import ValidationError

from exactsource.contracts import (
    MAX_PLAN_OPERATIONS,
    ClearRange,
    CopyRange,
    FillArrayFormula,
    FillFormula,
    SetArrayFormula,
    SetFormula,
    SetValue,
    SolvePlan,
    TaskSpec,
)
from exactsource.formula_safety import FormulaSafetyError, validate_formula_integrity
from exactsource.ranges import RangeSyntaxError, range_bounds

MAX_EXCEL_ROW = 1_048_576
MAX_EXCEL_COLUMN = 16_384
MAX_OPERATION_CELLS = 250_000
MAX_PLAN_CELL_WRITES = 500_000
_CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")
_RANGE_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)(?::\$?([A-Za-z]{1,3})\$?([1-9][0-9]*))?$"
)
_FENCED_JSON_RE = re.compile(r"^```(?:json)?\s*(\{.*\})\s*```$", re.IGNORECASE | re.DOTALL)
_MODERN_FUNCTION_PREFIXES = {
    "XLOOKUP": "_xlfn.XLOOKUP",
    "UNIQUE": "_xlfn.UNIQUE",
    "LET": "_xlfn.LET",
    "CHOOSECOLS": "_xlfn.CHOOSECOLS",
    "FILTER": "_xlfn._xlws.FILTER",
    "AGGREGATE": "_xlfn.AGGREGATE",
    "IFNA": "_xlfn.IFNA",
    "MAXIFS": "_xlfn.MAXIFS",
    "MINIFS": "_xlfn.MINIFS",
    "TEXTJOIN": "_xlfn.TEXTJOIN",
    "CONCAT": "_xlfn.CONCAT",
    "FILTERXML": "_xlfn.FILTERXML",
    "TEXTSPLIT": "_xlfn.TEXTSPLIT",
    "DROP": "_xlfn.DROP",
    "SORT": "_xlfn._xlws.SORT",
    "SORTBY": "_xlfn.SORTBY",
}
_MODERN_FUNCTION_RE = re.compile(
    r"(?<![A-Za-z0-9_.])(?:"
    + "|".join(re.escape(name) for name in sorted(_MODERN_FUNCTION_PREFIXES, key=len, reverse=True))
    + r")(?=\s*\()",
    flags=re.IGNORECASE,
)


class PlanError(RuntimeError):
    """Base class for plan parsing and application failures."""


class PlanParseError(PlanError):
    """Raised when model output is not exactly one valid SolvePlan object."""


class PlanApplicationError(PlanError):
    """Raised before or during a deterministic workbook operation."""


def _reject_json_constant(value: str) -> NoReturn:
    raise ValueError(f"non-standard JSON constant {value!r} is not allowed")


def parse_plan(text: str) -> SolvePlan:
    """Parse plain or singly fenced JSON and validate it as a strict SolvePlan.

    Leading explanations, trailing commentary, multiple objects, JSON5 and non-finite
    numbers are deliberately rejected so a plausible-looking response cannot be
    silently interpreted as a different workbook mutation.
    """

    if not isinstance(text, str) or not text.strip():
        raise PlanParseError("model response is empty")
    candidate = text.strip()
    fenced = _FENCED_JSON_RE.fullmatch(candidate)
    if fenced:
        candidate = fenced.group(1).strip()
    try:
        decoded = json.loads(candidate, parse_constant=_reject_json_constant)
    except (json.JSONDecodeError, ValueError) as exc:
        raise PlanParseError(f"model response is not strict JSON: {exc}") from None
    if not isinstance(decoded, dict):
        raise PlanParseError("model response must be one JSON object")
    try:
        return SolvePlan.model_validate(decoded)
    except ValidationError as exc:
        raise PlanParseError(f"model response does not match SolvePlan: {exc}") from None


def _cell_coordinate(value: str, *, field: str) -> tuple[str, int, int]:
    if not isinstance(value, str):
        raise PlanApplicationError(f"{field} must be a cell address string")
    match = _CELL_RE.fullmatch(value.strip())
    if not match:
        raise PlanApplicationError(f"{field} is not one A1 cell address: {value!r}")
    column = column_index_from_string(match.group(1).upper())
    row = int(match.group(2))
    if column > MAX_EXCEL_COLUMN or row > MAX_EXCEL_ROW:
        raise PlanApplicationError(f"{field} is outside Excel worksheet limits: {value!r}")
    return f"{get_column_letter(column)}{row}", row, column


def _range_coordinates(value: str, *, field: str) -> tuple[str, int, int, int, int, int]:
    if not isinstance(value, str):
        raise PlanApplicationError(f"{field} must be an A1 range string")
    match = _RANGE_RE.fullmatch(value.strip())
    if not match:
        raise PlanApplicationError(f"{field} is not one rectangular A1 range: {value!r}")
    min_col = column_index_from_string(match.group(1).upper())
    min_row = int(match.group(2))
    max_col = column_index_from_string((match.group(3) or match.group(1)).upper())
    max_row = int(match.group(4) or match.group(2))
    if min_col > max_col or min_row > max_row:
        raise PlanApplicationError(f"{field} must run from top-left to bottom-right")
    if max_col > MAX_EXCEL_COLUMN or max_row > MAX_EXCEL_ROW:
        raise PlanApplicationError(f"{field} is outside Excel worksheet limits: {value!r}")
    count = (max_row - min_row + 1) * (max_col - min_col + 1)
    if count > MAX_OPERATION_CELLS:
        raise PlanApplicationError(
            f"{field} contains {count:,} cells; limit is {MAX_OPERATION_CELLS:,}"
        )
    normalised = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return normalised, min_row, min_col, max_row, max_col, count


def _sheet(workbook: object, name: str, *, field: str):
    if not isinstance(name, str) or not name:
        raise PlanApplicationError(f"{field} must be a non-empty worksheet name")
    sheetnames = getattr(workbook, "sheetnames", ())
    if name not in sheetnames:
        raise PlanApplicationError(f"{field} names missing worksheet {name!r}")
    return workbook[name]  # type: ignore[index]


def _canonicalise_modern_functions(value: str) -> str:
    """Rewrite supported bare function calls outside Excel string literals.

    Stored xlsx formula names are intentionally canonical rather than preserving
    input case.  The negative look-behind in ``_MODERN_FUNCTION_RE`` excludes
    already-qualified names, while this scanner prevents formula-looking text in
    double-quoted strings, quoted sheet names and bracketed references from being
    edited. Doubled quote characters inside either kind of quoted token are escaped.
    """

    def replace(match: re.Match[str]) -> str:
        return _MODERN_FUNCTION_PREFIXES[match.group(0).upper()]

    pieces: list[str] = []
    outside_start = 0
    index = 0
    while index < len(value):
        delimiter = value[index]
        if delimiter in {'"', "'"}:
            pieces.append(_MODERN_FUNCTION_RE.sub(replace, value[outside_start:index]))
            token_start = index
            index += 1
            while index < len(value):
                if value[index] != delimiter:
                    index += 1
                    continue
                if index + 1 < len(value) and value[index + 1] == delimiter:
                    index += 2
                    continue
                index += 1
                break
            pieces.append(value[token_start:index])
            outside_start = index
            continue
        if delimiter == "[":
            pieces.append(_MODERN_FUNCTION_RE.sub(replace, value[outside_start:index]))
            reference_start = index
            depth = 1
            index += 1
            while index < len(value) and depth:
                if value[index] == "[":
                    depth += 1
                elif value[index] == "]":
                    depth -= 1
                index += 1
            pieces.append(value[reference_start:index])
            outside_start = index
            continue
        index += 1

    pieces.append(_MODERN_FUNCTION_RE.sub(replace, value[outside_start:]))
    return "".join(pieces)


def _formula(value: str, *, field: str, workbook: object | None = None) -> str:
    if not isinstance(value, str) or not value.startswith("=") or len(value) < 2:
        raise PlanApplicationError(f"{field} must be a non-empty Excel formula beginning with '='")
    if "\x00" in value:
        raise PlanApplicationError(f"{field} contains a null byte")
    if len(value) > 8_192:
        raise PlanApplicationError(f"{field} exceeds Excel's 8,192-character formula limit")
    canonicalised = _canonicalise_modern_functions(value)
    if len(canonicalised) > 8_192:
        raise PlanApplicationError(
            f"{field} exceeds Excel's 8,192-character formula limit after canonicalisation"
        )
    try:
        sheetnames = getattr(workbook, "sheetnames", None) if workbook is not None else None
        validate_formula_integrity(canonicalised, sheetnames=sheetnames)
    except FormulaSafetyError as exc:
        raise PlanApplicationError(f"{field} {exc}") from None
    return canonicalised


def _save_atomic(workbook: object, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_symlink():
        raise PlanApplicationError("destination workbook must not be a symbolic link")
    with tempfile.NamedTemporaryFile(
        prefix=f".{destination.stem}-",
        suffix=".xlsx",
        dir=destination.parent,
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)  # type: ignore[attr-defined]
        os.replace(temporary, destination)
        os.chmod(destination, 0o644)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def _write_literal(cell: object, value: object) -> None:
    cell.value = value  # type: ignore[attr-defined]
    if isinstance(value, str) and value.startswith("="):
        # A set_value operation is explicitly literal, even when its text resembles
        # a formula. Formula writes must use set_formula.
        cell.data_type = "s"  # type: ignore[attr-defined]


def _apply_copy(workbook: object, operation: CopyRange) -> tuple[int, set[str]]:
    source_sheet = _sheet(workbook, operation.source_sheet, field="source_sheet")
    destination_sheet = _sheet(workbook, operation.destination_sheet, field="destination_sheet")
    _, min_row, min_col, max_row, max_col, count = _range_coordinates(
        operation.source_range, field="source_range"
    )
    destination, destination_row, destination_col = _cell_coordinate(
        operation.destination_cell, field="destination_cell"
    )
    del destination
    row_count = max_row - min_row + 1
    column_count = max_col - min_col + 1
    if destination_row + row_count - 1 > MAX_EXCEL_ROW:
        raise PlanApplicationError("copy_range destination exceeds the last Excel row")
    if destination_col + column_count - 1 > MAX_EXCEL_COLUMN:
        raise PlanApplicationError("copy_range destination exceeds the last Excel column")

    # An OOXML array formula is one anchored object with a rectangular ref. Copying
    # only part of that rectangle would silently discard or corrupt its metadata.
    # Require the complete ref, then translate both its formula text and ref exactly
    # as the containing source rectangle moves.
    array_refs: dict[str, tuple[str, int, int, int, int]] = {}
    for anchor, array_ref in source_sheet.array_formulae.items():
        normalised_ref, array_min_row, array_min_col, array_max_row, array_max_col, _ = (
            _range_coordinates(array_ref, field=f"array formula {anchor} ref")
        )
        intersects = not (
            array_max_row < min_row
            or array_min_row > max_row
            or array_max_col < min_col
            or array_min_col > max_col
        )
        if not intersects:
            continue
        contained = (
            min_row <= array_min_row <= array_max_row <= max_row
            and min_col <= array_min_col <= array_max_col <= max_col
        )
        if not contained:
            raise PlanApplicationError(
                f"copy_range cannot copy only part of array formula {anchor}:{normalised_ref}; "
                "copy its complete ref or use set_array_formula/fill_array_formula"
            )
        array_refs[anchor] = (
            normalised_ref,
            array_min_row,
            array_min_col,
            array_max_row,
            array_max_col,
        )

    # Snapshot first so overlapping copies behave like Excel copy/paste rather than a
    # cascading series of assignments.
    snapshot: list[tuple[int, int, str, object, object | None]] = []
    for row_offset, source_row in enumerate(range(min_row, max_row + 1)):
        for column_offset, source_col in enumerate(range(min_col, max_col + 1)):
            source_cell = source_sheet.cell(source_row, source_col)
            style = copy.copy(source_cell._style) if operation.include_style else None
            snapshot.append(
                (row_offset, column_offset, source_cell.coordinate, source_cell.value, style)
            )

    for row_offset, column_offset, source_coordinate, value, style in snapshot:
        target = destination_sheet.cell(
            destination_row + row_offset, destination_col + column_offset
        )
        if isinstance(value, str) and value.startswith("="):
            try:
                value = _formula(
                    Translator(value, origin=source_coordinate).translate_formula(
                        target.coordinate
                    ),
                    field="copied formula",
                    workbook=workbook,
                )
            except PlanApplicationError:
                raise
            except Exception:
                # openpyxl's TokenizerError embeds the complete formula, which
                # may contain confidential workbook text. Keep this error fixed
                # and categorical so traces never echo the formula.
                raise PlanApplicationError("copied formula could not be translated") from None
        elif isinstance(value, ArrayFormula):
            array = array_refs.get(source_coordinate)
            if array is None:
                raise PlanApplicationError(
                    f"array formula anchor {source_coordinate} has no complete copied ref"
                )
            _, array_min_row, array_min_col, array_max_row, array_max_col = array
            _, source_anchor_row, source_anchor_column = _cell_coordinate(
                source_coordinate,
                field="array formula anchor",
            )
            row_delta = target.row - source_anchor_row
            column_delta = target.column - source_anchor_column
            translated_ref = (
                f"{get_column_letter(array_min_col + column_delta)}"
                f"{array_min_row + row_delta}:"
                f"{get_column_letter(array_max_col + column_delta)}"
                f"{array_max_row + row_delta}"
            )
            try:
                translated_text = Translator(
                    _formula(value.text, field="array formula text", workbook=workbook),
                    origin=source_coordinate,
                ).translate_formula(target.coordinate)
            except PlanApplicationError:
                raise
            except Exception:
                raise PlanApplicationError("copied array formula could not be translated") from None
            value = ArrayFormula(ref=translated_ref, text=translated_text)
        elif isinstance(value, DataTableFormula):
            # What-if data-table formulae carry anchor/ref metadata whose safe
            # relocation is not implemented here. Reusing the source object can
            # retain a stale ref and also bypass validation of r1/r2, so copying
            # one must fail closed instead of producing a corrupted workbook.
            raise PlanApplicationError("copy_range does not support data-table formulae")
        target.value = value
        if style is not None:
            target._style = copy.copy(style)
    return count, {operation.destination_sheet}


def _destination_bounds(operation: object) -> tuple[str, int, int, int, int]:
    if isinstance(operation, (SetValue, SetFormula, SetArrayFormula)):
        _, row, column = _cell_coordinate(operation.cell, field="cell")
        return operation.sheet, row, column, row, column
    if isinstance(operation, (FillFormula, FillArrayFormula, ClearRange)):
        _, min_row, min_col, max_row, max_col, _ = _range_coordinates(
            operation.range, field="range"
        )
        return operation.sheet, min_row, min_col, max_row, max_col
    if isinstance(operation, CopyRange):
        _, source_min_row, source_min_col, source_max_row, source_max_col, _ = _range_coordinates(
            operation.source_range, field="source_range"
        )
        _, destination_row, destination_col = _cell_coordinate(
            operation.destination_cell, field="destination_cell"
        )
        max_row = destination_row + (source_max_row - source_min_row)
        max_col = destination_col + (source_max_col - source_min_col)
        if max_row > MAX_EXCEL_ROW or max_col > MAX_EXCEL_COLUMN:
            raise PlanApplicationError("copy_range destination exceeds Excel worksheet limits")
        return (
            operation.destination_sheet,
            destination_row,
            destination_col,
            max_row,
            max_col,
        )
    raise PlanApplicationError(f"unsupported operation type {type(operation).__name__}")


def _answer_bounds(workbook: object, task: TaskSpec) -> dict[str, list[tuple[int, int, int, int]]]:
    allowed: dict[str, list[tuple[int, int, int, int]]] = {}
    for answer in task.answer_ranges:
        _sheet(workbook, answer.sheet, field="answer range sheet")
        try:
            min_col, min_row, max_col, max_row = range_bounds(
                answer.cells,
                # Whole-column and whole-row answer declarations express the
                # evaluator's allowed destination, not merely the input sheet's
                # currently materialised extent. A result tab may begin empty and
                # legitimately grow within that declared Excel-grid range.
                max_row=MAX_EXCEL_ROW,
                max_column=MAX_EXCEL_COLUMN,
            )
        except RangeSyntaxError as exc:
            raise PlanApplicationError(
                f"task declares invalid answer range {answer.sheet!r}!{answer.cells}: {exc}"
            ) from None
        allowed.setdefault(answer.sheet, []).append((min_row, min_col, max_row, max_col))
    if not allowed:
        raise PlanApplicationError("task has no declared answer ranges")
    return allowed


def _rectangle_is_covered(
    destination: tuple[int, int, int, int],
    allowed: list[tuple[int, int, int, int]],
) -> bool:
    min_row, min_col, max_row, max_col = destination
    for row in range(min_row, max_row + 1):
        intervals = sorted(
            (allowed_min_col, allowed_max_col)
            for allowed_min_row, allowed_min_col, allowed_max_row, allowed_max_col in allowed
            if allowed_min_row <= row <= allowed_max_row
            and allowed_max_col >= min_col
            and allowed_min_col <= max_col
        )
        cursor = min_col
        for interval_start, interval_end in intervals:
            if interval_end < cursor:
                continue
            if interval_start > cursor:
                break
            cursor = max(cursor, interval_end + 1)
            if cursor > max_col:
                break
        if cursor <= max_col:
            return False
    return True


def _validate_operation_scope(workbook: object, task: TaskSpec, plan: SolvePlan) -> None:
    allowed_by_sheet = _answer_bounds(workbook, task)
    for index, operation in enumerate(plan.operations):
        sheet, min_row, min_col, max_row, max_col = _destination_bounds(operation)
        allowed = allowed_by_sheet.get(sheet)
        if not allowed:
            raise PlanApplicationError(
                f"operation {index} destination worksheet {sheet!r} has no declared answer range"
            )
        if not _rectangle_is_covered((min_row, min_col, max_row, max_col), allowed):
            start = f"{get_column_letter(min_col)}{min_row}"
            end = f"{get_column_letter(max_col)}{max_row}"
            destination = start if start == end else f"{start}:{end}"
            raise PlanApplicationError(
                f"operation {index} destination {sheet!r}!{destination} "
                "falls outside declared answer ranges"
            )


def _validate_plan_resource_limits(plan: SolvePlan) -> None:
    """Reject oversized declarative plans before workbook loading or mutation.

    The operation-count check is intentionally repeated here even though Pydantic
    enforces it at the contract boundary. Callers can bypass model validation with
    ``model_construct``, and application safety must not depend on how a plan was
    instantiated.
    """

    operation_count = len(plan.operations)
    if operation_count > MAX_PLAN_OPERATIONS:
        raise PlanApplicationError(
            f"plan contains {operation_count:,} operations; limit is {MAX_PLAN_OPERATIONS:,}"
        )

    total_writes = 0
    for operation in plan.operations:
        _, min_row, min_col, max_row, max_col = _destination_bounds(operation)
        total_writes += (max_row - min_row + 1) * (max_col - min_col + 1)
        if total_writes > MAX_PLAN_CELL_WRITES:
            raise PlanApplicationError(
                f"plan writes {total_writes:,} cells; limit is {MAX_PLAN_CELL_WRITES:,}"
            )


def _apply_one(workbook: object, operation: object) -> tuple[int, set[str], dict[str, str] | None]:
    if isinstance(operation, SetValue):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        coordinate, _, _ = _cell_coordinate(operation.cell, field="cell")
        _write_literal(sheet[coordinate], operation.value)
        return 1, {operation.sheet}, None

    if isinstance(operation, SetFormula):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        coordinate, _, _ = _cell_coordinate(operation.cell, field="cell")
        sheet[coordinate] = _formula(operation.formula, field="formula", workbook=workbook)
        return 1, {operation.sheet}, None

    if isinstance(operation, FillFormula):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        _, min_row, min_col, max_row, max_col, count = _range_coordinates(
            operation.range, field="range"
        )
        origin = f"{get_column_letter(min_col)}{min_row}"
        formula = _formula(operation.formula, field="formula", workbook=workbook)
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                target = sheet.cell(row, column)
                try:
                    target.value = Translator(formula, origin=origin).translate_formula(
                        target.coordinate
                    )
                except Exception:
                    # TokenizerError includes the complete formula. Do not place
                    # model-generated literals or workbook text in traces.
                    raise PlanApplicationError("fill formula could not be translated") from None
        return count, {operation.sheet}, None

    if isinstance(operation, SetArrayFormula):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        coordinate, _, _ = _cell_coordinate(operation.cell, field="cell")
        formula = _formula(operation.formula, field="formula", workbook=workbook)
        sheet[coordinate] = ArrayFormula(ref=coordinate, text=formula)
        return (
            1,
            {operation.sheet},
            {
                "sheet": operation.sheet,
                "anchor": coordinate,
                "ref": coordinate,
                "text": formula,
            },
        )

    if isinstance(operation, FillArrayFormula):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        normalised, min_row, min_col, max_row, max_col, count = _range_coordinates(
            operation.range, field="range"
        )
        formula = _formula(operation.formula, field="formula", workbook=workbook)
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                sheet.cell(row, column).value = None
        anchor = f"{get_column_letter(min_col)}{min_row}"
        sheet[anchor] = ArrayFormula(ref=normalised, text=formula)
        return (
            count,
            {operation.sheet},
            {
                "sheet": operation.sheet,
                "anchor": anchor,
                "ref": normalised,
                "text": formula,
            },
        )

    if isinstance(operation, ClearRange):
        sheet = _sheet(workbook, operation.sheet, field="sheet")
        _, min_row, min_col, max_row, max_col, count = _range_coordinates(
            operation.range, field="range"
        )
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                sheet.cell(row, column).value = None
        return count, {operation.sheet}, None

    if isinstance(operation, CopyRange):
        count, sheets = _apply_copy(workbook, operation)
        return count, sheets, None

    raise PlanApplicationError(f"unsupported operation type {type(operation).__name__}")


def apply_operations(
    plan: SolvePlan,
    task: TaskSpec,
    source_path: str | Path,
    destination_path: str | Path,
) -> dict[str, object]:
    """Apply an operations plan to a workbook and atomically write its result."""

    if plan.route != "operations":
        raise PlanApplicationError("apply_operations requires an operations route")
    _validate_plan_resource_limits(plan)
    source = Path(source_path)
    destination = Path(destination_path)
    if not source.is_file():
        raise PlanApplicationError(f"source workbook does not exist: {source}")
    if source.resolve() == destination.resolve():
        raise PlanApplicationError("source and destination workbooks must be different paths")

    try:
        workbook = load_workbook(source, data_only=False, keep_links=True)
    except Exception as exc:
        raise PlanApplicationError(f"could not load source workbook: {exc}") from None

    writes = 0
    touched_sheets: set[str] = set()
    array_formulas: list[dict[str, str]] = []
    try:
        _validate_operation_scope(workbook, task, plan)
        for index, operation in enumerate(plan.operations):
            try:
                operation_writes, sheets, array_formula = _apply_one(workbook, operation)
            except PlanApplicationError as exc:
                raise PlanApplicationError(f"operation {index} failed: {exc}") from None
            except Exception as exc:
                raise PlanApplicationError(
                    f"operation {index} failed while writing the workbook: {exc}"
                ) from None
            writes += operation_writes
            touched_sheets.update(sheets)
            if array_formula is not None:
                array_formulas.append(array_formula)

        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
            calculation.calcMode = "auto"
        _save_atomic(workbook, destination)
    finally:
        workbook.close()

    digest = hashlib.sha256(destination.read_bytes()).hexdigest()
    return {
        "task_id": task.id,
        "route": "operations",
        "operations_applied": len(plan.operations),
        "cell_writes": writes,
        "touched_sheets": sorted(touched_sheets),
        "array_formulas": array_formulas,
        "output_bytes": destination.stat().st_size,
        "output_sha256": digest,
    }
