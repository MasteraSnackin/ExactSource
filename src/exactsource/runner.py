"""Failure-isolated task execution for the SpreadsheetBench judge contract."""

from __future__ import annotations

import inspect
import json
import os
import tempfile
import time
from collections.abc import Callable, Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from exactsource.artifacts import (
    ArtifactError,
    OutputLayout,
    Prediction,
    TraceRecorder,
    atomic_copy_workbook,
    prepare_output,
    promote_workbook,
    safe_task_id,
    temporary_workbook_path,
    validate_run,
    write_predictions,
    write_trace,
)
from exactsource.config import (
    API_KEY_ENV,
    CELL_MAX_OUTPUT_TOKENS,
    CELL_TRUNCATION_RECOVERY_REASONING_EFFORT,
    CONCURRENCY,
    MODEL_NAME,
    REASONING_EFFORT,
    SEMANTIC_REPAIRS,
    SHEET_MAX_OUTPUT_TOKENS,
)
from exactsource.contracts import ContextPack, ModelReply, SolvePlan, SolveResult, TaskSpec
from exactsource.metrics import write_run_metrics
from exactsource.model import ModelTruncationError


class TaskSolver(Protocol):
    """Injected single-task solver used by the runtime and deterministic tests."""

    def __call__(
        self,
        task: TaskSpec,
        working_xlsx: Path,
        trace: TraceRecorder,
    ) -> SolveResult | None: ...


LogSink = Callable[[str], None]


@dataclass(frozen=True, slots=True)
class RunSummary:
    total: int
    succeeded: int
    failed: int
    predictions_path: Path
    run_metrics_path: Path


@dataclass(frozen=True, slots=True)
class _TaskRun:
    prediction: Prediction
    latency_ms: int


class TaskSolveError(RuntimeError):
    """Turn a solver-declared failure into the ordinary fallback path."""


class ModelClient(Protocol):
    def complete(
        self,
        messages: list[dict[str, str]],
        *,
        max_output_tokens: int,
        reasoning_effort: bool,
        on_attempt: Callable[[dict[str, object]], None] | None = None,
    ) -> ModelReply: ...

    def close(self) -> None: ...


ContextBuilder = Callable[[TaskSpec], ContextPack]
MessageBuilder = Callable[[TaskSpec, ContextPack], list[dict[str, str]]]
PlanParser = Callable[[str], SolvePlan]
OperationApplier = Callable[[SolvePlan, TaskSpec, Path, Path], dict[str, object]]
TransformRunner = Callable[[str, Path, Path], dict[str, object]]


@dataclass(frozen=True, slots=True)
class _GenerationPolicy:
    name: str
    max_output_tokens: int
    reasoning_effort: bool
    recovery_reason: str | None = None


_CELL_POLICY = _GenerationPolicy(
    name="cell_reasoning",
    max_output_tokens=CELL_MAX_OUTPUT_TOKENS,
    reasoning_effort=REASONING_EFFORT,
)
_SHEET_POLICY = _GenerationPolicy(
    name="sheet_reasoning",
    max_output_tokens=SHEET_MAX_OUTPUT_TOKENS,
    reasoning_effort=REASONING_EFFORT,
)
_CELL_TRUNCATION_RECOVERY_POLICY = _GenerationPolicy(
    name="cell_max_tokens_no_think_recovery",
    max_output_tokens=CELL_MAX_OUTPUT_TOKENS,
    reasoning_effort=CELL_TRUNCATION_RECOVERY_REASONING_EFFORT,
    recovery_reason="max_tokens",
)


class DefaultTaskSolver:
    """Compose the fixed model, plan parser and two workbook execution routes."""

    def __init__(
        self,
        client: ModelClient | None = None,
        *,
        context_builder: ContextBuilder | None = None,
        message_builder: MessageBuilder | None = None,
        plan_parser: PlanParser | None = None,
        operation_applier: OperationApplier | None = None,
        transform_runner: TransformRunner | None = None,
        semantic_repairs: int = SEMANTIC_REPAIRS,
    ) -> None:
        # Lazy imports keep runner tests independent from network configuration and
        # make each stage directly replaceable in deterministic experiments.
        from exactsource.context import build_context
        from exactsource.model import TinkerClient
        from exactsource.plans import apply_operations, parse_plan
        from exactsource.prompts import build_messages
        from exactsource.sandbox import run_transform

        if semantic_repairs < 0:
            raise ValueError("semantic_repairs must not be negative")
        self.client = client or TinkerClient()
        self._owns_client = client is None
        self.context_builder = context_builder or build_context
        self.message_builder = message_builder or build_messages
        self.plan_parser = plan_parser or parse_plan
        self.operation_applier = operation_applier or apply_operations
        self.transform_runner = transform_runner or run_transform
        self.semantic_repairs = semantic_repairs

    def close(self) -> None:
        if self._owns_client:
            self.client.close()

    def __call__(
        self,
        task: TaskSpec,
        working_xlsx: Path,
        trace: TraceRecorder,
    ) -> SolveResult:
        context_started = time.monotonic()
        context = self.context_builder(task)
        context_build_latency_ms = _elapsed_ms(context_started)
        context_evidence = {
            "original_chars": context.original_chars,
            "emitted_chars": len(context.text),
            "truncated": context.truncated,
            "sha256": context.sha256,
        }
        messages_started = time.monotonic()
        messages = self.message_builder(task, context)
        message_build_latency_ms = _elapsed_ms(messages_started)
        prior_reply: str | None = None
        last_error: Exception | None = None
        default_policy = _CELL_POLICY if task.is_cell_level else _SHEET_POLICY
        next_call_is_truncation_recovery = False

        for semantic_attempt in range(self.semantic_repairs + 1):
            is_truncation_recovery = next_call_is_truncation_recovery
            if is_truncation_recovery:
                request_messages = _cell_truncation_recovery_messages(messages)
                policy = _CELL_TRUNCATION_RECOVERY_POLICY
            else:
                request_messages = (
                    messages
                    if semantic_attempt == 0
                    else _repair_messages(messages, prior_reply or "", last_error)
                )
                policy = default_policy
            prompt_text = json.dumps(request_messages, ensure_ascii=False, separators=(",", ":"))
            message_chars = sum(len(message["content"]) for message in request_messages)
            record_count_before_call = len(trace.records)
            logical_call_started = time.monotonic()

            def on_attempt(
                event: dict[str, object],
                *,
                semantic_attempt_number: int = semantic_attempt + 1,
                attempt_prompt: str = prompt_text,
                attempt_message_chars: int = message_chars,
                attempt_policy: _GenerationPolicy = policy,
            ) -> None:
                record = dict(event)
                record.setdefault("provider_status", record.get("status"))
                record.setdefault("message_chars", attempt_message_chars)
                record.update(
                    {
                        "phase": "model",
                        "event": "provider_attempt",
                        "model": MODEL_NAME,
                        "semantic_attempt": semantic_attempt_number,
                        "semantic_repair": semantic_attempt_number > 1,
                        "prompt": attempt_prompt,
                        "context": context_evidence,
                        "generation_policy": attempt_policy.name,
                        "max_output_tokens": attempt_policy.max_output_tokens,
                        "reasoning_effort": attempt_policy.reasoning_effort,
                    }
                )
                if attempt_policy.recovery_reason is not None:
                    record["recovery_reason"] = attempt_policy.recovery_reason
                if semantic_attempt_number == 1 and record.get("attempt") == 1:
                    record["context_build_latency_ms"] = context_build_latency_ms
                    record["message_build_latency_ms"] = message_build_latency_ms
                trace.record(record)

            try:
                reply = self.client.complete(
                    request_messages,
                    max_output_tokens=policy.max_output_tokens,
                    reasoning_effort=policy.reasoning_effort,
                    on_attempt=on_attempt,
                )
            except ModelTruncationError:
                if len(trace.records) > record_count_before_call:
                    trace.update_last(
                        logical_call_terminal=True,
                        logical_call_latency_ms=_elapsed_ms(logical_call_started),
                        plan_status="not_reached",
                    )
                can_recover = (
                    task.is_cell_level
                    and semantic_attempt == 0
                    and self.semantic_repairs >= 1
                    and not is_truncation_recovery
                )
                if can_recover:
                    next_call_is_truncation_recovery = True
                    continue
                raise
            except Exception:
                if len(trace.records) > record_count_before_call:
                    trace.update_last(
                        logical_call_terminal=True,
                        logical_call_latency_ms=_elapsed_ms(logical_call_started),
                        plan_status="not_reached",
                    )
                raise
            if len(trace.records) > record_count_before_call:
                trace.update_last(
                    logical_call_terminal=True,
                    logical_call_latency_ms=_non_negative_latency(
                        reply.latency_ms,
                        fallback_started=logical_call_started,
                    ),
                )
            prior_reply = reply.text

            parse_started = time.monotonic()
            try:
                plan = self.plan_parser(reply.text)
            except Exception as error:
                last_error = error
                trace.update_last(
                    event="model_call",
                    status="plan_rejected",
                    plan_status="parse_rejected",
                    semantic_attempt=semantic_attempt + 1,
                    tool="parse_and_apply_plan",
                    tool_input={"action": "parse_and_apply_returned_solve_plan"},
                    tool_output=None,
                    error=_safe_error(error),
                    plan_parse_latency_ms=_elapsed_ms(parse_started),
                )
                if is_truncation_recovery or semantic_attempt >= self.semantic_repairs:
                    raise
                continue

            plan_parse_latency_ms = _elapsed_ms(parse_started)
            apply_started = time.monotonic()
            try:
                evidence = self._apply_plan(plan, task, working_xlsx)
            except Exception as error:
                last_error = error
                trace.update_last(
                    event="model_call",
                    status="plan_rejected",
                    plan_status="apply_rejected",
                    semantic_attempt=semantic_attempt + 1,
                    tool="parse_and_apply_plan",
                    tool_input={"action": "parse_and_apply_returned_solve_plan"},
                    tool_output=None,
                    error=_safe_error(error),
                    plan_parse_latency_ms=plan_parse_latency_ms,
                    plan_apply_latency_ms=_elapsed_ms(apply_started),
                )
                if is_truncation_recovery or semantic_attempt >= self.semantic_repairs:
                    raise
                continue

            trace.update_last(
                event="model_call",
                plan_status="accepted",
                semantic_attempt=semantic_attempt + 1,
                tool="apply_operations" if plan.route == "operations" else "run_transform",
                tool_input=plan.model_dump(mode="json"),
                tool_output=evidence,
                error=None,
                plan_parse_latency_ms=plan_parse_latency_ms,
                plan_apply_latency_ms=_elapsed_ms(apply_started),
            )
            return SolveResult(status="ok", plan=plan)

        raise TaskSolveError(_safe_error(last_error or RuntimeError("plan repair exhausted")))

    def _apply_plan(
        self,
        plan: SolvePlan,
        task: TaskSpec,
        working_xlsx: Path,
    ) -> dict[str, object]:
        if plan.route == "python" and task.is_cell_level:
            raise TaskSolveError(
                "python route is not allowed for cell-level tasks; use declarative operations"
            )
        descriptor, candidate_name = tempfile.mkstemp(
            prefix=f".{task.id}.candidate.",
            suffix=".xlsx",
            dir=working_xlsx.parent,
        )
        os.close(descriptor)
        candidate = Path(candidate_name)
        try:
            if plan.route == "operations":
                evidence = self.operation_applier(plan, task, working_xlsx, candidate)
            else:
                assert plan.python_code is not None
                evidence = self.transform_runner(plan.python_code, working_xlsx, candidate)
            _validate_answer_sheets(task, candidate)
            promote_workbook(candidate, working_xlsx)
            return evidence
        finally:
            candidate.unlink(missing_ok=True)


def run_tasks(
    tasks: Sequence[TaskSpec],
    out_dir: Path,
    solver: TaskSolver,
    *,
    concurrency: int = CONCURRENCY,
    log: LogSink = print,
) -> RunSummary:
    """Run every task, preserving dataset order in ``predictions.jsonl``.

    Individual solver failures never stop the batch. Each failed task receives a
    readable copy of its init workbook, a prediction entry and a trace. Structural
    failures that make the judge contract impossible still fail the run loudly.
    """

    if concurrency < 1:
        raise ValueError("concurrency must be at least one")
    run_started = time.monotonic()
    task_list = list(tasks)
    _validate_tasks(task_list)
    layout = prepare_output(Path(out_dir))
    log(f"tasks={len(task_list)} concurrency={concurrency}")

    indexed_results: dict[int, _TaskRun] = {}
    if task_list:
        workers = min(concurrency, len(task_list))
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="exactsource") as pool:
            futures = {
                pool.submit(_run_one, task, layout, solver): index
                for index, task in enumerate(task_list)
            }
            for future in as_completed(futures):
                index = futures[future]
                task_run = future.result()
                indexed_results[index] = task_run
                log(f"{task_run.prediction.id:<12} {task_run.prediction.status}")

    task_runs = [indexed_results[index] for index in range(len(task_list))]
    predictions = [task_run.prediction for task_run in task_runs]
    write_predictions(layout, predictions)
    checked = validate_run(layout, (task.id for task in task_list))
    succeeded = sum(item.status == "ok" for item in checked)
    run_wall_time_ms = max(0, round((time.monotonic() - run_started) * 1_000))
    write_run_metrics(
        layout,
        checked,
        run_wall_time_ms=run_wall_time_ms,
        task_latencies_ms={task_run.prediction.id: task_run.latency_ms for task_run in task_runs},
    )
    summary = RunSummary(
        total=len(checked),
        succeeded=succeeded,
        failed=len(checked) - succeeded,
        predictions_path=layout.predictions_path,
        run_metrics_path=layout.run_metrics_path,
    )
    log(
        f"complete total={summary.total} ok={summary.succeeded} "
        f"failed={summary.failed} predictions={summary.predictions_path} "
        f"metrics={summary.run_metrics_path}"
    )
    return summary


def _run_one(
    task: TaskSpec,
    layout: OutputLayout,
    solver: TaskSolver,
) -> _TaskRun:
    recorder = TraceRecorder(task.id)
    destination = layout.output_path(task.id)
    working = temporary_workbook_path(layout, task.id)
    started = time.monotonic()
    status = "ok"
    runtime_error: str | None = None

    try:
        atomic_copy_workbook(task.init_xlsx, working)
        result = solver(task, working, recorder)
        if inspect.isawaitable(result):
            raise TypeError("TaskSolver must be synchronous")
        _raise_for_solver_result(result)
        promote_workbook(working, destination)
    except Exception as error:
        working.unlink(missing_ok=True)
        safe_error = _safe_error(error)
        runtime_error = safe_error
        status = f"error: {safe_error}"[:200]
        atomic_copy_workbook(task.init_xlsx, destination)
    finally:
        working.unlink(missing_ok=True)

    task_latency_ms = _elapsed_ms(started)
    if recorder.records:
        terminal_fields: dict[str, object] = {
            "task_status": "ok" if status == "ok" else "error",
            "task_latency_ms": task_latency_ms,
        }
        if runtime_error is not None:
            terminal_fields["runtime_error"] = runtime_error
        recorder.update_last(**terminal_fields)

    write_trace(layout, task.id, recorder)
    return _TaskRun(
        prediction=Prediction(
            id=task.id,
            output=layout.relative_output(task.id),
            status=status,
        ),
        latency_ms=task_latency_ms,
    )


def _raise_for_solver_result(result: SolveResult | None) -> None:
    if result is None:
        return
    if not isinstance(result, SolveResult):
        raise TypeError(f"solver returned unsupported result type {type(result).__name__}")
    if result.status.casefold() not in {"ok", "success"}:
        raise TaskSolveError(result.error or result.status)


def _repair_messages(
    base: list[dict[str, str]],
    previous_reply: str,
    error: Exception | None,
) -> list[dict[str, str]]:
    problem = _safe_error(error or RuntimeError("the previous plan was rejected"))
    instruction = (
        "The previous SolvePlan could not be parsed or applied safely. "
        f"Deterministic error: {problem}. Return one corrected SolvePlan JSON object only. "
        "Keep the original spreadsheet instruction and workbook context unchanged."
    )
    return [
        *base,
        {"role": "assistant", "content": previous_reply},
        {"role": "user", "content": instruction},
    ]


def _cell_truncation_recovery_messages(
    base: list[dict[str, str]],
) -> list[dict[str, str]]:
    """Request one fresh answer-only completion without replaying partial output."""

    messages = [dict(message) for message in base]
    final_user_index = next(
        (index for index in range(len(messages) - 1, -1, -1) if messages[index]["role"] == "user"),
        None,
    )
    if final_user_index is None:
        raise TaskSolveError("cell truncation recovery requires a user instruction")
    final_user = messages[final_user_index]
    messages[final_user_index] = {
        "role": "user",
        "content": f"{final_user['content']}\n\n/no_think",
    }
    messages.append({"role": "assistant", "content": "<think>\n\n</think>\n\n"})
    return messages


def _validate_tasks(tasks: Sequence[TaskSpec]) -> None:
    seen: set[str] = set()
    for index, task in enumerate(tasks):
        if not isinstance(task, TaskSpec):
            raise TypeError(f"task {index} is not a TaskSpec")
        task_id = safe_task_id(task.id)
        if task_id in seen:
            raise ArtifactError(f"duplicate task id: {task_id!r}")
        seen.add(task_id)
        if not Path(task.init_xlsx).is_file():
            raise ArtifactError(f"task {task_id!r} init workbook is missing")


def _validate_answer_sheets(task: TaskSpec, workbook_path: Path) -> None:
    """Require every declared result worksheet to exist after a successful plan."""

    try:
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise TaskSolveError(f"cannot inspect candidate workbook: {error}") from None
    try:
        missing = sorted({target.sheet for target in task.answer_ranges} - set(workbook.sheetnames))
    finally:
        workbook.close()
    if missing:
        raise TaskSolveError(
            "candidate workbook is missing required answer worksheet(s): "
            + ", ".join(repr(name) for name in missing)
        )


def _safe_error(error: Exception) -> str:
    try:
        from exactsource.model import redact_secrets

        message = redact_secrets(error)
    except Exception:
        message = str(error)
        key = os.environ.get(API_KEY_ENV)
        if key:
            message = message.replace(key, "[REDACTED]")
    message = " ".join(message.split())
    return f"{type(error).__name__}: {message}"[:500]


def _elapsed_ms(started: float) -> int:
    return max(0, round((time.monotonic() - started) * 1_000))


def _non_negative_latency(value: object, *, fallback_started: float) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        return _elapsed_ms(fallback_started)
    return value
