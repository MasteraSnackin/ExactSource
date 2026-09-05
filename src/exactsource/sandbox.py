"""Isolated execution for broad, model-generated workbook transformations.

This is a defence-in-depth boundary for code that already runs inside the submission
container. Generated code receives an in-memory openpyxl Workbook and an immutable
cached-value view; input and output paths remain exclusively in the trusted worker
wrapper.
"""

from __future__ import annotations

import ast
import hashlib
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import tempfile
import time as time_module
from copy import copy as copy_value
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from types import MappingProxyType
from typing import Any, NoReturn

from openpyxl import load_workbook

from exactsource.formula_safety import (
    snapshot_cell_hyperlinks,
    snapshot_formula_metadata,
    snapshot_formula_texts,
    validate_changed_cell_hyperlinks,
    validate_changed_formula_metadata_safety,
    validate_changed_formula_safety,
)

DEFAULT_TRANSFORM_TIMEOUT = 20.0
MAX_CODE_CHARS = 60_000
MAX_OUTPUT_BYTES = 256 * 1024 * 1024
_WORKER_FLAG = "--exactsource-transform-worker"

_BANNED_NODES = (
    ast.AsyncFunctionDef,
    ast.Await,
    ast.ClassDef,
    ast.Global,
    ast.Import,
    ast.ImportFrom,
    ast.Lambda,
    ast.Nonlocal,
    ast.While,
    ast.With,
    ast.AsyncWith,
    ast.Yield,
    ast.YieldFrom,
)
_BANNED_NAMES = frozenset(
    {
        "__builtins__",
        "breakpoint",
        "compile",
        "delattr",
        "dir",
        "eval",
        "exec",
        "exit",
        "getattr",
        "globals",
        "help",
        "input",
        "locals",
        "memoryview",
        "open",
        "os",
        "pathlib",
        "quit",
        "setattr",
        "socket",
        "subprocess",
        "sys",
        "type",
        "vars",
        "__import__",
    }
)
_BANNED_ATTRIBUTES = frozenset(
    {
        "chdir",
        "connect",
        "environ",
        "execl",
        "execle",
        "execlp",
        "execlpe",
        "execv",
        "execve",
        "execvp",
        "execvpe",
        "extract",
        "extractall",
        "filename",
        "fork",
        "forkpty",
        "format",
        "format_map",
        "getenv",
        "kill",
        "makedirs",
        "mkdir",
        "modules",
        "now",
        "open",
        "path",
        "popen",
        "read",
        "read_bytes",
        "read_text",
        "readline",
        "readlines",
        "rename",
        "replace",
        "rmdir",
        "save",
        "send",
        "sendall",
        "spawnl",
        "spawnle",
        "spawnlp",
        "spawnlpe",
        "spawnv",
        "spawnve",
        "spawnvp",
        "spawnvpe",
        "system",
        "sys",
        "today",
        "truncate",
        "unlink",
        "utcnow",
        "walk",
        "write",
        "write_bytes",
        "write_text",
        "writelines",
        "writestr",
    }
)


class SandboxError(RuntimeError):
    """Base class for generated-code validation and execution failures."""


class SandboxValidationError(SandboxError):
    """Raised when generated source violates the restricted code contract."""


class SandboxExecutionError(SandboxError):
    """Raised when screened code fails or exceeds an execution boundary."""


def _validation_error(message: str, node: ast.AST | None = None) -> NoReturn:
    location = ""
    if node is not None and getattr(node, "lineno", None):
        location = f" at line {node.lineno}"
    raise SandboxValidationError(f"generated transform rejected{location}: {message}")


def screen_transform(code: str) -> dict[str, int]:
    """Validate that code defines one restricted ``transform(wb)`` function."""

    if not isinstance(code, str) or not code.strip():
        raise SandboxValidationError("generated transform is empty")
    if len(code) > MAX_CODE_CHARS:
        raise SandboxValidationError(f"generated transform exceeds {MAX_CODE_CHARS:,} characters")
    if "\x00" in code:
        raise SandboxValidationError("generated transform contains a null byte")
    try:
        tree = ast.parse(code, mode="exec")
    except SyntaxError as exc:
        raise SandboxValidationError(
            f"generated transform has invalid Python syntax at line {exc.lineno}: {exc.msg}"
        ) from None

    # Report imports as the security violation they are, even though an import also
    # makes the module fail the one-function structural rule below.
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            _validation_error(f"{type(node).__name__} is not allowed", node)

    executable_body = [
        node
        for node in tree.body
        if not (
            isinstance(node, ast.Expr)
            and isinstance(node.value, ast.Constant)
            and isinstance(node.value.value, str)
        )
    ]
    if len(executable_body) != 1 or not isinstance(executable_body[0], ast.FunctionDef):
        _validation_error("module must contain exactly one function and optional docstring")
    function = executable_body[0]
    if function.name != "transform":
        _validation_error("function must be named transform", function)
    arguments = function.args
    if (
        len(arguments.posonlyargs) != 0
        or len(arguments.args) != 1
        or arguments.args[0].arg != "wb"
        or arguments.vararg is not None
        or arguments.kwarg is not None
        or arguments.kwonlyargs
        or arguments.defaults
        or arguments.kw_defaults
        or function.decorator_list
    ):
        _validation_error("transform must have the exact signature transform(wb)", function)

    function_count = 0
    node_count = 0
    for node in ast.walk(tree):
        node_count += 1
        if isinstance(node, ast.FunctionDef):
            function_count += 1
        if isinstance(node, _BANNED_NODES):
            _validation_error(f"{type(node).__name__} is not allowed", node)
        if isinstance(node, ast.Name) and (node.id.startswith("__") or node.id in _BANNED_NAMES):
            _validation_error(f"name {node.id!r} is not allowed", node)
        if isinstance(node, ast.Attribute) and (
            node.attr.startswith("_") or node.attr in _BANNED_ATTRIBUTES
        ):
            _validation_error(f"attribute {node.attr!r} is not allowed", node)
        if isinstance(node, ast.Constant) and isinstance(node.value, bytes):
            _validation_error("bytes literals are not allowed", node)
    if function_count != 1:
        _validation_error("nested or additional functions are not allowed")
    return {"ast_nodes": node_count, "code_chars": len(code)}


_SAFE_BUILTINS = MappingProxyType(
    {
        "Exception": Exception,
        "KeyError": KeyError,
        "TypeError": TypeError,
        "ValueError": ValueError,
        "abs": abs,
        "all": all,
        "any": any,
        "bool": bool,
        "dict": dict,
        "enumerate": enumerate,
        "filter": filter,
        "float": float,
        "int": int,
        "isinstance": isinstance,
        "iter": iter,
        "len": len,
        "list": list,
        "map": map,
        "max": max,
        "min": min,
        "next": next,
        "range": range,
        "reversed": reversed,
        "round": round,
        "set": set,
        "slice": slice,
        "sorted": sorted,
        "str": str,
        "sum": sum,
        "tuple": tuple,
        "zip": zip,
    }
)


class _SafeFacade:
    """Expose an immutable allow-list without exposing a Python module object.

    A module is not a safe capability container: seemingly harmless modules can
    retain references to ``sys``, ``os`` or imported helper modules in their global
    namespace.  The generated program therefore receives this deliberately tiny
    facade instead.  Only explicitly copied callables and constants are reachable.
    """

    __slots__ = ("_values",)

    def __init__(self, values: dict[str, Any]) -> None:
        object.__setattr__(self, "_values", MappingProxyType(dict(values)))

    def __getattribute__(self, name: str) -> Any:
        if name.startswith("_"):
            raise AttributeError(name)
        values = object.__getattribute__(self, "_values")
        try:
            return values[name]
        except KeyError:
            raise AttributeError(name) from None

    def __setattr__(self, name: str, value: Any) -> NoReturn:
        del value
        raise AttributeError(f"safe helper {name!r} is read-only")


_SAFE_RE = _SafeFacade(
    {
        "ASCII": re.ASCII,
        "DOTALL": re.DOTALL,
        "IGNORECASE": re.IGNORECASE,
        "MULTILINE": re.MULTILINE,
        "compile": re.compile,
        "escape": re.escape,
        "findall": re.findall,
        "finditer": re.finditer,
        "fullmatch": re.fullmatch,
        "match": re.match,
        "search": re.search,
        "split": re.split,
        "sub": re.sub,
        "subn": re.subn,
    }
)

_SAFE_MATH = _SafeFacade(
    {
        "acos": math.acos,
        "asin": math.asin,
        "atan": math.atan,
        "atan2": math.atan2,
        "ceil": math.ceil,
        "cos": math.cos,
        "degrees": math.degrees,
        "e": math.e,
        "exp": math.exp,
        "fabs": math.fabs,
        "factorial": math.factorial,
        "floor": math.floor,
        "fmod": math.fmod,
        "fsum": math.fsum,
        "gcd": math.gcd,
        "hypot": math.hypot,
        "inf": math.inf,
        "isclose": math.isclose,
        "isfinite": math.isfinite,
        "isinf": math.isinf,
        "isnan": math.isnan,
        "lcm": math.lcm,
        "log": math.log,
        "log10": math.log10,
        "nan": math.nan,
        "pi": math.pi,
        "pow": math.pow,
        "prod": math.prod,
        "radians": math.radians,
        "sin": math.sin,
        "sqrt": math.sqrt,
        "tan": math.tan,
        "tau": math.tau,
        "trunc": math.trunc,
    }
)

_SAFE_STATISTICS = _SafeFacade(
    {
        "fmean": statistics.fmean,
        "mean": statistics.mean,
        "median": statistics.median,
        "median_grouped": statistics.median_grouped,
        "median_high": statistics.median_high,
        "median_low": statistics.median_low,
        "mode": statistics.mode,
        "multimode": statistics.multimode,
        "pstdev": statistics.pstdev,
        "pvariance": statistics.pvariance,
        "stdev": statistics.stdev,
        "variance": statistics.variance,
    }
)


def _cached_value_view(workbook: object) -> MappingProxyType:
    """Freeze materialised, non-empty cell values by exact sheet and coordinate.

    ``Worksheet.iter_rows`` expands to the worksheet's reported dimensions and can
    create a large number of otherwise blank cells.  The trusted worker instead
    copies only cells already materialised by openpyxl's loader.  The outer and
    per-sheet mappings are both proxies so generated code cannot alter the view.
    """

    sheets: dict[str, MappingProxyType] = {}
    for worksheet in workbook.worksheets:
        values = {
            cell.coordinate: cell.value
            for cell in worksheet._cells.values()
            if cell.value is not None
        }
        sheets[worksheet.title] = MappingProxyType(values)
    return MappingProxyType(sheets)


def _safe_namespace(cached_values: MappingProxyType) -> dict[str, Any]:
    return {
        "__builtins__": _SAFE_BUILTINS,
        "cached_values": cached_values,
        "re": _SAFE_RE,
        "math": _SAFE_MATH,
        "statistics": _SAFE_STATISTICS,
        "datetime": datetime,
        "date": date,
        "time": time,
        "timedelta": timedelta,
        "Decimal": Decimal,
        "copy": copy_value,
    }


def _execute_transform(
    code: str,
    workbook: object,
    cached_values: MappingProxyType,
) -> None:
    screen_transform(code)
    namespace = _safe_namespace(cached_values)
    compiled = compile(code, "<generated-transform>", "exec", dont_inherit=True, optimize=2)
    exec(compiled, namespace, namespace)
    transform = namespace.get("transform")
    if not callable(transform):
        raise SandboxExecutionError("screened source did not define callable transform")
    result = transform(workbook)
    if result is not None:
        raise SandboxExecutionError("transform(wb) must mutate in place and return None")


def _apply_resource_limits(timeout: float) -> None:
    try:
        import resource
    except ImportError:  # pragma: no cover - only relevant on non-POSIX hosts
        return

    cpu_seconds = max(1, math.ceil(timeout))
    limits: list[tuple[int, tuple[int, int]]] = [
        (resource.RLIMIT_CPU, (cpu_seconds, cpu_seconds + 1)),
        (resource.RLIMIT_FSIZE, (MAX_OUTPUT_BYTES, MAX_OUTPUT_BYTES)),
        (resource.RLIMIT_NOFILE, (64, 64)),
    ]
    if sys.platform.startswith("linux") and hasattr(resource, "RLIMIT_AS"):
        limits.append((resource.RLIMIT_AS, (2 * 1024**3, 2 * 1024**3)))
    if hasattr(resource, "RLIMIT_NPROC"):
        # The worker never needs to create another process.  This remains useful
        # even if a future validator regression makes a process API reachable.
        limits.append((resource.RLIMIT_NPROC, (0, 0)))
    for resource_id, requested in limits:
        try:
            current_soft, current_hard = resource.getrlimit(resource_id)
            soft = requested[0] if current_soft < 0 else min(requested[0], current_soft)
            hard = requested[1] if current_hard < 0 else min(requested[1], current_hard)
            if soft > hard:
                soft = hard
            resource.setrlimit(resource_id, (soft, hard))
        except (OSError, ValueError):
            # A container runtime may impose a tighter immutable limit. Retaining that
            # tighter limit is safe, so execution can continue.
            continue


def _worker_main(arguments: list[str]) -> int:
    if len(arguments) != 5 or arguments[1] != _WORKER_FLAG:
        return 64
    source = Path(arguments[2])
    destination = Path(arguments[3])
    code_path = Path(arguments[4])
    try:
        timeout = float(os.environ.get("EXACTSOURCE_CHILD_TIMEOUT", DEFAULT_TRANSFORM_TIMEOUT))
        os.environ.clear()
        _apply_resource_limits(timeout)
        code = code_path.read_text(encoding="utf-8")
        workbook = load_workbook(source, data_only=False, keep_links=True)
        try:
            cached_workbook = load_workbook(source, data_only=True, keep_links=True)
            try:
                cached_values = _cached_value_view(cached_workbook)
            finally:
                cached_workbook.close()
            del cached_workbook
            before = list(workbook.sheetnames)
            formula_snapshot = snapshot_formula_texts(workbook)
            formula_metadata_snapshot = snapshot_formula_metadata(workbook)
            hyperlink_snapshot = snapshot_cell_hyperlinks(workbook)
            started = time_module.monotonic()
            _execute_transform(code, workbook, cached_values)
            formulae_checked = validate_changed_formula_safety(workbook, formula_snapshot)
            formula_metadata_checked = validate_changed_formula_metadata_safety(
                workbook, formula_metadata_snapshot
            )
            hyperlinks_checked = validate_changed_cell_hyperlinks(workbook, hyperlink_snapshot)
            calculation = getattr(workbook, "calculation", None)
            if calculation is not None:
                calculation.fullCalcOnLoad = True
                calculation.forceFullCalc = True
                calculation.calcMode = "auto"
            workbook.save(destination)
            after = list(workbook.sheetnames)
        finally:
            workbook.close()
        print(
            json.dumps(
                {
                    "status": "ok",
                    "sheets_before": before,
                    "sheets_after": after,
                    "formulae_checked": formulae_checked,
                    "formula_metadata_checked": formula_metadata_checked,
                    "hyperlinks_checked": hyperlinks_checked,
                    "effective_uid": (os.geteuid() if hasattr(os, "geteuid") else None),
                    "execution_ms": max(0, round((time_module.monotonic() - started) * 1_000)),
                },
                separators=(",", ":"),
            ),
            flush=True,
        )
        return 0
    except BaseException as exc:
        print(
            json.dumps(
                {
                    "status": "error",
                    "error_type": type(exc).__name__,
                    "error": str(exc)[:2_000],
                },
                separators=(",", ":"),
            ),
            flush=True,
        )
        return 2


def _child_environment(timeout: float) -> dict[str, str]:
    # No provider key or inherited user environment is passed to generated code.
    return {
        "PATH": os.defpath,
        "PYTHONHASHSEED": "0",
        "LANG": "C.UTF-8",
        "LC_ALL": "C.UTF-8",
        "EXACTSOURCE_CHILD_TIMEOUT": str(timeout),
    }


def _child_process_options(work_dir: Path, staged_files: tuple[Path, ...]) -> dict[str, Any]:
    """Return POSIX child options and drop root before executing generated code."""

    if os.name != "posix":  # pragma: no cover - the submission image is Linux
        return {}

    options: dict[str, Any] = {"umask": 0o077}
    get_effective_uid = getattr(os, "geteuid", None)
    if get_effective_uid is None or get_effective_uid() != 0:
        return options

    # Debian's unprivileged nobody/nogroup identity.  The generated worker owns
    # only its private temporary directory; the trusted parent retains /data and
    # /out access and performs the final atomic promotion.
    unprivileged_uid = 65_534
    unprivileged_gid = 65_534
    os.chown(work_dir, unprivileged_uid, unprivileged_gid)
    os.chmod(work_dir, 0o700)
    for staged_file in staged_files:
        os.chown(staged_file, unprivileged_uid, unprivileged_gid)
        os.chmod(staged_file, 0o400)
    options.update(
        {
            "user": unprivileged_uid,
            "group": unprivileged_gid,
            "extra_groups": (),
        }
    )
    return options


def _promote_atomic(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_symlink():
        raise SandboxExecutionError("destination workbook must not be a symbolic link")
    with tempfile.NamedTemporaryFile(
        prefix=f".{destination.stem}-",
        suffix=".xlsx",
        dir=destination.parent,
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
    try:
        shutil.copyfile(source, temporary)
        os.replace(temporary, destination)
        os.chmod(destination, 0o644)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def run_transform(
    code: str,
    source_path: str | Path,
    destination_path: str | Path,
    timeout: float = DEFAULT_TRANSFORM_TIMEOUT,
) -> dict[str, object]:
    """Run screened code in a resource-bounded child and write an atomic result."""

    screening = screen_transform(code)
    if not isinstance(timeout, (int, float)) or isinstance(timeout, bool):
        raise ValueError("timeout must be a positive number")
    timeout = float(timeout)
    if not 0 < timeout <= 300:
        raise ValueError("timeout must be greater than zero and at most 300 seconds")

    source = Path(source_path)
    destination = Path(destination_path)
    if not source.is_file():
        raise SandboxExecutionError(f"source workbook does not exist: {source}")
    if source.resolve() == destination.resolve():
        raise SandboxExecutionError("source and destination workbooks must be different paths")

    with tempfile.TemporaryDirectory(prefix="exactsource-transform-") as directory:
        work_dir = Path(directory)
        staged_source = work_dir / "input.xlsx"
        staged_destination = work_dir / "output.xlsx"
        staged_code = work_dir / "transform.py.txt"
        shutil.copyfile(source, staged_source)
        staged_code.write_text(code, encoding="utf-8")
        command = [
            sys.executable,
            "-I",
            str(Path(__file__).resolve()),
            _WORKER_FLAG,
            str(staged_source),
            str(staged_destination),
            str(staged_code),
        ]
        child_options = _child_process_options(
            work_dir,
            (staged_source, staged_code),
        )
        try:
            completed = subprocess.run(
                command,
                stdin=subprocess.DEVNULL,
                capture_output=True,
                text=True,
                timeout=timeout,
                env=_child_environment(timeout),
                cwd=work_dir,
                start_new_session=True,
                check=False,
                **child_options,
            )
        except subprocess.TimeoutExpired:
            raise SandboxExecutionError(
                f"generated transform exceeded the {timeout:g}-second wall-clock limit"
            ) from None

        stdout = completed.stdout.strip()
        try:
            worker_evidence = json.loads(stdout) if stdout else {}
        except json.JSONDecodeError:
            worker_evidence = {}
        if completed.returncode != 0:
            detail = worker_evidence.get("error") if isinstance(worker_evidence, dict) else None
            if not detail:
                detail = completed.stderr.strip()[:2_000] or "worker failed without details"
            raise SandboxExecutionError(f"generated transform failed: {detail}")
        if not isinstance(worker_evidence, dict) or worker_evidence.get("status") != "ok":
            raise SandboxExecutionError("generated transform worker returned malformed evidence")
        if hasattr(os, "geteuid") and os.geteuid() == 0:
            worker_uid = worker_evidence.get("effective_uid")
            if not isinstance(worker_uid, int) or worker_uid == 0:
                raise SandboxExecutionError(
                    "generated transform worker did not drop root privileges"
                )
        if not staged_destination.is_file():
            raise SandboxExecutionError("generated transform did not produce a workbook")
        if staged_destination.stat().st_size > MAX_OUTPUT_BYTES:
            raise SandboxExecutionError(
                "generated transform workbook exceeds the output size limit"
            )

        try:
            verification = load_workbook(staged_destination, read_only=True, data_only=False)
            verification.close()
        except Exception as exc:
            raise SandboxExecutionError(
                f"generated transform produced an invalid workbook: {exc}"
            ) from None
        digest = hashlib.sha256(staged_destination.read_bytes()).hexdigest()
        size = staged_destination.stat().st_size
        _promote_atomic(staged_destination, destination)

    return {
        "route": "python",
        **screening,
        "sheets_before": worker_evidence.get("sheets_before"),
        "sheets_after": worker_evidence.get("sheets_after"),
        "formulae_checked": worker_evidence.get("formulae_checked"),
        "formula_metadata_checked": worker_evidence.get("formula_metadata_checked"),
        "hyperlinks_checked": worker_evidence.get("hyperlinks_checked"),
        "effective_uid": worker_evidence.get("effective_uid"),
        "execution_ms": worker_evidence.get("execution_ms"),
        "output_bytes": size,
        "output_sha256": digest,
    }


if __name__ == "__main__":  # pragma: no cover - exercised through run_transform
    raise SystemExit(_worker_main(sys.argv))
