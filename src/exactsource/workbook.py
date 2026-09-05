"""Formula-preserving, sparse inspection of an input workbook.

Unlike a rectangular first-N-rows dump, this module keeps formulas visible,
records workbook structure and can inspect distant answer cells without walking
every blank cell in a worksheet's reported dimension.
"""

from __future__ import annotations

from collections.abc import Set
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from exactsource.contracts import QualifiedRange
from exactsource.ranges import normalise_a1_range, range_bounds, range_cell_count

CellPosition = tuple[int, int]


@dataclass(frozen=True, slots=True)
class CellSnapshot:
    sheet: str
    coordinate: str
    row: int
    column: int
    value: Any
    formula: str | None
    formula_ref: str | None
    cached_value: Any
    data_type: str
    number_format: str
    style_id: int

    @property
    def is_blank(self) -> bool:
        return self.value is None and self.formula is None


@dataclass(frozen=True, slots=True)
class TableSnapshot:
    name: str
    ref: str
    style: str | None
    columns: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class SheetManifest:
    name: str
    state: str
    declared_range: str
    effective_range: str | None
    max_row: int
    max_column: int
    materialised_cells: int
    nonempty_cells: int
    formula_cells: int
    merged_ranges: tuple[str, ...]
    tables: tuple[TableSnapshot, ...]
    auto_filter: str | None
    freeze_panes: str | None


@dataclass(frozen=True, slots=True)
class DefinedNameSnapshot:
    name: str
    refers_to: str
    local_sheet: str | None
    hidden: bool


@dataclass(frozen=True, slots=True)
class WorkbookManifest:
    sheets: tuple[SheetManifest, ...]
    defined_names: tuple[DefinedNameSnapshot, ...]


def _materialised_cells(worksheet: Worksheet) -> tuple[Cell, ...]:
    """Return instantiated cells without expanding a sparse worksheet grid."""

    cells = getattr(worksheet, "_cells", {})
    return tuple(sorted(cells.values(), key=lambda cell: (cell.row, cell.column)))


def _effective_range(cells: tuple[Cell, ...]) -> str | None:
    populated = [cell for cell in cells if cell.value is not None]
    if not populated:
        return None
    min_row = min(cell.row for cell in populated)
    max_row = max(cell.row for cell in populated)
    min_column = min(cell.column for cell in populated)
    max_column = max(cell.column for cell in populated)
    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _normalise_cached_value(value: Any) -> Any:
    # openpyxl already returns Python date/time objects.  Keeping them as such
    # lets callers choose a faithful display or comparison representation.
    if isinstance(value, (datetime, date, time, str, int, float, bool, type(None))):
        return value
    return value


def _formula_parts(value: Any, data_type: str) -> tuple[str | None, str | None]:
    if data_type != "f":
        return None, None
    if isinstance(value, str):
        return value, None
    # openpyxl represents dynamic and legacy array formulas as ArrayFormula
    # objects.  Their default repr contains a process-specific memory address,
    # so using .text is required for both fidelity and deterministic contexts.
    text = getattr(value, "text", None)
    ref = getattr(value, "ref", None)
    if isinstance(text, str):
        return text, str(ref) if ref is not None else None
    if ref is not None:
        return f"<{type(value).__name__}>", str(ref)
    return f"<{type(value).__name__}>", None


class WorkbookInspector:
    """Open the same workbook in formula and cached-value modes.

    Use as a context manager so both zip archives are closed promptly when 400
    workbooks are processed in succession.
    """

    def __init__(self, path: Path) -> None:
        self.path = Path(path)
        self._formula_workbook: openpyxl.Workbook | None = None
        self._cached_workbook: openpyxl.Workbook | None = None
        self._manifest: WorkbookManifest | None = None

    def __enter__(self) -> WorkbookInspector:
        self.open()
        return self

    def __exit__(self, _exc_type: object, _exc: object, _traceback: object) -> None:
        self.close()

    @property
    def workbook(self) -> openpyxl.Workbook:
        self.open()
        assert self._formula_workbook is not None
        return self._formula_workbook

    @property
    def cached_workbook(self) -> openpyxl.Workbook:
        self.open()
        assert self._cached_workbook is not None
        return self._cached_workbook

    def open(self) -> None:
        if self._formula_workbook is not None:
            return
        self._formula_workbook = openpyxl.load_workbook(
            self.path,
            data_only=False,
            read_only=False,
            keep_links=False,
        )
        try:
            self._cached_workbook = openpyxl.load_workbook(
                self.path,
                data_only=True,
                read_only=False,
                keep_links=False,
            )
        except Exception:
            self._formula_workbook.close()
            self._formula_workbook = None
            raise

    def close(self) -> None:
        if self._formula_workbook is not None:
            self._formula_workbook.close()
        if self._cached_workbook is not None:
            self._cached_workbook.close()
        self._formula_workbook = None
        self._cached_workbook = None

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(self.workbook.sheetnames)

    def resolve_sheet(self, requested: str) -> str:
        """Resolve exact, case-only and accidental edge-space variants.

        No active-sheet fallback is used for a genuinely absent named sheet;
        missing output sheets are meaningful for sheet-level tasks.
        """

        if requested in self.workbook.sheetnames:
            return requested
        case_matches = [
            name for name in self.workbook.sheetnames if name.casefold() == requested.casefold()
        ]
        if len(case_matches) == 1:
            return case_matches[0]
        stripped = requested.strip().casefold()
        whitespace_matches = [
            name for name in self.workbook.sheetnames if name.strip().casefold() == stripped
        ]
        if len(whitespace_matches) == 1:
            return whitespace_matches[0]
        raise KeyError(f"worksheet not found: {requested!r}")

    def worksheet(self, requested: str) -> Worksheet:
        return self.workbook[self.resolve_sheet(requested)]

    def _cached_worksheet(self, resolved: str) -> Worksheet:
        return self.cached_workbook[resolved]

    def _cell_snapshot(self, resolved_sheet: str, row: int, column: int) -> CellSnapshot:
        worksheet = self.workbook[resolved_sheet]
        cached_worksheet = self._cached_worksheet(resolved_sheet)
        cell = getattr(worksheet, "_cells", {}).get((row, column))
        cached_cell = getattr(cached_worksheet, "_cells", {}).get((row, column))
        coordinate = f"{get_column_letter(column)}{row}"
        value = None if cell is None else cell.value
        data_type = "n" if cell is None else cell.data_type
        formula, formula_ref = _formula_parts(value, data_type)
        cached_value = None if cached_cell is None else _normalise_cached_value(cached_cell.value)
        return CellSnapshot(
            sheet=resolved_sheet,
            coordinate=coordinate,
            row=row,
            column=column,
            value=value,
            formula=formula,
            formula_ref=formula_ref,
            cached_value=cached_value,
            data_type=data_type,
            number_format="General" if cell is None else cell.number_format,
            style_id=0 if cell is None else cell.style_id,
        )

    def read_range(
        self,
        sheet: str,
        cells: str,
        *,
        max_cells: int | None = None,
    ) -> tuple[CellSnapshot, ...]:
        """Read every cell in a range, including blanks, in row-major order."""

        resolved = self.resolve_sheet(sheet)
        worksheet = self.workbook[resolved]
        canonical = normalise_a1_range(cells)
        count = range_cell_count(
            canonical,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )
        if max_cells is not None and count > max_cells:
            raise ValueError(f"{resolved}!{canonical} has {count} cells; limit is {max_cells}")
        min_col, min_row, max_col, max_row = range_bounds(
            canonical,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )
        return tuple(
            self._cell_snapshot(resolved, row, column)
            for row in range(min_row, max_row + 1)
            for column in range(min_col, max_col + 1)
        )

    def sample_range(
        self,
        sheet: str,
        cells: str,
        *,
        limit: int,
        include_blank: bool = False,
        excluded: Set[CellPosition] = frozenset(),
    ) -> tuple[CellSnapshot, ...]:
        """Return a deterministic head/formula/spread/tail sample of a range."""

        if limit < 1:
            return ()
        resolved = self.resolve_sheet(sheet)
        worksheet = self.workbook[resolved]
        min_col, min_row, max_col, max_row = range_bounds(
            cells,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )
        total = (max_col - min_col + 1) * (max_row - min_row + 1)
        if total <= limit:
            snapshots = tuple(
                self._cell_snapshot(resolved, row, column)
                for row in range(min_row, max_row + 1)
                for column in range(min_col, max_col + 1)
                if (row, column) not in excluded
            )
            if include_blank:
                return snapshots
            return tuple(snapshot for snapshot in snapshots if not snapshot.is_blank)

        occupied = [
            cell
            for cell in _materialised_cells(worksheet)
            if min_row <= cell.row <= max_row
            and min_col <= cell.column <= max_col
            and cell.value is not None
            and (cell.row, cell.column) not in excluded
        ]
        positions: list[tuple[int, int]] = []

        def add(row: int, column: int) -> None:
            position = (row, column)
            if position not in excluded and position not in positions and len(positions) < limit:
                positions.append(position)

        # Existing formulas reveal local fill patterns even when they are far
        # from the rectangular top-left sample.
        formulas = [cell for cell in occupied if cell.data_type == "f"]
        for cell in _evenly_sample(formulas, max(1, limit // 3)):
            add(cell.row, cell.column)
        # Headers and first records are usually the most semantically useful.
        headers = [cell for cell in occupied if cell.row <= min_row + 2]
        for cell in _evenly_sample(headers, max(1, limit // 3)):
            add(cell.row, cell.column)
        # Add an even spread before reserving a tail sample.
        remaining = max(0, limit - len(positions))
        spread_quota = max(0, remaining * 2 // 3)
        for cell in _evenly_sample(occupied, spread_quota):
            add(cell.row, cell.column)
        for cell in reversed(occupied):
            add(cell.row, cell.column)

        if include_blank and len(positions) < limit:
            for row, column in _strategic_grid(min_col, min_row, max_col, max_row, limit):
                add(row, column)

        positions.sort()
        return tuple(self._cell_snapshot(resolved, row, column) for row, column in positions)

    def neighbourhood(
        self,
        target: QualifiedRange,
        *,
        row_margin: int = 2,
        column_margin: int = 2,
        limit: int = 180,
    ) -> tuple[CellSnapshot, ...]:
        """Inspect a target plus nearby labels, formulas and source values."""

        resolved = self.resolve_sheet(target.sheet)
        worksheet = self.workbook[resolved]
        min_col, min_row, max_col, max_row = range_bounds(
            target.cells,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )
        min_col = max(1, min_col - max(0, column_margin))
        min_row = max(1, min_row - max(0, row_margin))
        max_col += max(0, column_margin)
        max_row += max(0, row_margin)
        expanded = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        return self.sample_range(resolved, expanded, limit=limit, include_blank=True)

    def formulas_near(
        self,
        target: QualifiedRange,
        *,
        row_margin: int = 8,
        column_margin: int = 4,
        limit: int = 80,
        excluded: Set[CellPosition] = frozenset(),
    ) -> tuple[CellSnapshot, ...]:
        """Return materialised formulas close to a target range."""

        resolved = self.resolve_sheet(target.sheet)
        worksheet = self.workbook[resolved]
        min_col, min_row, max_col, max_row = range_bounds(
            target.cells,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )
        min_col = max(1, min_col - max(0, column_margin))
        min_row = max(1, min_row - max(0, row_margin))
        max_col += max(0, column_margin)
        max_row += max(0, row_margin)
        formulas = [
            cell
            for cell in _materialised_cells(worksheet)
            if cell.data_type == "f"
            and min_row <= cell.row <= max_row
            and min_col <= cell.column <= max_col
            and (cell.row, cell.column) not in excluded
        ]
        return tuple(
            self._cell_snapshot(resolved, cell.row, cell.column)
            for cell in _evenly_sample(formulas, limit)
        )

    def populated_sample(
        self,
        sheet: str,
        *,
        limit: int = 100,
        excluded: Set[CellPosition] = frozenset(),
    ) -> tuple[CellSnapshot, ...]:
        resolved = self.resolve_sheet(sheet)
        worksheet = self.workbook[resolved]
        cells = [
            cell
            for cell in _materialised_cells(worksheet)
            if cell.value is not None and (cell.row, cell.column) not in excluded
        ]
        selected: list[Cell] = []

        def add(cell: Cell) -> None:
            if cell not in selected and len(selected) < limit:
                selected.append(cell)

        formulas = [cell for cell in cells if cell.data_type == "f"]
        for cell in _evenly_sample(formulas, max(1, limit // 3)):
            add(cell)
        headers = [cell for cell in cells if cell.row <= 3]
        for cell in _evenly_sample(headers, max(1, limit // 3)):
            add(cell)
        for cell in _evenly_sample(cells, max(0, limit - len(selected)) * 2 // 3):
            add(cell)
        for cell in reversed(cells):
            add(cell)
        selected.sort(key=lambda cell: (cell.row, cell.column))
        return tuple(self._cell_snapshot(resolved, cell.row, cell.column) for cell in selected)

    def formula_sample(
        self,
        sheet: str,
        *,
        limit: int = 100,
        excluded: Set[CellPosition] = frozenset(),
    ) -> tuple[CellSnapshot, ...]:
        """Return a deterministic workbook-wide sample containing only formulas."""

        if limit < 1:
            return ()
        resolved = self.resolve_sheet(sheet)
        worksheet = self.workbook[resolved]
        formulas = [
            cell
            for cell in _materialised_cells(worksheet)
            if cell.data_type == "f" and (cell.row, cell.column) not in excluded
        ]
        return tuple(
            self._cell_snapshot(resolved, cell.row, cell.column)
            for cell in _evenly_sample(formulas, limit)
        )

    def manifest(self) -> WorkbookManifest:
        if self._manifest is not None:
            return self._manifest

        sheets: list[SheetManifest] = []
        for worksheet in self.workbook.worksheets:
            cells = _materialised_cells(worksheet)
            nonempty = tuple(cell for cell in cells if cell.value is not None)
            tables: list[TableSnapshot] = []
            for table_name in sorted(worksheet.tables, key=str.casefold):
                table = worksheet.tables[table_name]
                style_info = getattr(table, "tableStyleInfo", None)
                tables.append(
                    TableSnapshot(
                        name=str(table_name),
                        ref=str(table.ref),
                        style=None if style_info is None else style_info.name,
                        columns=tuple(str(column.name) for column in table.tableColumns),
                    )
                )
            freeze_panes = worksheet.freeze_panes
            if hasattr(freeze_panes, "coordinate"):
                freeze_panes = freeze_panes.coordinate
            sheets.append(
                SheetManifest(
                    name=worksheet.title,
                    state=worksheet.sheet_state,
                    declared_range=worksheet.calculate_dimension(),
                    effective_range=_effective_range(cells),
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                    materialised_cells=len(cells),
                    nonempty_cells=len(nonempty),
                    formula_cells=sum(cell.data_type == "f" for cell in nonempty),
                    merged_ranges=tuple(
                        sorted(str(item) for item in worksheet.merged_cells.ranges)
                    ),
                    tables=tuple(tables),
                    auto_filter=worksheet.auto_filter.ref,
                    freeze_panes=None if freeze_panes is None else str(freeze_panes),
                )
            )

        names: list[DefinedNameSnapshot] = []
        for name in sorted(
            self.workbook.defined_names.values(), key=lambda item: item.name.casefold()
        ):
            local_sheet: str | None = None
            local_id = getattr(name, "localSheetId", None)
            if isinstance(local_id, int) and 0 <= local_id < len(self.workbook.sheetnames):
                local_sheet = self.workbook.sheetnames[local_id]
            names.append(
                DefinedNameSnapshot(
                    name=name.name,
                    refers_to=str(getattr(name, "attr_text", "")),
                    local_sheet=local_sheet,
                    hidden=bool(getattr(name, "hidden", False)),
                )
            )
        self._manifest = WorkbookManifest(sheets=tuple(sheets), defined_names=tuple(names))
        return self._manifest


def _evenly_sample(items: list[Any], limit: int) -> list[Any]:
    if limit <= 0 or not items:
        return []
    if len(items) <= limit:
        return list(items)
    if limit == 1:
        return [items[0]]
    indexes = sorted({round(index * (len(items) - 1) / (limit - 1)) for index in range(limit)})
    return [items[index] for index in indexes]


def _strategic_grid(
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    limit: int,
) -> tuple[tuple[int, int], ...]:
    rows = _axis_points(min_row, max_row, max(1, int(limit**0.5)))
    columns = _axis_points(min_col, max_col, max(1, limit // max(1, len(rows))))
    return tuple((row, column) for row in rows for column in columns)[:limit]


def _axis_points(start: int, end: int, limit: int) -> tuple[int, ...]:
    if end <= start or limit <= 1:
        return (start,)
    length = end - start + 1
    if length <= limit:
        return tuple(range(start, end + 1))
    return tuple(
        sorted({start + round(index * (length - 1) / (limit - 1)) for index in range(limit)})
    )


def inspect_workbook(path: Path) -> WorkbookManifest:
    """Convenience wrapper for callers that only need structural metadata."""

    with WorkbookInspector(path) as inspector:
        return inspector.manifest()


def read_exact_range(path: Path, target: QualifiedRange) -> tuple[CellSnapshot, ...]:
    """Convenience wrapper for a formula-preserving exact range read."""

    with WorkbookInspector(path) as inspector:
        return inspector.read_range(target.sheet, target.cells)
