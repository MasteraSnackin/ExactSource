from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import openpyxl

from exactsource.context import build_context
from exactsource.contracts import QualifiedRange, TaskSpec
from exactsource.model import serialise_request_payload
from exactsource.prompts import (
    CELL_PROMPT_PLAN_SCHEMA_TEXT,
    SHEET_PROMPT_PLAN_SCHEMA_TEXT,
    build_messages,
)


def _module():
    path = Path(__file__).parents[1] / "tools" / "profile_prompts.py"
    spec = importlib.util.spec_from_file_location("profile_prompts", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _task(path: Path, task_id: str, *, sheet_level: bool = False) -> TaskSpec:
    return TaskSpec(
        id=task_id,
        instruction_type=("Sheet-Level Manipulation" if sheet_level else "Cell-Level Manipulation"),
        instruction="Fill the answer from the source values.",
        spreadsheet_path=f"spreadsheet/{task_id}",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Data", "C2"),),
        data_position="Data!A1:B3",
    )


def test_profile_task_uses_the_production_request_builder(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "init.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Left", "Right", "Answer"])
    sheet.append([2, 3, None])
    sheet.append([5, 7, None])
    workbook.save(workbook_path)

    task = _task(workbook_path, "synthetic-profile")
    profile = module.profile_task(task)

    assert profile.task_id == "synthetic-profile"
    assert profile.kind == "cell"
    assert profile.context_original_chars == profile.context_emitted_chars
    assert not profile.context_truncated
    assert profile.message_content_chars > profile.user_content_chars
    assert profile.user_content_chars > profile.context_emitted_chars
    assert profile.request_chars > profile.context_emitted_chars
    assert profile.request_utf8_bytes >= profile.request_chars
    assert profile.renderer_input_tokens is None
    messages = build_messages(task, build_context(task))
    _validated, _payload, serialised = serialise_request_payload(messages)
    assert profile.request_chars == len(serialised)
    assert profile.request_utf8_bytes == len(serialised.encode("utf-8"))


def test_exact_renderer_token_counts_are_optional_and_reported(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "Source"
    workbook.save(workbook_path)
    seen_messages: list[list[dict[str, str]]] = []

    def count(messages: list[dict[str, str]]) -> int:
        seen_messages.append(messages)
        return 123

    profile = module.profile_task(
        _task(workbook_path, "token-profile"),
        renderer_token_counter=count,
    )
    report = module.build_report(
        [profile],
        dataset_sha256="b" * 64,
        selection_name="synthetic",
        renderer_metadata={"renderer": "test-renderer"},
    )

    assert len(seen_messages) == 1
    assert [message["role"] for message in seen_messages[0]] == ["system", "user"]
    assert profile.renderer_input_tokens == 123
    assert report["schema_version"] == 3
    assert report["summary"]["renderer_input_tokens"] == {
        "min": 123,
        "mean": 123.0,
        "p50": 123,
        "p90": 123,
        "p95": 123,
        "p99": 123,
        "max": 123,
    }
    token_config = report["configuration"]["renderer_token_counting"]
    assert token_config == {
        "definition": "generation prompt before model output",
        "enabled": True,
        "renderer": "test-renderer",
    }


def test_report_identifies_each_route_specific_solve_plan_schema(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.save(workbook_path)
    profiles = [
        module.profile_task(_task(workbook_path, "cell-schema")),
        module.profile_task(_task(workbook_path, "sheet-schema", sheet_level=True)),
    ]

    report = module.build_report(
        profiles,
        dataset_sha256="d" * 64,
        selection_name="synthetic",
    )

    schema_config = report["configuration"]["solve_plan_schemas"]
    assert schema_config == {
        "cell": {
            "chars": len(CELL_PROMPT_PLAN_SCHEMA_TEXT),
            "sha256": module._sha256_text(CELL_PROMPT_PLAN_SCHEMA_TEXT),
        },
        "sheet": {
            "chars": len(SHEET_PROMPT_PLAN_SCHEMA_TEXT),
            "sha256": module._sha256_text(SHEET_PROMPT_PLAN_SCHEMA_TEXT),
        },
    }
    assert schema_config["cell"] != schema_config["sheet"]
    assert "solve_plan_schema_chars" not in report["configuration"]
    assert "solve_plan_schema_sha256" not in report["configuration"]


def test_report_rejects_partially_populated_renderer_counts(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.save(workbook_path)
    counted = module.profile_task(
        _task(workbook_path, "counted"), renderer_token_counter=lambda _messages: 7
    )
    uncounted = module.profile_task(_task(workbook_path, "uncounted"))

    try:
        module.build_report(
            [counted, uncounted],
            dataset_sha256="c" * 64,
            selection_name="synthetic",
        )
    except ValueError as exc:
        assert "every profile or none" in str(exc)
    else:
        raise AssertionError("partial renderer token counts should fail")


def test_report_is_deterministic_and_uses_nearest_rank_percentiles(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = "Source"
    workbook.save(workbook_path)
    profiles = [
        module.profile_task(_task(workbook_path, "b", sheet_level=True)),
        module.profile_task(_task(workbook_path, "a")),
    ]

    first = module.build_report(
        profiles,
        dataset_sha256="a" * 64,
        selection_name="split.json:development_ids",
    )
    second = module.build_report(
        list(reversed(profiles)),
        dataset_sha256="a" * 64,
        selection_name="split.json:development_ids",
    )

    assert first == second
    assert first["summary"]["tasks_by_kind"] == {"cell": 1, "sheet": 1}
    assert [item["task_id"] for item in first["task_profiles"]] == ["a", "b"]
    assert first["method"].endswith("no model calls or golden access")
    assert module.distribution([1, 2, 3, 100]) == {
        "min": 1,
        "mean": 26.5,
        "p50": 2,
        "p90": 100,
        "p95": 100,
        "p99": 100,
        "max": 100,
    }


def test_selection_rejects_duplicates_and_reads_integer_ids(tmp_path: Path) -> None:
    module = _module()
    selection = tmp_path / "split.json"
    selection.write_text(json.dumps({"development_ids": [1, "2"]}), encoding="utf-8")

    selected, label = module._selection(selection, "development_ids")

    assert selected == {"1", "2"}
    assert label == "split.json:development_ids"

    selection.write_text(json.dumps({"development_ids": ["1", "1"]}), encoding="utf-8")
    try:
        module._selection(selection, "development_ids")
    except ValueError as exc:
        assert "duplicate" in str(exc)
    else:
        raise AssertionError("duplicate selection ids should fail")
