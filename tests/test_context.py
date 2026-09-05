from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

from exactsource.context import _cell_line, _clip_section, build_context
from exactsource.contracts import QualifiedRange, TaskSpec
from exactsource.prompts import build_messages
from exactsource.workbook import CellSnapshot, WorkbookInspector, inspect_workbook, read_exact_range


def _rich_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inputs"
    sheet.append(["Item", "Amount", "Tax", "Total"])
    sheet.append(["Alpha", 10, 0.2, "=B2*(1+C2)"])
    sheet.append(["Beta", 20, 0.1, "=B3*(1+C3)"])
    sheet["B2"].number_format = "£#,##0.00"
    sheet.merge_cells("A5:B5")
    sheet["A5"] = "Merged note"
    sheet["Z200"] = "distant source"
    sheet["F2"] = ArrayFormula(ref="F2:F3", text="=_xlfn.SEQUENCE(2)")
    table = Table(displayName="InputTable", ref="A1:D3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    workbook.defined_names.add(DefinedName("TaxCell", attr_text="'Inputs'!$C$2"))

    output = workbook.create_sheet("Output")
    output["A1"] = "Result"
    output["A2"] = "=Inputs!D2"
    workbook.save(path)


def _task(path: Path) -> TaskSpec:
    return TaskSpec(
        id="synthetic-1",
        instruction_type="Cell-Level Manipulation",
        instruction="Fill Output!A3 from the matching input row.",
        spreadsheet_path="spreadsheet/synthetic-1",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Output", "A3"),),
        data_position="'Inputs'!A1:D3,'Inputs'!Z200",
    )


def test_workbook_inspection_preserves_formulas_and_structure(tmp_path: Path) -> None:
    path = tmp_path / "init.xlsx"
    _rich_workbook(path)

    manifest = inspect_workbook(path)
    inputs = manifest.sheets[0]
    assert inputs.formula_cells == 3
    assert inputs.effective_range == "A1:Z200"
    assert inputs.merged_ranges == ("A5:B5",)
    assert inputs.tables[0].name == "InputTable"
    assert inputs.tables[0].columns == ("Item", "Amount", "Tax", "Total")
    assert manifest.defined_names[0].name == "TaxCell"

    cells = read_exact_range(path, QualifiedRange("Inputs", "D2:D3"))
    assert [cell.formula for cell in cells] == ["=B2*(1+C2)", "=B3*(1+C3)"]


def test_exact_range_reads_blanks_without_materialising_them(tmp_path: Path) -> None:
    path = tmp_path / "init.xlsx"
    _rich_workbook(path)
    with WorkbookInspector(path) as inspector:
        before = inspector.manifest().sheets[1].materialised_cells
        cells = inspector.read_range("output", "A1:B3")
        after = inspector.manifest().sheets[1].materialised_cells
    assert len(cells) == 6
    assert cells[-1].coordinate == "B3"
    assert cells[-1].is_blank
    assert before == after


def test_context_is_formula_aware_deterministic_and_budgeted(tmp_path: Path) -> None:
    path = tmp_path / "init.xlsx"
    _rich_workbook(path)
    task = _task(path)

    first = build_context(task, char_budget=50_000)
    second = build_context(task, char_budget=50_000)

    assert first == second
    assert first.sha256 == hashlib.sha256(first.text.encode("utf-8")).hexdigest()
    assert 'formula="=B2*(1+C2)"' in first.text
    assert 'formula="=_xlfn.SEQUENCE(2)"' in first.text
    assert 'array_ref="F2:F3"' in first.text
    assert "openpyxl.worksheet.formula" not in first.text
    assert "distant source" in first.text
    assert "InputTable" in first.text
    assert 'columns=["Item", "Amount", "Tax", "Total"]' in first.text
    assert "TaxCell" in first.text
    assert "A5:B5" in first.text
    assert not first.truncated
    assert first.text.startswith("## Answer-target context")
    assert "# ExactSource task context" not in first.text
    assert "## Instruction" not in first.text
    assert "## Graded answer ranges" not in first.text

    clipped = build_context(task, char_budget=900)
    assert len(clipped.text) <= 900
    assert clipped.original_chars == first.original_chars
    assert clipped.truncated
    assert "CONTEXT TRUNCATED" in clipped.text

    one_character = build_context(task, char_budget=1)
    assert len(one_character.text) == 1
    assert one_character.text.strip()
    assert one_character.truncated
    assert build_messages(task, one_character)[1]["content"]


def test_context_reports_a_required_output_sheet_that_is_not_yet_present(tmp_path: Path) -> None:
    path = tmp_path / "init.xlsx"
    _rich_workbook(path)
    task = _task(path)
    task = TaskSpec(
        id=task.id,
        instruction_type="Sheet-Level Manipulation",
        instruction=task.instruction,
        spreadsheet_path=task.spreadsheet_path,
        init_xlsx=task.init_xlsx,
        answer_ranges=(QualifiedRange("New Results", "A1:C3"),),
        data_position=task.data_position,
    )

    context = build_context(task)

    assert "requested worksheet does not exist" in context.text
    assert "may need to create it" in context.text


def _declared_source_section(text: str) -> str:
    return text.split("## Declared source regions", 1)[1].split(
        "## Workbook-wide formula patterns", 1
    )[0]


def _sheet_level_task(
    path: Path,
    *,
    instruction: str,
    data_position: str,
    answer_sheet: str = "New Output",
) -> TaskSpec:
    return TaskSpec(
        id="sheet-source-test",
        instruction_type="Sheet-Level Manipulation",
        instruction=instruction,
        spreadsheet_path="spreadsheet/sheet-source-test",
        init_xlsx=path,
        answer_ranges=(QualifiedRange(answer_sheet, "A1:B2"),),
        data_position=data_position,
    )


def test_sheet_level_unqualified_source_uses_the_only_existing_input_sheet(
    tmp_path: Path,
) -> None:
    path = tmp_path / "single-input.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Inputs"
    workbook.active["A1"] = "source header"
    workbook.active["A2"] = 42
    workbook.save(path)
    task = _sheet_level_task(
        path,
        instruction="Create the New Output worksheet from Inputs.",
        data_position="A1:B2",
    )

    context = build_context(task, char_budget=50_000)
    sources = _declared_source_section(context.text)

    assert "### Inputs!A1:B2" in sources
    assert "source header" in sources
    assert "### 'New Output'!A1:B2" not in sources
    assert "requested worksheet does not exist" in context.text


def test_sheet_level_unqualified_source_samples_all_sheets_with_mentions_first(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multiple-inputs.xlsx"
    workbook = openpyxl.Workbook()
    archive = workbook.active
    archive.title = "Archive"
    archive["A1"] = "archive value"
    inputs = workbook.create_sheet("Input Data")
    inputs["A1"] = "priority value"
    lookup = workbook.create_sheet("Lookup")
    lookup["A1"] = "lookup value"
    workbook.save(path)
    task = _sheet_level_task(
        path,
        instruction="Create New Output using Input Data and retain the lookup values.",
        data_position="A1:B2",
    )

    first = build_context(task, char_budget=50_000)
    second = build_context(task, char_budget=50_000)
    sources = _declared_source_section(first.text)

    assert first == second
    assert "### Archive!A1:B2" in sources
    assert "### 'Input Data'!A1:B2" in sources
    assert "### Lookup!A1:B2" in sources
    assert sources.index("### 'Input Data'!A1:B2") < sources.index("### Archive!A1:B2")
    assert sources.index("### Lookup!A1:B2") < sources.index("### Archive!A1:B2")
    assert "priority value" in sources
    assert "archive value" in sources
    assert "lookup value" in sources


def test_sheet_level_qualified_source_is_not_expanded_to_other_sheets(tmp_path: Path) -> None:
    path = tmp_path / "qualified-source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Archive"
    workbook.active["A1"] = "not declared"
    inputs = workbook.create_sheet("Input Data")
    inputs["A1"] = "declared"
    workbook.save(path)
    task = _sheet_level_task(
        path,
        instruction="Create New Output using Input Data.",
        data_position="'Input Data'!A1:B2",
    )

    sources = _declared_source_section(build_context(task, char_budget=50_000).text)

    assert "### 'Input Data'!A1:B2" in sources
    assert "### Archive!A1:B2" not in sources
    assert sources.count("### ") == 1


def test_manifest_context_lists_columns_for_every_table(tmp_path: Path) -> None:
    path = tmp_path / "many-tables.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    for index in range(81):
        column = get_column_letter(index + 1)
        header = f"Header {index:03d}"
        sheet[f"{column}1"] = header
        sheet[f"{column}2"] = index
        sheet.add_table(
            Table(
                displayName=f"Catalog_{index:03d}",
                ref=f"{column}1:{column}2",
            )
        )
    workbook.save(path)
    task = TaskSpec(
        id="table-manifest-test",
        instruction_type="Cell-Level Manipulation",
        instruction="Read the catalog.",
        spreadsheet_path="spreadsheet/table-manifest-test",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Catalog", "A2"),),
        data_position="Catalog!A1:A2",
    )

    manifest = inspect_workbook(path)
    context = build_context(task, char_budget=100_000)
    last_table_line = next(
        line for line in context.text.splitlines() if 'table="Catalog_080"' in line
    )

    assert len(manifest.sheets[0].tables) == 81
    assert manifest.sheets[0].tables[-1].columns == ("Header 080",)
    assert 'columns=["Header 080"]' in last_table_line
    assert "tables_omitted" not in context.text


def test_oversized_context_reserves_every_evidence_section_and_full_manifest(
    tmp_path: Path,
) -> None:
    path = tmp_path / "oversized.xlsx"
    workbook = openpyxl.Workbook()
    signals = workbook.active
    signals.title = "Signals"
    signals["A1"] = "sparse sentinel"
    signals["B1"] = "=1+1"

    bulk = workbook.create_sheet("Bulk")
    bulk.append(["Long source", "Calculated value"])
    for row in range(2, 222):
        bulk.append([f"bulk-{row}-" + "x" * 400, f"=LEN(A{row})"])
    bulk.add_table(Table(displayName="BulkTable", ref="A1:B221"))
    workbook.save(path)
    task = _sheet_level_task(
        path,
        instruction="Create New Output from the Bulk worksheet.",
        data_position="Bulk!A1:B221",
    )

    first = build_context(task, char_budget=12_000)
    second = build_context(task, char_budget=12_000)

    assert first == second
    assert first.truncated
    assert len(first.text) <= 12_000
    assert first.sha256 == hashlib.sha256(first.text.encode("utf-8")).hexdigest()
    for heading in (
        "## Answer-target context",
        "## Workbook structure",
        "## Declared source regions",
        "## Workbook-wide formula patterns",
        "## Other populated workbook cells",
    ):
        assert heading in first.text

    assert "## Instruction" not in first.text
    assert "## Graded answer ranges" not in first.text

    manifest = first.text.split("## Workbook structure", 1)[1].split(
        "## Declared source regions", 1
    )[0]
    sources = _declared_source_section(first.text)
    formulas = first.text.split("## Workbook-wide formula patterns", 1)[1].split(
        "## Other populated workbook cells", 1
    )[0]
    sparse = first.text.split("## Other populated workbook cells", 1)[1]

    assert 'table="BulkTable"' in manifest
    assert 'columns=["Long source", "Calculated value"]' in manifest
    assert sources.index("### Bulk!A1:B221") < sources.index("bulk-2-")
    assert 'formula="=1+1"' in formulas
    assert "sparse sentinel" in sparse
    assert "SECTION TRUNCATED" in first.text
    assert "CONTEXT TRUNCATED" in first.text


def test_cell_evidence_prefers_target_over_overlapping_declared_source(
    tmp_path: Path,
) -> None:
    path = tmp_path / "overlapping-evidence.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A2"] = 21
    sheet["B2"] = "=A2*2"
    sheet["Z10"] = "distant declared source"
    workbook.save(path)
    task = TaskSpec(
        id="overlapping-evidence",
        instruction_type="Cell-Level Manipulation",
        instruction="Repair the formula in Data!B2.",
        spreadsheet_path="spreadsheet/overlapping-evidence",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Data", "B2"),),
        data_position="Data!B2,Data!Z10",
    )

    context = build_context(task, char_budget=50_000).text
    target_section = context.split("## Answer-target context", 1)[1].split(
        "## Workbook structure", 1
    )[0]
    source_section = _declared_source_section(context)

    assert 'B2: formula="=A2*2"' in target_section
    assert 'B2: formula="=A2*2"' not in source_section
    assert context.count('B2: formula="=A2*2"') == 1
    assert "distant declared source" in source_section
    assert "All sampled formulas already appear in higher-priority context." in context
    assert "No existing formulas were found" not in context


def test_same_coordinate_on_different_sheets_remains_distinct(tmp_path: Path) -> None:
    path = tmp_path / "same-coordinate.xlsx"
    workbook = openpyxl.Workbook()
    primary = workbook.active
    primary.title = "Primary"
    primary["A1"] = "primary coordinate evidence"
    secondary = workbook.create_sheet("Secondary")
    secondary["A1"] = "secondary coordinate evidence"
    workbook.save(path)
    task = TaskSpec(
        id="same-coordinate",
        instruction_type="Cell-Level Manipulation",
        instruction="Update Primary!A1 using Secondary!A1.",
        spreadsheet_path="spreadsheet/same-coordinate",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Primary", "A1"),),
        data_position="Secondary!A1",
    )

    context = build_context(task, char_budget=50_000).text
    source_section = _declared_source_section(context)

    assert context.count('A1: value="primary coordinate evidence"') == 1
    assert context.count('A1: value="secondary coordinate evidence"') == 1
    assert 'A1: value="secondary coordinate evidence"' in source_section


def test_unique_formula_and_value_evidence_survive_section_deduplication(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unique-evidence.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    sheet["Y50"] = "unique populated value"
    sheet["Z50"] = "=LEN(Y50)"
    workbook.save(path)
    task = TaskSpec(
        id="unique-evidence",
        instruction_type="Cell-Level Manipulation",
        instruction="Populate Evidence!A1.",
        spreadsheet_path="spreadsheet/unique-evidence",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Evidence", "A1"),),
    )

    context = build_context(task, char_budget=50_000).text
    formula_section = context.split("## Workbook-wide formula patterns", 1)[1].split(
        "## Other populated workbook cells", 1
    )[0]
    sparse_section = context.split("## Other populated workbook cells", 1)[1]

    assert 'Z50: formula="=LEN(Y50)"' in formula_section
    assert context.count('Z50: formula="=LEN(Y50)"') == 1
    assert 'Y50: value="unique populated value"' in sparse_section
    assert context.count('Y50: value="unique populated value"') == 1


def test_blank_target_neighbourhood_omits_only_uninformative_blank_cells(
    tmp_path: Path,
) -> None:
    path = tmp_path / "blank-neighbourhood.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Output"
    sheet["A1"] = "nearby label"
    sheet["D3"].number_format = "0.00"
    workbook.save(path)
    task = TaskSpec(
        id="blank-neighbourhood",
        instruction_type="Cell-Level Manipulation",
        instruction="Calculate Output!C3.",
        spreadsheet_path="spreadsheet/blank-neighbourhood",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Output", "C3"),),
    )

    context = build_context(task, char_budget=50_000).text
    target_section = context.split("## Answer-target context", 1)[1].split(
        "## Workbook structure", 1
    )[0]

    assert "- C3: value=null" in target_section
    assert '- D3: value=null; format="0.00"' in target_section
    assert "- B2: value=null" not in target_section
    assert "- E5: value=null" not in target_section
    assert target_section.count("value=null") == 2


def test_later_formula_sample_refills_after_target_overlap(tmp_path: Path) -> None:
    path = tmp_path / "formula-refill.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dense"
    for row in range(1, 401):
        sheet.cell(row, 1).value = f"=ROW()+{row}"
        sheet.cell(row, 2).value = f"label-{row}"
        sheet.cell(row, 3).value = row
    workbook.save(path)
    task = TaskSpec(
        id="formula-refill",
        instruction_type="Cell-Level Manipulation",
        instruction="Repair the formulae in Dense!A1:A200.",
        spreadsheet_path="spreadsheet/formula-refill",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("Dense", "A1:A200"),),
    )

    context = build_context(task, char_budget=100_000).text
    target_section = context.split("## Answer-target context", 1)[1].split(
        "## Workbook structure", 1
    )[0]
    formula_section = context.split("## Workbook-wide formula patterns", 1)[1].split(
        "## Other populated workbook cells", 1
    )[0]

    target_formula_lines = {line for line in target_section.splitlines() if "formula=" in line}
    catalogue_formula_lines = {
        line.strip() for line in formula_section.splitlines() if "formula=" in line
    }
    normalised_target_lines = {line.strip().removeprefix("- ") for line in target_formula_lines}
    normalised_catalogue_lines = {line.removeprefix("- ") for line in catalogue_formula_lines}
    assert len(normalised_catalogue_lines) == 60
    assert normalised_catalogue_lines.isdisjoint(normalised_target_lines)


def _formula_snapshot(cached_value: object) -> CellSnapshot:
    return CellSnapshot(
        sheet="Data",
        coordinate="A1",
        row=1,
        column=1,
        value="=1+1",
        formula="=1+1",
        formula_ref=None,
        cached_value=cached_value,
        data_type="f",
        number_format="General",
        style_id=0,
    )


def test_formula_lines_omit_only_unavailable_cached_values() -> None:
    assert "cached=" not in _cell_line(_formula_snapshot(None))
    assert "cached=0" in _cell_line(_formula_snapshot(0))
    assert "cached=false" in _cell_line(_formula_snapshot(False))
    assert 'cached=""' in _cell_line(_formula_snapshot(""))


def test_child_aware_clipping_water_fills_bodies_after_all_headings_and_statuses() -> None:
    text = "## Dense ranges\n\n" + "\n\n".join(
        (
            f"### Range {index}\n"
            f"- Graded cells={index * 100}\n"
            f"- sentinel-{index}: " + chr(96 + index) * 700
        )
        for index in range(1, 5)
    )
    budget = 720

    first = _clip_section(text, budget)
    second = _clip_section(text, budget)

    assert first == second
    assert len(first) <= budget
    assert "SECTION TRUNCATED" in first
    for index in range(1, 5):
        assert f"### Range {index}" in first
        assert f"- Graded cells={index * 100}" in first
        assert f"- sentinel-{index}:" in first
    assert "CHILD BLOCKS OMITTED" not in first


def test_child_aware_clipping_counts_blocks_when_all_headings_cannot_fit() -> None:
    text = "## Many ranges\n\n" + "\n\n".join(
        f"### Range {index} " + "x" * 70 + f"\n- sentinel-{index}" for index in range(1, 9)
    )

    first = _clip_section(text, 280)
    second = _clip_section(text, 280)

    assert first == second
    assert len(first) <= 280
    assert "[CHILD BLOCKS OMITTED; count=6]" in first
    assert "### Range 1" in first
    assert "### Range 2" in first
    assert "### Range 3" not in first
    assert "SECTION TRUNCATED" in first


def test_clipped_formula_and_sparse_sections_keep_every_worksheet(tmp_path: Path) -> None:
    path = tmp_path / "multi-sheet-evidence.xlsx"
    workbook = openpyxl.Workbook()
    sheet_names = ("North", "South", "East", "West")
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        sheet.title = sheet_name
        for row in range(1, 141):
            sheet.cell(row, 1).value = f"=ROW()+{sheet_index}"
            sheet.cell(row, 2).value = f"{sheet_name}-value-{row}-" + "v" * 90
    workbook.save(path)
    task = TaskSpec(
        id="multi-sheet-evidence",
        instruction_type="Cell-Level Manipulation",
        instruction="Create the result on a new worksheet.",
        spreadsheet_path="spreadsheet/multi-sheet-evidence",
        init_xlsx=path,
        answer_ranges=(QualifiedRange("New Output", "A1"),),
    )

    first = build_context(task, char_budget=9_000)
    second = build_context(task, char_budget=9_000)
    formulas = first.text.split("## Workbook-wide formula patterns", 1)[1].split(
        "## Other populated workbook cells", 1
    )[0]
    sparse = first.text.split("## Other populated workbook cells", 1)[1]

    assert first == second
    assert first.truncated
    assert len(first.text) <= 9_000
    for sheet_name in sheet_names:
        marker = f'### Worksheet "{sheet_name}"'
        assert marker in formulas
        assert marker in sparse
        assert f'formula="=ROW()+{sheet_names.index(sheet_name)}"' in formulas
        assert f"{sheet_name}-value-" in sparse


def test_task_41_47_shaped_context_keeps_every_target_and_source_block(
    tmp_path: Path,
) -> None:
    path = tmp_path / "task-41-47-shaped.xlsx"
    workbook = openpyxl.Workbook()
    cmsn = workbook.active
    cmsn.title = "CMSN"
    cstr = workbook.create_sheet("CSTR")
    expro = workbook.create_sheet("EXPRO")
    output = workbook.create_sheet("OUT CAS")

    for sheet, rows in ((cmsn, 320), (cstr, 260), (expro, 8)):
        for row in range(1, rows + 1):
            for column in range(1, 5):
                sheet.cell(row, column).value = (
                    f"{sheet.title}-r{row}-c{column}-" + sheet.title.lower() * 18
                )

    for row in range(1, 321):
        for column in range(1, 16):
            output.cell(row, column).value = f"OUT-r{row}-c{column}-" + "o" * 70
    workbook.save(path)
    task = TaskSpec(
        id="41-47-shaped",
        instruction_type="Sheet-Level Manipulation",
        instruction="Aggregate CMSN, CSTR and EXPRO into the four OUT CAS tables.",
        spreadsheet_path="spreadsheet/41-47-shaped",
        init_xlsx=path,
        answer_ranges=(
            QualifiedRange("OUT CAS", "A2:C1529"),
            QualifiedRange("OUT CAS", "E2:G586"),
            QualifiedRange("OUT CAS", "I2:K13"),
            QualifiedRange("OUT CAS", "L2:O8"),
        ),
        data_position=("CMSN!A1:D1529,CSTR!A1:D592,EXPRO!A1:D8,'OUT CAS'!A1:O1529"),
    )

    first = build_context(task, char_budget=12_000)
    second = build_context(task, char_budget=12_000)
    targets = first.text.split("## Answer-target context", 1)[1].split("## Workbook structure", 1)[
        0
    ]
    sources = _declared_source_section(first.text)

    assert first == second
    assert first.truncated
    assert len(first.text) <= 12_000
    assert first.sha256 == hashlib.sha256(first.text.encode("utf-8")).hexdigest()
    expected_targets = (
        ("### 'OUT CAS'!A2:C1529", "- Graded cells=4584"),
        ("### 'OUT CAS'!E2:G586", "- Graded cells=1755"),
        ("### 'OUT CAS'!I2:K13", "- Graded cells=36"),
        ("### 'OUT CAS'!L2:O8", "- Graded cells=28"),
    )
    expected_sources = (
        ("### CMSN!A1:D1529", "- Region cells=6116"),
        ("### CSTR!A1:D592", "- Region cells=2368"),
        ("### EXPRO!A1:D8", "- Region cells=32"),
        ("### 'OUT CAS'!A1:O1529", "- Region cells=22935"),
    )
    for heading, status in expected_targets:
        assert heading in targets
        assert status in targets
    for heading, status in expected_sources:
        assert heading in sources
        assert status in sources
    assert "CHILD BLOCKS OMITTED" not in targets
    assert "CHILD BLOCKS OMITTED" not in sources
