from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from exactsource.contracts import (
    MAX_PLAN_OPERATIONS,
    ClearRange,
    CopyRange,
    FillArrayFormula,
    FillFormula,
    QualifiedRange,
    SetArrayFormula,
    SetFormula,
    SetValue,
    SolvePlan,
    TaskSpec,
)
from exactsource.plans import (
    MAX_PLAN_CELL_WRITES,
    PlanApplicationError,
    PlanParseError,
    _validate_plan_resource_limits,
    apply_operations,
    parse_plan,
)


def _task(source: Path) -> TaskSpec:
    return TaskSpec(
        id="synthetic-1",
        instruction_type="Cell-level manipulation",
        instruction="Populate the requested cells.",
        spreadsheet_path="synthetic-1",
        init_xlsx=source,
        answer_ranges=(QualifiedRange("Output", "A2:D4"),),
    )


def _source_workbook(path: Path) -> None:
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    for row, value in enumerate((10, 20, 30), start=2):
        data.cell(row, 1, value)
    data["B2"] = "=A2*2"
    data["B2"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    output = workbook.create_sheet("Output")
    output["A3"] = "remove me"
    workbook.save(path)
    workbook.close()


def test_parse_plan_accepts_plain_and_single_fenced_json() -> None:
    payload = (
        '{"route":"operations","summary":"Write total",'
        '"operations":[{"op":"set_formula","sheet":"Output",'
        '"cell":"B2","formula":"=SUM(Data!A2:A4)"}]}'
    )

    assert parse_plan(payload).route == "operations"
    assert parse_plan(f"```json\n{payload}\n```").operations[0].cell == "B2"


@pytest.mark.parametrize(
    "response",
    [
        "Here is the plan: {}",
        "{} trailing",
        "[]",
        '{"route":"operations","summary":"x","operations":[],"extra":NaN}',
        "```json\n{}\n``` extra",
    ],
)
def test_parse_plan_rejects_ambiguous_or_invalid_output(response: str) -> None:
    with pytest.raises(PlanParseError):
        parse_plan(response)


def test_apply_operations_independently_rejects_too_many_constructed_operations(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    source.write_bytes(b"resource validation must run before workbook loading")
    operation = SetValue(op="set_value", sheet="Output", cell="A2", value=1)
    plan = SolvePlan.model_construct(
        route="operations",
        summary="Bypass contract validation to exercise the runtime guard.",
        operations=[operation] * (MAX_PLAN_OPERATIONS + 1),
        python_code=None,
    )

    with pytest.raises(PlanApplicationError, match="129 operations; limit is 128"):
        apply_operations(plan, _task(source), source, destination)

    assert source.read_bytes() == b"resource validation must run before workbook loading"
    assert not destination.exists()


def test_plan_cell_write_budget_counts_destination_rectangles_and_single_cells(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    source.write_bytes(b"resource validation must run before workbook loading")
    exactly_at_limit = SolvePlan(
        route="operations",
        summary="Exercise the aggregate write budget boundary.",
        operations=[
            ClearRange(op="clear_range", sheet="Output", range="A1:B100000"),
            ClearRange(op="clear_range", sheet="Output", range="A1:B100000"),
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="A1:A100000",
                destination_sheet="Output",
                destination_cell="A1",
            ),
        ],
    )

    _validate_plan_resource_limits(exactly_at_limit)

    over_limit = SolvePlan(
        route="operations",
        summary="Exceed the aggregate write budget by one cell.",
        operations=[
            *exactly_at_limit.operations,
            SetValue(op="set_value", sheet="Output", cell="A2", value=1),
        ],
    )
    with pytest.raises(
        PlanApplicationError,
        match=rf"{MAX_PLAN_CELL_WRITES + 1:,} cells; limit is {MAX_PLAN_CELL_WRITES:,}",
    ):
        apply_operations(over_limit, _task(source), source, destination)

    assert source.read_bytes() == b"resource validation must run before workbook loading"
    assert not destination.exists()


def test_per_operation_cell_limit_remains_independent_of_plan_budget() -> None:
    plan = SolvePlan(
        route="operations",
        summary="Exceed the per-operation range limit.",
        operations=[ClearRange(op="clear_range", sheet="Output", range="A1:B125001")],
    )

    with pytest.raises(
        PlanApplicationError,
        match="contains 250,002 cells; limit is 250,000",
    ):
        _validate_plan_resource_limits(plan)


def test_apply_operations_writes_literals_formulas_fills_copies_and_clears(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Exercise each deterministic operation.",
        operations=[
            SetValue(op="set_value", sheet="Output", cell="$A$2", value="=literal"),
            SetFormula(op="set_formula", sheet="Output", cell="B2", formula="=Data!A2*2"),
            FillFormula(
                op="fill_formula",
                sheet="Output",
                range="B2:B4",
                formula="=Data!A2*$D$1",
            ),
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="B2",
                destination_sheet="Output",
                destination_cell="D4",
                include_style=True,
            ),
            ClearRange(op="clear_range", sheet="Output", range="A3:A3"),
        ],
    )

    evidence = apply_operations(plan, _task(source), source, destination)

    assert destination.stat().st_mode & 0o777 == 0o644
    result = load_workbook(destination, data_only=False)
    try:
        output = result["Output"]
        assert output["A2"].value == "=literal"
        assert output["A2"].data_type == "s"
        assert output["B2"].value == "=Data!A2*$D$1"
        assert output["B3"].value == "=Data!A3*$D$1"
        assert output["B4"].value == "=Data!A4*$D$1"
        assert output["D4"].value == "=C4*2"
        assert output["D4"].fill.fgColor.rgb == "0000FF00"
        assert output["A3"].value is None
    finally:
        result.close()

    assert evidence["task_id"] == "synthetic-1"
    assert evidence["operations_applied"] == 5
    assert evidence["cell_writes"] == 7
    assert evidence["output_sha256"]


def test_formula_operation_rejects_external_capability_without_leaking_uri(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Attempt an unsafe external formula.",
        operations=[
            SetFormula(
                op="set_formula",
                sheet="Output",
                cell="B2",
                formula='=WEBSERVICE("https://secret.invalid/path")',
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "external-capability function" in str(caught.value)
    assert "secret.invalid" not in str(caught.value)
    assert not destination.exists()


@pytest.mark.parametrize(
    ("formula", "category"),
    [
        ("=SUM(Data!A2:A4", "unbalanced formula delimiter"),
        ("=Missing!A1", "missing worksheet reference"),
        ("='Missing Sheet'!A1", "missing worksheet reference"),
    ],
)
def test_formula_operation_rejects_confident_integrity_errors_without_leaking_names(
    tmp_path: Path,
    formula: str,
    category: str,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Attempt a statically invalid formula.",
        operations=[SetFormula(op="set_formula", sheet="Output", cell="B2", formula=formula)],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert category in str(caught.value)
    assert "Missing" not in str(caught.value)
    assert not destination.exists()


def test_formula_operation_preserves_modern_spill_syntax(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "spill.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Reference an existing spill range.",
        operations=[SetFormula(op="set_formula", sheet="Output", cell="B2", formula="=Data!A2#")],
    )

    apply_operations(plan, _task(source), source, destination)

    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["Output"]["B2"].value == "=Data!A2#"
    finally:
        workbook.close()


def test_copy_range_rejects_new_copy_of_preexisting_external_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = '=WEBSERVICE("https://legacy.invalid/path")'
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Attempt to copy an external formula into the answer range.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="C2",
                destination_sheet="Output",
                destination_cell="B2",
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "external-capability function" in str(caught.value)
    assert "legacy.invalid" not in str(caught.value)
    assert not destination.exists()


def test_copy_range_rejects_data_table_formula_without_leaking_references(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = DataTableFormula(
        ref="C2",
        r1="CONFIDENTIAL_MISSING_SHEET!A1",
    )
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Attempt to copy a what-if data table formula.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="C2",
                destination_sheet="Output",
                destination_cell="B2",
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "does not support data-table formulae" in str(caught.value)
    assert "CONFIDENTIAL_MISSING_SHEET" not in str(caught.value)
    assert not destination.exists()


def test_copy_formula_translation_error_does_not_echo_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = '=CONFIDENTIAL_FORMULA_MARKER"x"'
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Attempt to copy a formula the translator cannot tokenise.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="C2",
                destination_sheet="Output",
                destination_cell="B2",
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "copied formula could not be translated" in str(caught.value)
    assert "CONFIDENTIAL_FORMULA_MARKER" not in str(caught.value)
    assert not destination.exists()


def test_fill_formula_translation_error_does_not_echo_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Attempt to fill a formula the translator cannot tokenise.",
        operations=[
            FillFormula(
                op="fill_formula",
                sheet="Output",
                range="B2:B3",
                formula='=CONFIDENTIAL_FILL_MARKER"x"',
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "fill formula could not be translated" in str(caught.value)
    assert "CONFIDENTIAL_FILL_MARKER" not in str(caught.value)
    assert not destination.exists()


def test_copy_array_formula_translation_error_does_not_echo_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = ArrayFormula(
        ref="C2:D3",
        text='=CONFIDENTIAL_ARRAY_MARKER"x"',
    )
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Attempt to copy an array formula the translator cannot tokenise.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="C2:D3",
                destination_sheet="Output",
                destination_cell="A2",
            )
        ],
    )

    with pytest.raises(PlanApplicationError) as caught:
        apply_operations(plan, _task(source), source, destination)

    assert "copied array formula could not be translated" in str(caught.value)
    assert "CONFIDENTIAL_ARRAY_MARKER" not in str(caught.value)
    assert not destination.exists()


def test_operations_preserve_unchanged_preexisting_external_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = '=WEBSERVICE("https://legacy.invalid/path")'
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Make a safe change without altering the legacy formula.",
        operations=[SetFormula(op="set_formula", sheet="Output", cell="B2", formula="=1+1")],
    )

    apply_operations(plan, _task(source), source, destination)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Data"]["C2"].value == '=WEBSERVICE("https://legacy.invalid/path")'
        assert result["Output"]["B2"].value == "=1+1"
    finally:
        result.close()


def test_overlapping_copy_uses_a_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook["Data"].append([1, 2, 3])
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Shift cells right.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="A1:C1",
                destination_sheet="Data",
                destination_cell="B1",
            )
        ],
    )

    task = TaskSpec(
        id="overlap",
        instruction_type="Cell-level manipulation",
        instruction="Shift the requested cells.",
        spreadsheet_path="overlap",
        init_xlsx=source,
        answer_ranges=(QualifiedRange("Data", "B1:D1"),),
    )
    apply_operations(plan, task, source, destination)

    result = load_workbook(destination)
    try:
        assert [result["Data"].cell(1, column).value for column in range(1, 5)] == [
            1,
            1,
            2,
            3,
        ]
    finally:
        result.close()


@pytest.mark.parametrize(
    "operation, expected",
    [
        (
            SetValue(op="set_value", sheet="Missing", cell="A2", value=1),
            "destination worksheet",
        ),
        (
            SetValue(op="set_value", sheet="Output", cell="A2:B2", value=1),
            "not one A1 cell",
        ),
        (
            SetFormula(op="set_formula", sheet="Output", cell="A2", formula="SUM(A:A)"),
            "beginning with '='",
        ),
        (
            ClearRange(op="clear_range", sheet="Output", range="B2:A1"),
            "top-left to bottom-right",
        ),
        (
            ClearRange(op="clear_range", sheet="Output", range="XFE2"),
            "outside Excel worksheet limits",
        ),
    ],
)
def test_apply_operations_fails_closed_on_invalid_targets(
    tmp_path: Path, operation: object, expected: str
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Invalid synthetic operation.",
        operations=[operation],
    )

    with pytest.raises(PlanApplicationError, match=expected):
        apply_operations(plan, _task(source), source, destination)

    assert not destination.exists()


@pytest.mark.parametrize(
    ("formula", "stored"),
    [
        ("=XLOOKUP(A1,B:B,C:C)", "=_xlfn.XLOOKUP(A1,B:B,C:C)"),
        ("=UNIQUE(A1:A9)", "=_xlfn.UNIQUE(A1:A9)"),
        ("=LET(x,A1,x*2)", "=_xlfn.LET(x,A1,x*2)"),
        ("=CHOOSECOLS(A1:C4,2)", "=_xlfn.CHOOSECOLS(A1:C4,2)"),
        (
            "=FILTER(A1:C4,A1:A4>0)",
            "=_xlfn._xlws.FILTER(A1:C4,A1:A4>0)",
        ),
        ("=AGGREGATE(9,4,A1:A9)", "=_xlfn.AGGREGATE(9,4,A1:A9)"),
        ("=IFNA(A1,0)", "=_xlfn.IFNA(A1,0)"),
        ('=MAXIFS(A:A,B:B,">0")', '=_xlfn.MAXIFS(A:A,B:B,">0")'),
        ('=MINIFS(A:A,B:B,">0")', '=_xlfn.MINIFS(A:A,B:B,">0")'),
        (
            '=TEXTJOIN(",",TRUE,A1:A9)',
            '=_xlfn.TEXTJOIN(",",TRUE,A1:A9)',
        ),
        ("=CONCAT(A1:A9)", "=_xlfn.CONCAT(A1:A9)"),
        (
            '=FILTERXML(A1,"//item")',
            '=_xlfn.FILTERXML(A1,"//item")',
        ),
        ('=TEXTSPLIT(A1,",")', '=_xlfn.TEXTSPLIT(A1,",")'),
        ("=DROP(A1:C4,1)", "=_xlfn.DROP(A1:C4,1)"),
        ("=SORT(A1:C4)", "=_xlfn._xlws.SORT(A1:C4)"),
        ("=SORTBY(A1:C4,A1:A4)", "=_xlfn.SORTBY(A1:C4,A1:A4)"),
    ],
)
def test_bare_modern_functions_are_canonicalised_to_xlsx_storage_names(
    tmp_path: Path, formula: str, stored: str
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "modern.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Write a modern function.",
        operations=[SetFormula(op="set_formula", sheet="Output", cell="C2", formula=formula)],
    )

    apply_operations(plan, _task(source), source, destination)

    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["Output"]["C2"].value == stored
    finally:
        workbook.close()


def test_prefixed_modern_function_is_accepted(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "modern-prefixed.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Write a stored modern function.",
        operations=[
            SetFormula(
                op="set_formula",
                sheet="Output",
                cell="C2",
                formula="=_xlfn.XLOOKUP(A2,Data!A:A,Data!B:B)",
            )
        ],
    )

    apply_operations(plan, _task(source), source, destination)

    workbook = load_workbook(destination, data_only=False)
    assert workbook["Output"]["C2"].value == "=_xlfn.XLOOKUP(A2,Data!A:A,Data!B:B)"
    workbook.close()


def test_modern_function_canonicalisation_skips_strings_and_existing_prefixes(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "modern-strings.xlsx"
    _source_workbook(source)
    formula = (
        '="XLOOKUP(A1) ""FILTER(A1)"""&xlookup(A2,Data!A:A,Data!B:B)'
        "&_xlfn.UNIQUE(Data!A2:A4)&_xlfn._xlws.FILTER(Data!A2:A4,Data!A2:A4>0)"
    )
    plan = SolvePlan(
        route="operations",
        summary="Canonicalise only bare function calls.",
        operations=[SetFormula(op="set_formula", sheet="Output", cell="C2", formula=formula)],
    )

    apply_operations(plan, _task(source), source, destination)

    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook["Output"]["C2"].value == (
            '="XLOOKUP(A1) ""FILTER(A1)"""&_xlfn.XLOOKUP(A2,Data!A:A,Data!B:B)'
            "&_xlfn.UNIQUE(Data!A2:A4)&_xlfn._xlws.FILTER(Data!A2:A4,Data!A2:A4>0)"
        )
    finally:
        workbook.close()


def test_fill_formula_canonicalises_before_relative_translation(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "modern-fill.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Fill a modern formula.",
        operations=[
            FillFormula(
                op="fill_formula",
                sheet="Output",
                range="C2:C4",
                formula="=IFNA(Data!A2,0)",
            )
        ],
    )

    apply_operations(plan, _task(source), source, destination)

    workbook = load_workbook(destination, data_only=False)
    try:
        assert [workbook["Output"].cell(row, 3).value for row in range(2, 5)] == [
            "=_xlfn.IFNA(Data!A2,0)",
            "=_xlfn.IFNA(Data!A3,0)",
            "=_xlfn.IFNA(Data!A4,0)",
        ]
    finally:
        workbook.close()


def test_array_formula_operations_preserve_text_ref_and_evidence(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "arrays.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source)
    workbook["Output"]["C3"] = "stale spill value"
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Write one-cell and multi-cell array formulae.",
        operations=[
            SetArrayFormula(
                op="set_array_formula",
                sheet="Output",
                cell="$A$2",
                formula="=IFNA(Data!A2,0)",
            ),
            FillArrayFormula(
                op="fill_array_formula",
                sheet="Output",
                range="$B$2:$D$4",
                formula="=SORT(Data!A2:C4)",
            ),
        ],
    )

    evidence = apply_operations(plan, _task(source), source, destination)

    result = load_workbook(destination, data_only=False)
    try:
        output = result["Output"]
        single = output["A2"].value
        multi = output["B2"].value
        assert isinstance(single, ArrayFormula)
        assert single.text == "=_xlfn.IFNA(Data!A2,0)"
        assert single.ref == "A2"
        assert isinstance(multi, ArrayFormula)
        assert multi.text == "=_xlfn._xlws.SORT(Data!A2:C4)"
        assert multi.ref == "B2:D4"
        assert output.array_formulae == {"A2": "A2", "B2": "B2:D4"}
        assert output["C3"].value is None
        assert not isinstance(output["D4"].value, ArrayFormula)
    finally:
        result.close()

    assert evidence["cell_writes"] == 10
    assert evidence["array_formulas"] == [
        {
            "sheet": "Output",
            "anchor": "A2",
            "ref": "A2",
            "text": "=_xlfn.IFNA(Data!A2,0)",
        },
        {
            "sheet": "Output",
            "anchor": "B2",
            "ref": "B2:D4",
            "text": "=_xlfn._xlws.SORT(Data!A2:C4)",
        },
    ]


def test_multi_cell_array_ref_must_be_inside_declared_answer_ranges(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "array-outside.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Attempt an out-of-scope array formula.",
        operations=[
            FillArrayFormula(
                op="fill_array_formula",
                sheet="Output",
                range="D4:E5",
                formula="=Data!A2:B3",
            )
        ],
    )

    with pytest.raises(PlanApplicationError, match="falls outside declared answer ranges"):
        apply_operations(plan, _task(source), source, destination)

    assert not destination.exists()


def test_copy_range_translates_complete_array_formula_text_and_ref(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "copied-array.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = ArrayFormula(ref="C2:D3", text="=C2:D3*2")
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Copy the complete array formula.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="C2:D3",
                destination_sheet="Output",
                destination_cell="A2",
            )
        ],
    )

    apply_operations(plan, _task(source), source, destination)

    result = load_workbook(destination, data_only=False)
    try:
        copied = result["Output"]["A2"].value
        assert isinstance(copied, ArrayFormula)
        assert copied.ref == "A2:B3"
        assert copied.text == "=A2:B3*2"
    finally:
        result.close()


def test_copy_range_rejects_partial_array_formula_ref(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "partial-array.xlsx"
    _source_workbook(source)
    workbook = load_workbook(source, data_only=False)
    workbook["Data"]["C2"] = ArrayFormula(ref="C2:D3", text="=C2:D3*2")
    workbook.save(source)
    workbook.close()
    plan = SolvePlan(
        route="operations",
        summary="Attempt a partial array copy.",
        operations=[
            CopyRange(
                op="copy_range",
                source_sheet="Data",
                source_range="D2:D3",
                destination_sheet="Output",
                destination_cell="A2",
            )
        ],
    )

    with pytest.raises(PlanApplicationError, match="cannot copy only part of array formula"):
        apply_operations(plan, _task(source), source, destination)

    assert not destination.exists()


@pytest.mark.parametrize(
    "operation",
    [
        SetValue(op="set_value", sheet="Output", cell="A1", value="unrelated"),
        FillFormula(op="fill_formula", sheet="Output", range="B2:B5", formula="=Data!A2"),
        ClearRange(op="clear_range", sheet="Data", range="A2:A3"),
        CopyRange(
            op="copy_range",
            source_sheet="Data",
            source_range="A2:A4",
            destination_sheet="Output",
            destination_cell="D3",
        ),
    ],
)
def test_operations_cannot_write_outside_declared_answer_ranges(
    tmp_path: Path, operation: object
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    plan = SolvePlan(
        route="operations",
        summary="Attempt an out-of-scope write.",
        operations=[operation],
    )

    with pytest.raises(PlanApplicationError, match="declared answer range"):
        apply_operations(plan, _task(source), source, destination)

    assert not destination.exists()


def test_whole_column_answer_range_allows_result_to_grow_beyond_input_dimensions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    task = TaskSpec(
        id="whole-column",
        instruction_type="Cell-level manipulation",
        instruction="Write inside the declared answer column.",
        spreadsheet_path="whole-column",
        init_xlsx=source,
        answer_ranges=(QualifiedRange("Data", "B:B"),),
    )
    inside = SolvePlan(
        route="operations",
        summary="Write below the input's current used row boundary.",
        operations=[SetValue(op="set_value", sheet="Data", cell="B32", value=99)],
    )

    apply_operations(inside, task, source, destination)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Data"]["B32"].value == 99
    finally:
        result.close()

    wrong_column = SolvePlan(
        route="operations",
        summary="Write outside the declared answer column.",
        operations=[SetValue(op="set_value", sheet="Data", cell="C32", value=99)],
    )
    with pytest.raises(PlanApplicationError, match="falls outside"):
        apply_operations(wrong_column, task, source, tmp_path / "outside.xlsx")


def test_whole_row_answer_range_allows_result_to_grow_beyond_input_dimensions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _source_workbook(source)
    task = TaskSpec(
        id="whole-row",
        instruction_type="Cell-level manipulation",
        instruction="Write inside the declared answer row.",
        spreadsheet_path="whole-row",
        init_xlsx=source,
        answer_ranges=(QualifiedRange("Data", "2:2"),),
    )
    plan = SolvePlan(
        route="operations",
        summary="Write beyond the input's current used column boundary.",
        operations=[SetValue(op="set_value", sheet="Data", cell="Z2", value=99)],
    )

    apply_operations(plan, task, source, destination)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Data"]["Z2"].value == 99
    finally:
        result.close()
