import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableFormula

from exactsource.formula_safety import (
    FormulaSafetyError,
    snapshot_cell_hyperlinks,
    snapshot_formula_metadata,
    snapshot_formula_texts,
    validate_changed_cell_hyperlinks,
    validate_changed_formula_metadata_safety,
    validate_changed_formula_safety,
    validate_formula_integrity,
    validate_formula_safety,
)


@pytest.mark.parametrize(
    "formula",
    [
        "=SUM(A1:A3)",
        '="WEBSERVICE(""https://example.invalid"")"',
        '="text|not-dde"',
        "=SUM(Sales[Amount])",
        "=SUM(Sales[[#Data],[Amount]])",
        "='WEBSERVICE(archive)'!A1",
        "='Budget | Forecast'!A1",
        '=INDIRECT("A"&1)',
    ],
)
def test_validate_formula_safety_allows_internal_calculations(formula: str) -> None:
    validate_formula_safety(formula)


@pytest.mark.parametrize(
    ("formula", "sheetnames"),
    [
        ("=Data!A1", ("Data",)),
        ("=dAtA!$A$1", ("Data",)),
        ("='Cash Flow'!A1", ("Cash Flow",)),
        ("='O''Brien'!A1", ("O'Brien",)),
        ("=SUM('Jan 2024:Dec 2024'!B2)", ("Jan 2024", "Dec 2024")),
        ('="([{}]) ""quoted"""&Table1[[#Data],[Amount]]', ("Sheet",)),
        ("=SUM(Table1['#OfItems])", ("Sheet",)),
        ("=SUM(Table1[Column ']])", ("Sheet",)),
        ("=INDIRECT(\"'Missing'!A1\")", ("Sheet",)),
        ("=First:Missing!A1", ("First",)),
        ("=A1#", ("Sheet",)),
    ],
)
def test_formula_integrity_accepts_sound_static_constructs(
    formula: str,
    sheetnames: tuple[str, ...],
) -> None:
    validate_formula_integrity(formula, sheetnames=sheetnames)


@pytest.mark.parametrize(
    ("formula", "category"),
    [
        ("=SUM(A1:A2", "unbalanced formula delimiter"),
        ("=SUM(A1:A2))", "unbalanced formula delimiter"),
        ("={1,2", "unbalanced formula delimiter"),
        ("=Table1[[Amount]", "unbalanced formula delimiter"),
        ('="unterminated', "unterminated formula string"),
        ("='unterminated!A1", "unterminated quoted identifier"),
        ("=([A1)]", "unbalanced formula delimiter"),
    ],
)
def test_formula_integrity_rejects_confidently_malformed_structure(
    formula: str,
    category: str,
) -> None:
    with pytest.raises(FormulaSafetyError) as caught:
        validate_formula_integrity(formula, sheetnames=("Sheet",))

    assert caught.value.category == category


@pytest.mark.parametrize("formula", ["=Missing!A1", "='Missing Sheet'!$B$2"])
def test_formula_integrity_rejects_confident_missing_sheet_reference(formula: str) -> None:
    with pytest.raises(FormulaSafetyError) as caught:
        validate_formula_integrity(formula, sheetnames=("Data", "Output"))

    assert caught.value.category == "missing worksheet reference"
    assert "Missing" not in str(caught.value)


@pytest.mark.parametrize(
    "formula",
    [
        "=VLOOKUP(A2,B2:D20,3,FALSE)",
        "=HLOOKUP(A2,B2:D6,5,FALSE)",
        "=VLOOKUP(A2,$B$2:$D$20,1,FALSE)",
        "=VLOOKUP(A2,Data!B2:D20,3,FALSE)",
        "=HLOOKUP(A2,'Reference Data'!$B$2:$D$6,5,FALSE)",
        "=+VLOOKUP(A2,B2:D20,3,FALSE)",
        "=@HLOOKUP(A2,B2:D6,5,FALSE)",
        "=_xlfn.VLOOKUP(A2,B2:D20,3,FALSE)",
        "=_xlws.HLOOKUP(A2,B2:D6,5,FALSE)",
        "=+@_xlfn.VLOOKUP(A2,B2:D20,3,FALSE)",
    ],
)
def test_formula_integrity_accepts_literal_lookup_index_within_static_range(
    formula: str,
) -> None:
    validate_formula_integrity(
        formula,
        sheetnames=("Data", "Reference Data"),
    )


@pytest.mark.parametrize(
    "formula",
    [
        "=VLOOKUP(A2,B2:D20,4,FALSE)",
        "=  vlookup(A2,B2:D20,4,FALSE)",
        "= + VLOOKUP(A2,B2:D20,4,FALSE)",
        "=@VLOOKUP(A2,B2:D20,4,FALSE)",
        "=_xlfn.VLOOKUP(A2,B2:D20,4,FALSE)",
        "=_xlws.HLOOKUP(A2,B2:D6,6,FALSE)",
        "=+@_xlws.HLOOKUP(A2,B2:D6,6,FALSE)",
        "=VLOOKUP(A2,$B$2:$D$20,+4,FALSE)",
        "=VLOOKUP(A2,Data!B2:D20,0,FALSE)",
        "=VLOOKUP(A2,'Reference Data'!$B$2:$D$20,-1,FALSE)",
        "=HLOOKUP(A2,B2:D6,6,FALSE)",
        "=HLOOKUP(A2,'Reference Data'!$B$2:$D$6,000,FALSE)",
        "=VLOOKUP(A2,B2,2,FALSE)",
        f"=VLOOKUP(A2,B2:D20,{'9' * 5000},FALSE)",
    ],
)
def test_formula_integrity_rejects_literal_lookup_index_outside_static_range(
    formula: str,
) -> None:
    with pytest.raises(FormulaSafetyError) as caught:
        validate_formula_integrity(
            formula,
            sheetnames=("Data", "Reference Data"),
        )

    assert caught.value.category == "literal lookup index outside static range"
    assert str(caught.value) == (
        "formula rejected: literal lookup index outside static range is not allowed"
    )
    assert "Reference Data" not in str(caught.value)


@pytest.mark.parametrize(
    "formula",
    [
        "=IFERROR(VLOOKUP(A2,B2:D20,4,FALSE),0)",
        "=IFERROR(+@_xlfn.VLOOKUP(A2,B2:D20,4,FALSE),0)",
        "=VLOOKUP(A2,B2:D20,4,FALSE)+0",
        "=+VLOOKUP(A2,B2:D20,4,FALSE)+0",
        "=-VLOOKUP(A2,B2:D20,4,FALSE)",
        "=++VLOOKUP(A2,B2:D20,4,FALSE)",
        "=VLOOKUP(A2,LookupTable,4,FALSE)",
        "=VLOOKUP(A2,Table1[[#Data],[Key]:[Value]],4,FALSE)",
        "=VLOOKUP(A2,INDIRECT(F2),4,FALSE)",
        "=VLOOKUP(A2,CHOOSE({1,2},B:B,D:D),4,FALSE)",
        "=VLOOKUP(A2,(B2:D20),4,FALSE)",
        "=VLOOKUP(A2,B2:D20,COLUMNS(B2:E2),FALSE)",
        "=VLOOKUP(A2,B2:D20,1+3,FALSE)",
        "=VLOOKUP(A2,B2:D20,(4),FALSE)",
        "=VLOOKUP(A2,B2:D20,E2,FALSE)",
        "=VLOOKUP(A2,B2:D20,4.0,FALSE)",
        "=VLOOKUP(A2,B2:D20,4%,FALSE)",
        "=VLOOKUP(A2,Sheet1:Sheet3!B2:D20,4,FALSE)",
        "=VLOOKUP(A2,D20:B2,4,FALSE)",
        "=VLOOKUP(A2,B:D,4,FALSE)",
        "=XLOOKUP(A2,B2:B20,D2:D20)",
    ],
)
def test_formula_integrity_leaves_ambiguous_or_non_top_level_lookups_untouched(
    formula: str,
) -> None:
    validate_formula_integrity(formula)


def test_formula_integrity_fails_soft_for_oversized_static_range_row() -> None:
    oversized_row = "9" * 5000

    validate_formula_integrity(
        f"=VLOOKUP(A2,B2:D{oversized_row},4,FALSE)",
    )


@pytest.mark.parametrize(
    "formula",
    ["=NO_SUCH_FUNCTION(A1)", "=#REF!", "=A1+"],
)
def test_formula_integrity_does_not_claim_to_validate_excel_semantics(formula: str) -> None:
    validate_formula_integrity(formula, sheetnames=("Sheet",))


@pytest.mark.parametrize(
    ("formula", "category"),
    [
        ('=WEBSERVICE("https://private.invalid/path")', "external-capability function"),
        ('=_xlfn.WEBSERVICE("https://private.invalid")', "external-capability function"),
        ('=RTD("server",,A1)', "external-capability function"),
        ('=_xlfn.CALL("library","entry","J")', "external-capability function"),
        ('=REGISTER("library","entry","J")', "external-capability function"),
        ('=REGISTER.ID("library","entry","J")', "external-capability function"),
        ('=EXEC("notepad")', "external-capability function"),
        ('=RUN("MacroName")', "external-capability function"),
        ('=EVALUATE("WEBSERVICE(A1)")', "external-capability function"),
        ("=program|'topic'!A1", "DDE link"),
        ("='program'|'topic'!A1", "DDE link"),
        ("=[book.xlsx]Sheet1!A1", "external workbook link"),
        ("='[book.xlsx]Cash Flow'!$A$1", "external workbook link"),
        ("='C:\\funds\\[book.xlsx]Cash Flow'!$A$1", "external workbook link"),
        ('="file:///private/report.xlsx"', "file or UNC reference"),
        ('="\\\\server\\share\\report.xlsx"', "file or UNC reference"),
        ('=HYPERLINK("#\'Summary\'!A1","Jump")', "hyperlink or image function"),
        ('=HYPERLINK(A1,"Open")', "hyperlink or image function"),
        ("=_xlfn.IMAGE(A1)", "hyperlink or image function"),
        ('=INDIRECT("[book.xlsx]Sheet1!A1")', "external workbook link"),
        (
            '=INDIRECT("["&"book.xlsx"&"]"&"Sheet1"&"!A1")',
            "external workbook link",
        ),
        ('=INDIRECT("["&A1&"]Sheet1!A1")', "external workbook link"),
    ],
)
def test_validate_formula_safety_rejects_external_capabilities_without_echoing_input(
    formula: str,
    category: str,
) -> None:
    with pytest.raises(FormulaSafetyError) as caught:
        validate_formula_safety(formula)

    assert caught.value.category == category
    assert "private.invalid" not in str(caught.value)
    assert "server\\share" not in str(caught.value)


def test_changed_formula_validation_grandfathers_only_unchanged_formulae() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = '=WEBSERVICE("https://legacy.invalid")'
    sheet["A2"] = ArrayFormula(ref="A2", text="=SUM(B2:B3)")
    before = snapshot_formula_texts(workbook)

    assert validate_changed_formula_safety(workbook, before) == 0

    sheet["A2"].value.text = '=WEBSERVICE("https://new.invalid")'
    with pytest.raises(FormulaSafetyError, match="external-capability function"):
        validate_changed_formula_safety(workbook, before)


@pytest.mark.parametrize(
    ("formula", "category"),
    [
        ("=SUM(A1:A2", "unbalanced formula delimiter"),
        ("=Missing!A1", "missing worksheet reference"),
    ],
)
def test_changed_formula_validation_checks_structure_and_final_sheetnames(
    formula: str,
    category: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    before = snapshot_formula_texts(workbook)
    worksheet["A1"] = formula

    with pytest.raises(FormulaSafetyError) as caught:
        validate_changed_formula_safety(workbook, before)

    assert caught.value.category == category


def test_changed_array_formula_cannot_hide_function_in_discarded_prefix() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = ArrayFormula(ref="A1", text="=SUM(B1:B2)")
    before = snapshot_formula_texts(workbook)

    # openpyxl writes ArrayFormula.text[1:] into the worksheet XML. A malformed
    # leading character must not hide the formula which would actually be saved.
    sheet["A1"].value.text = 'xWEBSERVICE("https://array.invalid")'

    with pytest.raises(FormulaSafetyError, match="external-capability function"):
        validate_changed_formula_safety(workbook, before)


def test_expanding_preexisting_unsafe_array_formula_is_not_grandfathered() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = ArrayFormula(
        ref="A1",
        text='=WEBSERVICE("https://legacy-array.invalid")',
    )
    before = snapshot_formula_texts(workbook)

    sheet["A1"].value.ref = "A1:A10"

    with pytest.raises(FormulaSafetyError, match="external-capability function"):
        validate_changed_formula_safety(workbook, before)


def test_formula_snapshot_does_not_expand_sparse_worksheet_dimensions() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["XFD1048576"] = "=1+1"

    before_cells = len(sheet._cells)
    snapshot = snapshot_formula_texts(workbook)

    assert len(sheet._cells) == before_cells == 1
    assert snapshot == {("Sheet", "XFD1048576"): ("formula", "=1+1")}


def test_moved_preexisting_external_formula_is_not_grandfathered() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '=WEBSERVICE("https://legacy.invalid")'
    before = snapshot_formula_texts(workbook)

    sheet["B1"] = sheet["A1"].value
    sheet["A1"] = None

    with pytest.raises(FormulaSafetyError, match="external-capability function"):
        validate_changed_formula_safety(workbook, before)


def _workbook_with_formula_metadata() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Amount", "Calculated"])
    sheet.append([1, 2])
    sheet.append([3, 4])

    workbook.defined_names.add(DefinedName("SafeName", attr_text="SUM(Sheet1!A2:A3)"))
    table = Table(displayName="CashTable", ref="A1:B3")
    table._initialise_columns()
    table.tableColumns[0].name = "Amount"
    table.tableColumns[1].name = "Calculated"
    table.tableColumns[1].calculatedColumnFormula = TableFormula(attr_text="SUM(CashTable[Amount])")
    table.tableColumns[1].totalsRowFormula = TableFormula(attr_text="SUM(CashTable[Calculated])")
    sheet.add_table(table)

    validation = DataValidation(type="custom", formula1="SUM(A2:A3)>0")
    validation.add("A2:A3")
    sheet.add_data_validation(validation)
    sheet.conditional_formatting.add("A2:A3", FormulaRule(formula=["SUM(A2:A3)>0"]))
    return workbook


@pytest.mark.parametrize(
    "mutate",
    [
        lambda workbook: setattr(
            workbook.defined_names["SafeName"],
            "attr_text",
            'WEBSERVICE("https://name.invalid")',
        ),
        lambda workbook: setattr(
            workbook["Sheet1"].tables["CashTable"].tableColumns[1].calculatedColumnFormula,
            "attr_text",
            'WEBSERVICE("https://table.invalid")',
        ),
        lambda workbook: setattr(
            workbook["Sheet1"].tables["CashTable"].tableColumns[1].totalsRowFormula,
            "attr_text",
            'HYPERLINK("#A1","Jump")',
        ),
        lambda workbook: setattr(
            workbook["Sheet1"].data_validations.dataValidation[0],
            "formula1",
            'INDIRECT("[book.xlsx]Sheet1!A1")',
        ),
        lambda workbook: setattr(
            next(iter(workbook["Sheet1"].conditional_formatting._cf_rules.values()))[0],
            "formula",
            ['WEBSERVICE("https://format.invalid")'],
        ),
    ],
)
def test_changed_formula_metadata_is_checked_without_leaking_content(mutate) -> None:
    workbook = _workbook_with_formula_metadata()
    before = snapshot_formula_metadata(workbook)

    mutate(workbook)

    with pytest.raises(FormulaSafetyError) as caught:
        validate_changed_formula_metadata_safety(workbook, before)
    assert ".invalid" not in str(caught.value)


def test_unchanged_legacy_formula_metadata_is_grandfathered() -> None:
    workbook = _workbook_with_formula_metadata()
    workbook.defined_names["SafeName"].attr_text = 'WEBSERVICE("https://legacy.invalid")'
    table_formula = workbook["Sheet1"].tables["CashTable"].tableColumns[1].calculatedColumnFormula
    table_formula.attr_text = 'WEBSERVICE("https://legacy.invalid")'
    before = snapshot_formula_metadata(workbook)

    assert validate_changed_formula_metadata_safety(workbook, before) == {
        "conditional_formatting": 0,
        "data_validations": 0,
        "defined_names": 0,
        "table_formulae": 0,
    }


@pytest.mark.parametrize("flag", ["xlm", "function", "vbProcedure"])
def test_changed_defined_name_with_executable_flag_is_rejected(flag: str) -> None:
    workbook = _workbook_with_formula_metadata()
    before = snapshot_formula_metadata(workbook)

    setattr(workbook.defined_names["SafeName"], flag, True)

    with pytest.raises(FormulaSafetyError) as caught:
        validate_changed_formula_metadata_safety(workbook, before)
    assert caught.value.category == "executable defined name"


def test_new_external_cell_hyperlink_is_rejected_but_internal_location_is_allowed() -> None:
    workbook = Workbook()
    sheet = workbook.active
    before = snapshot_cell_hyperlinks(workbook)

    sheet["A1"].hyperlink = "https://private.invalid/path"
    with pytest.raises(FormulaSafetyError) as caught:
        validate_changed_cell_hyperlinks(workbook, before)
    assert caught.value.category == "external cell hyperlink"
    assert "private.invalid" not in str(caught.value)

    sheet["A1"].hyperlink = "#'Sheet'!B2"
    assert validate_changed_cell_hyperlinks(workbook, before) == 1


def test_external_hyperlink_location_cannot_hide_behind_internal_target() -> None:
    workbook = Workbook()
    sheet = workbook.active
    before = snapshot_cell_hyperlinks(workbook)
    sheet["A1"].hyperlink = "#Sheet!B2"
    sheet["A1"].hyperlink.location = "https://private.invalid/?payload=secret"

    with pytest.raises(FormulaSafetyError) as caught:
        validate_changed_cell_hyperlinks(workbook, before)

    assert caught.value.category == "external cell hyperlink"
    assert "private.invalid" not in str(caught.value)


def test_unchanged_legacy_external_hyperlink_is_grandfathered() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"].hyperlink = "https://legacy.invalid/path"
    before = snapshot_cell_hyperlinks(workbook)

    sheet["A1"].hyperlink.tooltip = "Updated display metadata"

    assert validate_changed_cell_hyperlinks(workbook, before) == 0


def test_moved_preexisting_external_cell_hyperlink_is_not_grandfathered() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"].hyperlink = "https://legacy.invalid/path"
    before = snapshot_cell_hyperlinks(workbook)

    sheet["B1"].hyperlink = sheet["A1"].hyperlink
    sheet["A1"].hyperlink = None

    with pytest.raises(FormulaSafetyError, match="external cell hyperlink"):
        validate_changed_cell_hyperlinks(workbook, before)
