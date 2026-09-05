from __future__ import annotations

import hashlib
import json
import time
from pathlib import Path

import openpyxl
import pytest

from exactsource.artifacts import ArtifactError, TraceRecorder, read_jsonl
from exactsource.cli import run_cli
from exactsource.config import (
    CELL_MAX_OUTPUT_TOKENS,
    CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS,
    SHEET_MAX_OUTPUT_TOKENS,
)
from exactsource.contracts import (
    ContextPack,
    ModelReply,
    QualifiedRange,
    SetValue,
    SolvePlan,
    SolveResult,
    TaskSpec,
)
from exactsource.model import ModelTruncationError
from exactsource.plans import PlanParseError
from exactsource.runner import DefaultTaskSolver, TaskSolveError, run_tasks


def _task(tmp_path: Path, task_id: str, value: str) -> TaskSpec:
    task_dir = tmp_path / "data" / "spreadsheet" / task_id
    task_dir.mkdir(parents=True)
    init = task_dir / f"1_{task_id}_init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Sheet1"
    workbook.active["A1"] = value
    workbook.save(init)
    workbook.close()
    return TaskSpec(
        id=task_id,
        instruction_type="Cell-Level Manipulation",
        instruction="Set A1 to solved",
        spreadsheet_path=f"spreadsheet/{task_id}",
        init_xlsx=init,
        answer_ranges=(QualifiedRange(sheet="Sheet1", cells="A1"),),
    )


def _sheet_task(tmp_path: Path, task_id: str, value: str) -> TaskSpec:
    cell_task = _task(tmp_path, task_id, value)
    return TaskSpec(
        id=cell_task.id,
        instruction_type="Sheet-Level Manipulation",
        instruction="Update the requested workbook output.",
        spreadsheet_path=cell_task.spreadsheet_path,
        init_xlsx=cell_task.init_xlsx,
        answer_ranges=cell_task.answer_ranges,
    )


def _cell(path: Path) -> object:
    workbook = openpyxl.load_workbook(path, data_only=False)
    value = workbook["Sheet1"]["A1"].value
    workbook.close()
    return value


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_runner_is_ordered_even_when_workers_finish_out_of_order(tmp_path: Path) -> None:
    tasks = [_task(tmp_path, "slow", "one"), _task(tmp_path, "fast", "two")]
    source_digests = {task.id: _digest(task.init_xlsx) for task in tasks}

    def solve(task: TaskSpec, working: Path, trace: TraceRecorder) -> SolveResult:
        if task.id == "slow":
            time.sleep(0.05)
        workbook = openpyxl.load_workbook(working)
        workbook["Sheet1"]["A1"] = f"solved-{task.id}"
        workbook.save(working)
        workbook.close()
        trace.record(model="fake", prompt="p", response="r", error=None)
        return SolveResult(status="ok", plan=None)

    lines: list[str] = []
    summary = run_tasks(tasks, tmp_path / "out", solve, concurrency=2, log=lines.append)

    predictions = read_jsonl(summary.predictions_path)
    assert [item["id"] for item in predictions] == ["slow", "fast"]
    assert summary.total == 2
    assert summary.succeeded == 2
    assert summary.failed == 0
    assert _cell(tmp_path / "out" / "outputs" / "slow.xlsx") == "solved-slow"
    assert _cell(tmp_path / "out" / "outputs" / "fast.xlsx") == "solved-fast"
    assert {_task.id: _digest(_task.init_xlsx) for _task in tasks} == source_digests
    assert lines[-1].startswith("complete total=2 ok=2 failed=0")
    metrics = json.loads(summary.run_metrics_path.read_text(encoding="utf-8"))
    assert metrics["tasks"] == {"total": 2, "succeeded": 2, "failed": 0}
    assert metrics["model"]["calls"] == 2
    assert metrics["model"]["attempts"] == 2
    assert metrics["usage"]["input_tokens"]["unknown_attempts"] == 2
    assert metrics["latency_ms"]["task"]["all"]["known_count"] == 2
    assert metrics["latency_ms"]["task"]["by_outcome"]["succeeded"]["known_count"] == 2


def test_runner_falls_back_to_init_and_continues_after_task_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    tasks = [_task(tmp_path, "broken", "original"), _task(tmp_path, "good", "other")]
    secret = "tinker-super-secret-test-value"
    monkeypatch.setenv("TINKER_API_KEY", secret)

    def solve(task: TaskSpec, working: Path, trace: TraceRecorder) -> None:
        trace.record(model="fake", prompt="p", response=None, error=None)
        if task.id == "broken":
            working.write_text("partial corrupt output", encoding="utf-8")
            raise RuntimeError(f"provider rejected {secret}")
        workbook = openpyxl.load_workbook(working)
        workbook["Sheet1"]["A1"] = "solved"
        workbook.save(working)
        workbook.close()

    summary = run_tasks(tasks, tmp_path / "out", solve, concurrency=2, log=lambda _: None)

    predictions = read_jsonl(summary.predictions_path)
    assert summary.succeeded == 1
    assert summary.failed == 1
    assert predictions[0]["id"] == "broken"
    assert predictions[0]["status"].startswith("error: RuntimeError")
    assert secret not in json.dumps(predictions)
    assert _cell(tmp_path / "out" / "outputs" / "broken.xlsx") == "original"
    assert _cell(tmp_path / "out" / "outputs" / "good.xlsx") == "solved"
    traces = read_jsonl(tmp_path / "out" / "traces" / "broken.jsonl")
    assert [record["step"] for record in traces] == [1]
    assert traces[-1]["task_status"] == "error"
    assert "RuntimeError" in traces[-1]["runtime_error"]
    assert secret not in json.dumps(traces)


def test_runner_treats_unreadable_solver_output_as_failure(tmp_path: Path) -> None:
    task = _task(tmp_path, "corrupt", "safe")

    def solve(_task: TaskSpec, working: Path, _trace: TraceRecorder) -> None:
        working.write_bytes(b"not a workbook")

    summary = run_tasks([task], tmp_path / "out", solve, log=lambda _: None)

    prediction = read_jsonl(summary.predictions_path)[0]
    assert prediction["status"].startswith("error: ArtifactError")
    assert _cell(tmp_path / "out" / "outputs" / "corrupt.xlsx") == "safe"


def test_runner_preserves_unrelated_existing_output(tmp_path: Path) -> None:
    task = _task(tmp_path, "new", "initial")
    out = tmp_path / "out"
    out.mkdir()
    sentinel = out / "analyst-notes.txt"
    sentinel.write_text("preserve", encoding="utf-8")

    def solve(_task: TaskSpec, _working: Path, trace: TraceRecorder) -> None:
        trace.record(model="fake", prompt="p", response="r", error=None)

    run_tasks([task], out, solve, log=lambda _: None)

    assert sentinel.read_text(encoding="utf-8") == "preserve"


def test_runner_rejects_duplicate_task_ids_before_writing(tmp_path: Path) -> None:
    task = _task(tmp_path, "duplicate", "initial")

    with pytest.raises(ArtifactError, match="duplicate task id"):
        run_tasks([task, task], tmp_path / "out", lambda *_: None, log=lambda _: None)
    assert not (tmp_path / "out").exists()


class _FakeClient:
    def __init__(self, replies: list[str | Exception]) -> None:
        self.replies = replies
        self.calls = 0
        self.requests: list[dict[str, object]] = []

    def complete(
        self,
        messages,
        *,
        max_output_tokens,
        reasoning_effort,
        on_attempt=None,
    ) -> ModelReply:
        reply = self.replies[self.calls]
        self.calls += 1
        self.requests.append(
            {
                "messages": [dict(message) for message in messages],
                "max_output_tokens": max_output_tokens,
                "reasoning_effort": reasoning_effort,
            }
        )
        truncated = isinstance(reply, ModelTruncationError)
        if truncated:
            response: str | None = "partial-response-must-not-be-replayed"
        elif isinstance(reply, Exception):
            response = None
        else:
            response = reply
        if on_attempt is not None:
            on_attempt(
                {
                    "attempt": 1,
                    "model": "ignored-by-runtime",
                    "temperature": 0,
                    "max_output_tokens": max_output_tokens,
                    "reasoning_effort": reasoning_effort,
                    "status": "error" if isinstance(reply, Exception) else "success",
                    "retryable": False,
                    "http_status": 200,
                    "response": response,
                    "error": str(reply) if isinstance(reply, Exception) else None,
                    "input_tokens": 10,
                    "output_tokens": reply.output_tokens if truncated else 5,
                    "stop_reason": "max_tokens" if truncated else "end_turn",
                    "latency_ms": 2,
                }
            )
        if isinstance(reply, Exception):
            raise reply
        return ModelReply(reply, 10, 5, 2)

    def close(self) -> None:
        pass


def _policy_test_plan() -> SolvePlan:
    return SolvePlan(
        route="operations",
        summary="Set the answer",
        operations=[SetValue(op="set_value", sheet="Sheet1", cell="A1", value="done")],
    )


def _policy_test_solver(
    client: _FakeClient,
    *,
    semantic_repairs: int = 1,
) -> DefaultTaskSolver:
    return DefaultTaskSolver(
        client,
        context_builder=lambda _: ContextPack(
            text="workbook", original_chars=8, truncated=False, sha256="abc"
        ),
        message_builder=lambda *_: [
            {"role": "system", "content": "return json"},
            {"role": "user", "content": "solve"},
        ],
        semantic_repairs=semantic_repairs,
    )


def test_cell_max_tokens_uses_one_fresh_no_think_recovery(tmp_path: Path) -> None:
    task = _task(tmp_path, "capped-cell", "initial")
    partial = ModelTruncationError(
        "provider stopped at max_tokens",
        output_tokens=CELL_MAX_OUTPUT_TOKENS,
    )
    client = _FakeClient([partial, _policy_test_plan().model_dump_json()])
    solver = _policy_test_solver(client)
    working = tmp_path / "working-capped-cell.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    result = solver(task, working, trace)

    assert result.status == "ok"
    assert client.calls == 2
    assert client.requests[0] == {
        "messages": [
            {"role": "system", "content": "return json"},
            {"role": "user", "content": "solve"},
        ],
        "max_output_tokens": CELL_MAX_OUTPUT_TOKENS,
        "reasoning_effort": True,
    }
    recovery_messages = client.requests[1]["messages"]
    assert recovery_messages == [
        {"role": "system", "content": "return json"},
        {"role": "user", "content": "solve\n\n/no_think"},
        {"role": "assistant", "content": "<think>\n\n</think>\n\n"},
    ]
    assert "partial-response-must-not-be-replayed" not in json.dumps(recovery_messages)
    assert client.requests[0]["max_output_tokens"] == CELL_MAX_OUTPUT_TOKENS == 16_000
    assert (
        client.requests[1]["max_output_tokens"]
        == CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS
        == 32_000
    )
    assert client.requests[1]["reasoning_effort"] is False
    assert _cell(working) == "done"
    assert len(trace.records) == 2
    assert trace.records[0]["semantic_attempt"] == 1
    assert trace.records[0]["semantic_repair"] is False
    assert trace.records[0]["generation_policy"] == "cell_reasoning"
    assert trace.records[0]["stop_reason"] == "max_tokens"
    assert trace.records[0]["output_tokens"] == CELL_MAX_OUTPUT_TOKENS
    assert trace.records[0]["plan_status"] == "not_reached"
    assert trace.records[1]["semantic_attempt"] == 2
    assert trace.records[1]["semantic_repair"] is True
    assert trace.records[1]["generation_policy"] == "cell_max_tokens_no_think_recovery"
    assert trace.records[1]["recovery_reason"] == "max_tokens"
    assert trace.records[1]["max_output_tokens"] == CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS
    assert trace.records[1]["reasoning_effort"] is False
    assert trace.records[1]["plan_status"] == "accepted"


def test_cell_max_tokens_without_repair_budget_fails_after_one_call(tmp_path: Path) -> None:
    task = _task(tmp_path, "capped-cell-no-budget", "initial")
    client = _FakeClient(
        [
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client, semantic_repairs=0)
    working = tmp_path / "working-capped-cell-no-budget.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())

    with pytest.raises(ModelTruncationError):
        solver(task, working, TraceRecorder(task.id))

    assert client.calls == 1


def test_cell_truncation_recovery_is_counted_as_one_semantic_repair(
    tmp_path: Path,
) -> None:
    task = _task(tmp_path, "capped-cell-metrics", "initial")
    client = _FakeClient(
        [
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            _policy_test_plan().model_dump_json(),
        ]
    )

    summary = run_tasks(
        [task],
        tmp_path / "capped-cell-metrics-out",
        _policy_test_solver(client),
        log=lambda _: None,
    )

    metrics = json.loads(summary.run_metrics_path.read_text(encoding="utf-8"))
    assert metrics["tasks"] == {"total": 1, "succeeded": 1, "failed": 0}
    assert metrics["model"]["calls"] == 2
    assert metrics["model"]["attempts"] == 2
    assert metrics["usage"]["output_tokens"]["known_sum"] == (CELL_MAX_OUTPUT_TOKENS + 5)
    assert metrics["reliability"]["semantic_repairs"] == 1
    assert metrics["reliability"]["provider_status_counts"] == {
        "error": 1,
        "success": 1,
    }


def test_sheet_max_tokens_does_not_trigger_capped_recovery(tmp_path: Path) -> None:
    task = _sheet_task(tmp_path, "capped-sheet", "initial")
    client = _FakeClient(
        [
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=SHEET_MAX_OUTPUT_TOKENS,
            ),
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client)
    working = tmp_path / "working-capped-sheet.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    with pytest.raises(ModelTruncationError):
        solver(task, working, trace)

    assert client.calls == 1
    assert client.requests[0]["max_output_tokens"] == SHEET_MAX_OUTPUT_TOKENS
    assert client.requests[0]["reasoning_effort"] is True
    assert trace.records[0]["generation_policy"] == "sheet_reasoning"
    assert trace.records[0]["plan_status"] == "not_reached"


def test_rejected_cell_truncation_recovery_never_makes_a_third_call(tmp_path: Path) -> None:
    task = _task(tmp_path, "capped-cell-rejected", "initial")
    client = _FakeClient(
        [
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            "not a solve plan",
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client, semantic_repairs=2)
    working = tmp_path / "working-capped-cell-rejected.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())

    with pytest.raises(PlanParseError):
        solver(task, working, TraceRecorder(task.id))

    assert client.calls == 2
    assert client.requests[1]["max_output_tokens"] == CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS
    assert client.requests[1]["reasoning_effort"] is False


def test_capped_cell_truncation_recovery_never_makes_a_third_call(tmp_path: Path) -> None:
    task = _task(tmp_path, "twice-capped-cell", "initial")
    client = _FakeClient(
        [
            ModelTruncationError(
                "initial response stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            ModelTruncationError(
                "recovery response stopped at max_tokens",
                output_tokens=CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS,
            ),
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client, semantic_repairs=2)
    working = tmp_path / "working-twice-capped-cell.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    with pytest.raises(ModelTruncationError):
        solver(task, working, trace)

    assert client.calls == 2
    assert [request["max_output_tokens"] for request in client.requests] == [
        CELL_MAX_OUTPUT_TOKENS,
        CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS,
    ]
    assert [request["reasoning_effort"] for request in client.requests] == [True, False]
    assert [record["generation_policy"] for record in trace.records] == [
        "cell_reasoning",
        "cell_max_tokens_no_think_recovery",
    ]
    assert [record["stop_reason"] for record in trace.records] == [
        "max_tokens",
        "max_tokens",
    ]
    assert _cell(working) == "initial"


def test_unsafe_cell_truncation_recovery_fails_without_changing_workbook(
    tmp_path: Path,
) -> None:
    task = _task(tmp_path, "capped-cell-unsafe", "initial")
    unsafe_plan = SolvePlan(
        route="python",
        summary="Use a disallowed route",
        python_code="def transform(wb):\n    wb.active['A1'] = 'unsafe'",
    )
    client = _FakeClient(
        [
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            unsafe_plan.model_dump_json(),
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client, semantic_repairs=2)
    working = tmp_path / "working-capped-cell-unsafe.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())

    with pytest.raises(TaskSolveError, match="not allowed for cell-level"):
        solver(task, working, TraceRecorder(task.id))

    assert client.calls == 2
    assert client.requests[1]["max_output_tokens"] == CELL_TRUNCATION_RECOVERY_MAX_OUTPUT_TOKENS
    assert _cell(working) == "initial"


def test_truncated_ordinary_cell_repair_never_makes_a_third_call(tmp_path: Path) -> None:
    task = _task(tmp_path, "capped-cell-repair", "initial")
    client = _FakeClient(
        [
            "not a solve plan",
            ModelTruncationError(
                "provider stopped at max_tokens",
                output_tokens=CELL_MAX_OUTPUT_TOKENS,
            ),
            _policy_test_plan().model_dump_json(),
        ]
    )
    solver = _policy_test_solver(client, semantic_repairs=2)
    working = tmp_path / "working-capped-cell-repair.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())

    with pytest.raises(ModelTruncationError):
        solver(task, working, TraceRecorder(task.id))

    assert client.calls == 2
    assert client.requests[1]["max_output_tokens"] == CELL_MAX_OUTPUT_TOKENS
    assert client.requests[1]["reasoning_effort"] is True
    repair_messages = client.requests[1]["messages"]
    assert isinstance(repair_messages, list)
    assert repair_messages[-2] == {"role": "assistant", "content": "not a solve plan"}


def test_sheet_completed_rejection_keeps_one_reasoning_repair_at_32k(tmp_path: Path) -> None:
    task = _sheet_task(tmp_path, "sheet-repair-policy", "initial")
    client = _FakeClient(["not a solve plan", _policy_test_plan().model_dump_json()])
    solver = _policy_test_solver(client)
    working = tmp_path / "working-sheet-repair-policy.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    result = solver(task, working, trace)

    assert result.status == "ok"
    assert client.calls == 2
    assert (
        [request["max_output_tokens"] for request in client.requests]
        == [
            SHEET_MAX_OUTPUT_TOKENS,
            SHEET_MAX_OUTPUT_TOKENS,
        ]
        == [32_000, 32_000]
    )
    assert [request["reasoning_effort"] for request in client.requests] == [True, True]
    assert [record["generation_policy"] for record in trace.records] == [
        "sheet_reasoning",
        "sheet_reasoning",
    ]
    assert [record["semantic_repair"] for record in trace.records] == [False, True]


def test_default_solver_repairs_one_rejected_plan_and_traces_each_call(tmp_path: Path) -> None:
    task = _task(tmp_path, "repair", "initial")
    valid_plan = SolvePlan(
        route="operations",
        summary="Set the answer",
        operations=[SetValue(op="set_value", sheet="Sheet1", cell="A1", value="done")],
    )
    client = _FakeClient(["bad", valid_plan.model_dump_json()])

    def context_builder(_task: TaskSpec) -> ContextPack:
        return ContextPack(text="workbook", original_chars=8, truncated=False, sha256="abc")

    def parser(text: str) -> SolvePlan:
        if text == "bad":
            raise ValueError("invalid first plan")
        return valid_plan

    def apply(
        _plan: SolvePlan,
        _task: TaskSpec,
        source: Path,
        destination: Path,
    ) -> dict[str, object]:
        workbook = openpyxl.load_workbook(source)
        workbook["Sheet1"]["A1"] = "done"
        workbook.save(destination)
        workbook.close()
        return {"route": "operations", "cell_writes": 1}

    solver = DefaultTaskSolver(
        client,
        context_builder=context_builder,
        message_builder=lambda *_: [{"role": "user", "content": "solve"}],
        plan_parser=parser,
        operation_applier=apply,
        semantic_repairs=1,
    )
    working = tmp_path / "working.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    result = solver(task, working, trace)

    assert result.status == "ok"
    assert client.calls == 2
    assert (
        [request["max_output_tokens"] for request in client.requests]
        == [
            CELL_MAX_OUTPUT_TOKENS,
            CELL_MAX_OUTPUT_TOKENS,
        ]
        == [16_000, 16_000]
    )
    assert [request["reasoning_effort"] for request in client.requests] == [True, True]
    repaired_messages = client.requests[1]["messages"]
    assert isinstance(repaired_messages, list)
    assert repaired_messages[-2] == {"role": "assistant", "content": "bad"}
    assert repaired_messages[-1]["role"] == "user"
    assert "invalid first plan" in repaired_messages[-1]["content"]
    assert _cell(working) == "done"
    provider_records = [row for row in trace.records if row.get("event") == "provider_attempt"]
    assert provider_records == []
    assert len(trace.records) == 2
    assert [row["semantic_attempt"] for row in trace.records] == [1, 2]
    assert all(row["model"] == "tinker:Qwen/Qwen3.8-27B" for row in trace.records)
    assert trace.records[0]["status"] == "plan_rejected"
    assert trace.records[0]["provider_status"] == "success"
    assert trace.records[0]["plan_status"] == "parse_rejected"
    assert trace.records[0]["semantic_repair"] is False
    assert trace.records[0]["generation_policy"] == "cell_reasoning"
    assert trace.records[0]["max_output_tokens"] == CELL_MAX_OUTPUT_TOKENS
    assert trace.records[0]["reasoning_effort"] is True
    assert trace.records[0]["logical_call_terminal"] is True
    assert trace.records[0]["logical_call_latency_ms"] == 2
    assert isinstance(trace.records[0]["context_build_latency_ms"], int)
    assert isinstance(trace.records[0]["message_build_latency_ms"], int)
    assert isinstance(trace.records[0]["plan_parse_latency_ms"], int)
    assert trace.records[0]["tool"] == "parse_and_apply_plan"
    assert trace.records[1]["provider_status"] == "success"
    assert trace.records[1]["plan_status"] == "accepted"
    assert trace.records[1]["semantic_repair"] is True
    assert trace.records[1]["generation_policy"] == "cell_reasoning"
    assert trace.records[1]["max_output_tokens"] == CELL_MAX_OUTPUT_TOKENS
    assert trace.records[1]["reasoning_effort"] is True
    assert isinstance(trace.records[1]["plan_apply_latency_ms"], int)
    assert trace.records[1]["tool"] == "apply_operations"


def test_cli_dependency_injection_needs_no_provider_key(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    task = _task(tmp_path, "cli", "initial")
    data_dir = tmp_path / "judge-data"
    data_dir.mkdir()
    out_dir = tmp_path / "judge-out"
    monkeypatch.delenv("TINKER_API_KEY", raising=False)

    def solve(_task: TaskSpec, working: Path, trace: TraceRecorder) -> None:
        workbook = openpyxl.load_workbook(working)
        workbook["Sheet1"]["A1"] = "from-cli"
        workbook.save(working)
        workbook.close()
        trace.record(model="fake", prompt="p", response="r", error=None)

    exit_code = run_cli(
        ["--data-dir", str(data_dir), "--out-dir", str(out_dir)],
        task_loader=lambda _: [task],
        solver=solve,
    )

    assert exit_code == 0
    assert _cell(out_dir / "outputs" / "cli.xlsx") == "from-cli"
    assert "ExactSource finished: 1/1 tasks succeeded" in (out_dir / "run.log").read_text(
        encoding="utf-8"
    )


def test_cli_development_ids_preserve_dataset_order_and_reject_unknown(
    tmp_path: Path,
) -> None:
    tasks = [
        _task(tmp_path, "first", "one"),
        _task(tmp_path, "second", "two"),
        _task(tmp_path, "third", "three"),
    ]
    data_dir = tmp_path / "judge-data"
    data_dir.mkdir()

    def solve(task: TaskSpec, _working: Path, trace: TraceRecorder) -> None:
        trace.record(model="fake", prompt="p", response="r", error=None)

    exit_code = run_cli(
        [
            "--data-dir",
            str(data_dir),
            "--out-dir",
            str(tmp_path / "selected-out"),
            "--ids",
            "third,first",
        ],
        task_loader=lambda _: tasks,
        solver=solve,
    )

    assert exit_code == 0
    predictions = [
        json.loads(line)
        for line in (tmp_path / "selected-out" / "predictions.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]
    assert [prediction["id"] for prediction in predictions] == ["first", "third"]

    bad_exit = run_cli(
        [
            "--data-dir",
            str(data_dir),
            "--out-dir",
            str(tmp_path / "unknown-out"),
            "--ids",
            "missing",
        ],
        task_loader=lambda _: tasks,
        solver=solve,
    )
    assert bad_exit == 1
    assert "unknown task ids" in (tmp_path / "unknown-out" / "run.log").read_text()


def test_cli_rejects_overlapping_input_and_output(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    data_dir.mkdir()

    exit_code = run_cli(
        ["--data-dir", str(data_dir), "--out-dir", str(data_dir / "out")],
        task_loader=lambda _: [],
        solver=lambda *_: None,
    )

    assert exit_code == 2
    assert not (data_dir / "out").exists()


def test_end_to_end_runtime_with_real_context_and_fake_model(tmp_path: Path) -> None:
    task_dir = tmp_path / "data" / "spreadsheet" / "e2e"
    task_dir.mkdir(parents=True)
    init = task_dir / "1_e2e_init.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Calculations"
    workbook.active["A1"] = 3
    workbook.active["B1"] = 4
    workbook.active["C1"] = None
    workbook.save(init)
    workbook.close()
    task = TaskSpec(
        id="e2e",
        instruction_type="Cell-Level Manipulation",
        instruction="In C1, add A1 and B1 using a formula.",
        spreadsheet_path="spreadsheet/e2e",
        init_xlsx=init,
        answer_ranges=(QualifiedRange(sheet="Calculations", cells="C1"),),
        data_position="Calculations!A1:B1",
    )
    plan = SolvePlan(
        route="operations",
        summary="Add the two source cells in C1",
        operations=[
            {
                "op": "set_formula",
                "sheet": "Calculations",
                "cell": "C1",
                "formula": "=SUM(A1:B1)",
            }
        ],
    )
    solver = DefaultTaskSolver(_FakeClient([plan.model_dump_json()]), semantic_repairs=1)

    summary = run_tasks([task], tmp_path / "out", solver, log=lambda _: None)

    assert summary.succeeded == 1
    result = openpyxl.load_workbook(tmp_path / "out" / "outputs" / "e2e.xlsx", data_only=False)
    assert result["Calculations"]["C1"].value == "=SUM(A1:B1)"
    result.close()
    traces = read_jsonl(tmp_path / "out" / "traces" / "e2e.jsonl")
    assert len(traces) == 1
    assert traces[0]["event"] == "model_call"
    assert traces[0]["tool"] == "apply_operations"
    assert traces[0]["context"]["sha256"]
    assert traces[0]["provider_status"] == "success"
    assert traces[0]["plan_status"] == "accepted"
    assert traces[0]["task_status"] == "ok"
    assert isinstance(traces[0]["task_latency_ms"], int)
    assert isinstance(traces[0]["context_build_latency_ms"], int)
    assert isinstance(traces[0]["message_build_latency_ms"], int)
    assert isinstance(traces[0]["plan_parse_latency_ms"], int)
    assert isinstance(traces[0]["plan_apply_latency_ms"], int)


def test_default_solver_rejects_python_route_for_cell_level_task(tmp_path: Path) -> None:
    task = _task(tmp_path, "no-python", "initial")
    plan = SolvePlan(
        route="python",
        summary="Attempt broad code",
        python_code="def transform(wb):\n    wb.active['A1'] = 'wrong'",
    )
    transform_called = False

    def transform(_code: str, _source: Path, _destination: Path) -> dict[str, object]:
        nonlocal transform_called
        transform_called = True
        return {}

    solver = DefaultTaskSolver(
        _FakeClient([plan.model_dump_json()]),
        context_builder=lambda _: ContextPack(
            text="workbook", original_chars=8, truncated=False, sha256="abc"
        ),
        message_builder=lambda *_: [{"role": "user", "content": "solve"}],
        transform_runner=transform,
        semantic_repairs=0,
    )
    working = tmp_path / "working-no-python.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())
    trace = TraceRecorder(task.id)

    with pytest.raises(TaskSolveError, match="not allowed for cell-level"):
        solver(task, working, trace)
    assert transform_called is False
    assert _cell(working) == "initial"
    assert trace.records[0]["provider_status"] == "success"
    assert trace.records[0]["plan_status"] == "apply_rejected"
    assert isinstance(trace.records[0]["plan_parse_latency_ms"], int)
    assert isinstance(trace.records[0]["plan_apply_latency_ms"], int)


def test_default_solver_rejects_missing_declared_result_sheet(tmp_path: Path) -> None:
    base = _task(tmp_path, "missing-result", "initial")
    task = TaskSpec(
        id=base.id,
        instruction_type="Sheet-Level Manipulation",
        instruction="Create a result worksheet.",
        spreadsheet_path=base.spreadsheet_path,
        init_xlsx=base.init_xlsx,
        answer_ranges=(QualifiedRange(sheet="Result", cells="A1"),),
    )
    plan = SolvePlan(
        route="python",
        summary="Incorrectly leave the result sheet absent",
        python_code="def transform(wb):\n    pass",
    )

    def transform(_code: str, source: Path, destination: Path) -> dict[str, object]:
        destination.write_bytes(source.read_bytes())
        return {"route": "python"}

    solver = DefaultTaskSolver(
        _FakeClient([plan.model_dump_json()]),
        context_builder=lambda _: ContextPack(
            text="workbook", original_chars=8, truncated=False, sha256="abc"
        ),
        message_builder=lambda *_: [{"role": "user", "content": "solve"}],
        transform_runner=transform,
        semantic_repairs=0,
    )
    working = tmp_path / "working-missing-result.xlsx"
    working.write_bytes(task.init_xlsx.read_bytes())

    with pytest.raises(TaskSolveError, match="missing required answer worksheet"):
        solver(task, working, TraceRecorder(task.id))
