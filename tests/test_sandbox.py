from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableFormula

from exactsource.sandbox import (
    SandboxExecutionError,
    SandboxValidationError,
    run_transform,
    screen_transform,
)


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row, value in enumerate((2, 4, 6), start=2):
        sheet.cell(row, 1, value)
    workbook.save(path)
    workbook.close()


def _workbook_with_cached_formula(path: Path) -> None:
    """Write a controlled XLSX whose formula has a cached scalar result."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cash Flow 2026"
    sheet["A1"] = 7
    sheet["B1"] = 5
    sheet["C1"] = "=A1+B1"
    workbook.save(path)
    workbook.close()

    with ZipFile(path, "r") as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    worksheet_path = "xl/worksheets/sheet1.xml"
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ElementTree.register_namespace("", namespace)
    worksheet_xml = ElementTree.fromstring(members[worksheet_path])
    formula_cell = worksheet_xml.find(f".//{{{namespace}}}c[@r='C1']")
    assert formula_cell is not None
    formula = formula_cell.find(f"{{{namespace}}}f")
    assert formula is not None and formula.text == "A1+B1"
    cached_value = formula_cell.find(f"{{{namespace}}}v")
    assert cached_value is not None
    cached_value.text = "12"
    members[worksheet_path] = ElementTree.tostring(worksheet_xml, encoding="utf-8")
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
        for filename, contents in members.items():
            archive.writestr(filename, contents)


def _workbook_with_formula_metadata(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Amount", "Calculated"])
    sheet.append([1, 2])
    sheet.append([3, 4])

    workbook.defined_names.add(DefinedName("SafeName", attr_text="SUM(Sheet1!A2:A3)"))
    table = Table(displayName="CashTable", ref="A1:B3")
    table._initialise_columns()
    table.tableColumns[0].name = "Amount"
    table.tableColumns[1].name = "Calculated"
    table.tableColumns[1].calculatedColumnFormula = TableFormula(attr_text="SUM(CashTable[Amount])")
    table.tableColumns[1].totalsRowFormula = TableFormula(attr_text="SUM(CashTable[Calculated])")
    sheet.add_table(table)
    workbook.save(path)
    workbook.close()


def test_run_transform_executes_in_child_and_promotes_valid_workbook(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    ws = wb["Sheet1"]
    for row in range(2, 5):
        ws.cell(row=row, column=2).value = ws.cell(row=row, column=1).value * 3
"""

    evidence = run_transform(code, source, destination, timeout=10)

    assert destination.stat().st_mode & 0o777 == 0o644
    result = load_workbook(destination)
    try:
        assert [result["Sheet1"].cell(row, 2).value for row in range(2, 5)] == [6, 12, 18]
    finally:
        result.close()
    assert evidence["route"] == "python"
    assert evidence["ast_nodes"] > 0
    assert evidence["output_sha256"]
    assert evidence["sheets_before"] == ["Sheet1"]
    assert evidence["formulae_checked"] == 0
    assert evidence["formula_metadata_checked"] == {
        "conditional_formatting": 0,
        "data_validations": 0,
        "defined_names": 0,
        "table_formulae": 0,
    }
    assert evidence["hyperlinks_checked"] == 0


def test_python_transform_rejects_new_external_formula_without_leaking_uri(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    wb["Sheet1"]["B2"] = '=WEBSERVICE("https://secret.invalid/path")'
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external-capability function" in str(caught.value)
    assert "secret.invalid" not in str(caught.value)
    assert not destination.exists()


def test_python_transform_preserves_unchanged_preexisting_external_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = '=WEBSERVICE("https://legacy.invalid/path")'
    workbook.save(source)
    workbook.close()
    code = """\
def transform(wb):
    wb["Sheet1"]["B1"] = "safe change"
"""

    evidence = run_transform(code, source, destination, timeout=10)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Sheet1"]["A1"].value == '=WEBSERVICE("https://legacy.invalid/path")'
        assert result["Sheet1"]["B1"].value == "safe change"
    finally:
        result.close()
    assert evidence["formulae_checked"] == 0


def test_python_transform_checks_changed_array_formula_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = ArrayFormula(ref="A1", text="=SUM(B1:B2)")
    workbook.save(source)
    workbook.close()
    code = """\
def transform(wb):
    wb["Sheet1"]["A1"].value.text = '=WEBSERVICE("https://array.invalid")'
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external-capability function" in str(caught.value)
    assert "array.invalid" not in str(caught.value)
    assert not destination.exists()


def test_python_transform_rejects_changed_defined_name_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook_with_formula_metadata(source)
    code = """\
def transform(wb):
    wb.defined_names["SafeName"].attr_text = '=WEBSERVICE("https://name.invalid")'
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external-capability function" in str(caught.value)
    assert "name.invalid" not in str(caught.value)
    assert not destination.exists()


def test_python_transform_rejects_xlm_exec_defined_name_without_leaking_payload(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook_with_formula_metadata(source)
    code = """\
def transform(wb):
    defined_name = wb.defined_names["SafeName"]
    defined_name.attr_text = '=EXEC("notepad-secret-payload")'
    defined_name.xlm = True
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "executable defined name" in str(caught.value)
    assert "notepad-secret-payload" not in str(caught.value)
    assert not destination.exists()


@pytest.mark.parametrize("attribute", ["calculatedColumnFormula", "totalsRowFormula"])
def test_python_transform_rejects_changed_table_column_formula(
    tmp_path: Path,
    attribute: str,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook_with_formula_metadata(source)
    code = f"""\
def transform(wb):
    formula = wb["Sheet1"].tables["CashTable"].tableColumns[1].{attribute}
    formula.attr_text = '=WEBSERVICE("https://table.invalid")'
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external-capability function" in str(caught.value)
    assert "table.invalid" not in str(caught.value)
    assert not destination.exists()


def test_python_transform_rejects_new_external_cell_hyperlink(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    wb["Sheet1"]["A1"].hyperlink = "https://link.invalid/?payload=workbook-data"
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external cell hyperlink" in str(caught.value)
    assert "link.invalid" not in str(caught.value)
    assert not destination.exists()


def test_python_transform_rejects_external_location_on_internal_hyperlink(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    cell = wb["Sheet1"]["A1"]
    cell.hyperlink = "#Sheet1!B2"
    cell.hyperlink.location = "https://location.invalid/?payload=secret"
"""

    with pytest.raises(SandboxExecutionError) as caught:
        run_transform(code, source, destination, timeout=10)

    assert "external cell hyperlink" in str(caught.value)
    assert "location.invalid" not in str(caught.value)
    assert not destination.exists()


@pytest.mark.parametrize(
    "code, expected",
    [
        ("import os\ndef transform(wb):\n    pass", "Import"),
        ("def transform(wb):\n    wb.save('/tmp/stolen.xlsx')", "attribute 'save'"),
        ("def transform(wb):\n    open('/tmp/stolen', 'w')", "name 'open'"),
        ("def transform(wb):\n    return wb.__class__", "attribute '__class__'"),
        ("def transform(wb):\n    while True:\n        pass", "While"),
        ("def wrong(wb):\n    pass", "named transform"),
        ("def transform(wb, path):\n    pass", "exact signature"),
        (
            "def helper():\n    pass\ndef transform(wb):\n    pass",
            "exactly one function",
        ),
    ],
)
def test_screen_transform_rejects_escape_surfaces(code: str, expected: str) -> None:
    with pytest.raises(SandboxValidationError, match=expected):
        screen_transform(code)


@pytest.mark.parametrize(
    "code, expected",
    [
        (
            "def transform(wb):\n"
            "    wb.active['B2'] = str(statistics.sys.modules['os'].listdir('/etc'))",
            "attribute '(modules|sys)'",
        ),
        (
            "def transform(wb):\n    wb.active['B2'] = '{0.__class__}'.format(wb)",
            "attribute 'format'",
        ),
        (
            "def transform(wb):\n"
            "    wb.active['B2'] = '{value.__class__}'.format_map({'value': wb})",
            "attribute 'format_map'",
        ),
    ],
)
def test_screen_transform_rejects_indirect_object_graph_escapes(
    code: str,
    expected: str,
) -> None:
    with pytest.raises(SandboxValidationError, match=expected):
        screen_transform(code)


def test_safe_helper_facades_retain_only_documented_operations(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    ws = wb["Sheet1"]
    ws["B2"] = math.sqrt(81)
    ws["B3"] = statistics.mean([2, 4, 6])
    ws["B4"] = re.sub("[0-9]", "x", "A1")
"""

    run_transform(code, source, destination, timeout=10)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Sheet1"]["B2"].value == 9
        assert result["Sheet1"]["B3"].value == 4
        assert result["Sheet1"]["B4"].value == "Ax"
    finally:
        result.close()


def test_cached_values_reads_formula_result_without_replacing_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook_with_cached_formula(source)
    code = """\
def transform(wb):
    ws = wb["Cash Flow 2026"]
    ws["D1"] = cached_values.get("Cash Flow 2026", {}).get("C1")
"""

    run_transform(code, source, destination, timeout=10)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Cash Flow 2026"]["C1"].value == "=A1+B1"
        assert result["Cash Flow 2026"]["D1"].value == 12
    finally:
        result.close()


def test_cached_values_outer_and_sheet_mappings_are_read_only(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook_with_cached_formula(source)
    code = """\
def transform(wb):
    ws = wb["Cash Flow 2026"]
    blocked = []
    try:
        cached_values["Cash Flow 2026"] = {}
    except TypeError:
        blocked.append("outer")
    try:
        cached_values["Cash Flow 2026"]["C1"] = 99
    except TypeError:
        blocked.append("sheet")
    ws["D1"] = ",".join(blocked)
"""

    run_transform(code, source, destination, timeout=10)

    result = load_workbook(destination, data_only=False)
    try:
        assert result["Cash Flow 2026"]["C1"].value == "=A1+B1"
        assert result["Cash Flow 2026"]["D1"].value == "outer,sheet"
    finally:
        result.close()


def test_transform_must_return_none(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)

    with pytest.raises(SandboxExecutionError, match="must mutate in place and return None"):
        run_transform("def transform(wb):\n    return wb", source, destination, timeout=10)

    assert not destination.exists()


def test_transform_has_a_wall_clock_timeout(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    code = """\
def transform(wb):
    total = 0
    for number in range(1000000000):
        total = total + number
"""

    with pytest.raises(SandboxExecutionError, match="wall-clock limit"):
        run_transform(code, source, destination, timeout=0.25)

    assert not destination.exists()


def test_failed_transform_does_not_replace_existing_destination(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    _workbook(source)
    destination.write_bytes(b"existing result")

    with pytest.raises(SandboxExecutionError, match="synthetic failure"):
        run_transform(
            "def transform(wb):\n    raise ValueError('synthetic failure')",
            source,
            destination,
            timeout=10,
        )

    assert destination.read_bytes() == b"existing result"
